from openpyxl import load_workbook

# Load the template to see the income statement structure
try:
    wb = load_workbook('Credit Rating Template NEW.xlsx')
    print('Worksheets:', wb.sheetnames)
    
    # Look for income statement sheet
    for sheet_name in wb.sheetnames:
        if 'income' in sheet_name.lower() or 'is' in sheet_name.lower():
            ws = wb[sheet_name]
            print(f'\nIncome Statement sheet: {sheet_name}')
            print('First 30 rows:')
            for row in range(1, 31):
                values = []
                for col in range(1, 5):  # First 4 columns
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        values.append(str(cell_value))
                if values:
                    print(f'Row {row}: {" | ".join(values)}')
            break
    
    # If no income sheet found, check the main sheets
    if 'BS' in wb.sheetnames:
        # Check if there's an IS sheet or income statement in BS sheet
        ws = wb['BS']
        print(f'\nChecking BS sheet for income statement rows...')
        for row in range(45, 75):  # Income statement might be below balance sheet
            values = []
            for col in range(1, 5):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and 'revenue' in str(cell_value).lower() or 'income' in str(cell_value).lower() or 'sales' in str(cell_value).lower():
                    values.append(str(cell_value))
            if values:
                print(f'Row {row}: {" | ".join(values)}')
                
except Exception as e:
    print(f'Error: {e}') 