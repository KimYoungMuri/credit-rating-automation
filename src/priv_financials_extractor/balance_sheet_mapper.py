import pandas as pd
import numpy as np
from pathlib import Path
import re
import shutil
from openpyxl import load_workbook
from collections import defaultdict
import openpyxl
from datetime import datetime
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import sys
from typing import Dict, List, Tuple, Optional
from transformers import AutoTokenizer, AutoModel
import torch
import subprocess
import json

class BalanceSheetMapper:
    def __init__(self):
        self.used_items = set()
        
    def is_total_or_net_row(self, description: str, value: float = None, prev_values: list = None, tol: float = 1e-2, year: str = None) -> bool:
        """Exclude all lines containing 'total' except for 'Total Common Shareholders' Equity' (case and punctuation insensitive)."""
        desc_lower = description.lower().strip()
        # Normalize apostrophes and remove extra spaces for robust matching
        desc_norm = re.sub(r"[’'`´]", "'", desc_lower)
        desc_norm = re.sub(r"\s+", " ", desc_norm)
        if desc_norm in ('', '-', 'n/a', 'na', 'none'):
            return True
        # Exclude all lines with 'total' except for 'total common shareholders' equity' (case/punct insensitive)
        if 'total' in desc_norm:
            # Remove all non-alphanumeric except spaces and apostrophes for comparison
            def clean(s):
                return re.sub(r"[^a-z0-9' ]", '', s)
            if clean(desc_norm) != clean("total common shareholders' equity"):
                return True
        return False

    def apply_rule_based_mapping(self, description: str, section: str) -> tuple[str, float]:
        """Apply rule-based mapping using ONLY exact template names for each section"""
        import re
        desc_lower = description.lower()
        
        # Current Assets mapping
        if section == 'current_assets':
            if re.search(r'cash\s+(?:and\s+)?(?:cash\s+)?equivalents?', desc_lower):
                return 'Cash and equivalents', 0.95
            elif re.search(r'accounts?\s+receivable|notes?\s+receivable', desc_lower):
                return 'Accounts Receivable', 0.95
            elif re.search(r'prepaid', desc_lower):
                return 'Prepaid Expenses', 0.95
            elif re.search(r'inventor', desc_lower):
                return 'Inventory', 0.95
            elif re.search(r'margin\s+deposit|derivative\s+asset|investment', desc_lower):
                return 'Investments', 0.95
            else:
                return 'Other', 0.3  # Lower confidence to allow semantic fallback
        
        # Noncurrent Assets mapping
        elif section == 'noncurrent_assets':
            if re.search(r'net\s+ppe|property|equipment|right of use|lease asset|finance lease asset', desc_lower):
                return 'Net PPE', 0.95
            elif re.search(r'goodwill', desc_lower):
                return 'Goodwill', 0.95
            elif re.search(r'intangible', desc_lower):
                return 'Intangibles', 0.95
            else:
                return 'Other', 0.3  # Lower confidence to allow semantic fallback
        
        # Current Liabilities mapping
        elif section == 'current_liabilities':
            if re.search(r'accounts?\s+payable', desc_lower):
                return 'Accounts Payable', 0.95
            elif re.search(r'accrued|interest', desc_lower):
                return 'Accrued Interest', 0.95
            elif re.search(r'short[- ]term\s+borrow|revolving\s+lines?\s+of\s+credit', desc_lower):
                return 'Short term Borrowing', 0.95
            elif re.search(r'current\s+portion.*long[- ]term\s+debt|long[- ]term\s+debt.*current\s+portion', desc_lower):
                return 'Current Portion of Long Term Debt', 0.95
            elif re.search(r'taxes\s+payable|property\s+taxes\s+payable|excise\s+taxes\s+payable|sales\s+taxes\s+payable', desc_lower):
                return 'Accounts Payable', 0.9
            else:
                return 'Other', 0.3  # Lower confidence to allow semantic fallback
        
        # Noncurrent Liabilities mapping
        elif section == 'noncurrent_liabilities':
            if re.search(r'long[- ]term debt', desc_lower):
                return 'Long Term Debt', 0.95
            elif re.search(r'deferred.*tax', desc_lower):  # Only match deferred tax specifically
                return 'Deferred income taxes', 0.95
            elif re.search(r'finance lease liability', desc_lower):
                return 'Other', 0.9
            elif re.search(r'operating lease liability', desc_lower):
                return 'Other', 0.9
            else:
                return 'Other', 0.3  # Lower confidence to allow semantic fallback
        
        # Equity mapping
        elif section == 'equity':
            if re.search(r'common\s+stock', desc_lower):
                return 'Common Stock', 0.95
            elif re.search(r'retained\s+earnings', desc_lower):
                return 'Retained Earnings', 0.95
            elif re.search(r'paid[- ]in\s+capital', desc_lower):
                return 'Paid in Capital', 0.95
            elif re.search(r'total\s+common\s+shareholders.*equity', desc_lower):
                return 'Common Stock', 0.9  # Map to Common Stock as closest match
            else:
                return 'Other', 0.3  # Lower confidence to allow semantic fallback
        
        return None, 0.0

    def get_semantic_match(self, description: str, template_items: list) -> tuple[str, float]:
        """Get semantic match using sentence transformers"""
        if not template_items:
            return None, 0.0
        
        try:
            desc_embedding = self.model.encode([description])
            template_embeddings = self.model.encode(template_items)
            similarities = cosine_similarity(desc_embedding, template_embeddings)[0]
            best_idx = np.argmax(similarities)
            best_score = similarities[best_idx]
            
            if best_score > 0.5:  # Higher threshold for semantic matching
                return template_items[best_idx], best_score
            
            return None, 0.0
            
        except Exception as e:
            print(f"Error in semantic matching: {e}")
            return None, 0.0

    def call_ollama_disambiguate(self, line_item, section_above, section_below, template_above, template_below, model="mistral"):
        prompt = (
            f"You are an expert accountant. Given the following ambiguous balance sheet line item, decide which section it belongs to and which template line item is the best match.\n\n"
            f"Line item: \"{line_item}\"\n\n"
            f"Section above: \"{section_above}\"\n"
            f"Template line items for {section_above}:\n- " + "\n- ".join(template_above) + "\n\n"
            f"Section below: \"{section_below}\"\n"
            f"Template line items for {section_below}:\n- " + "\n- ".join(template_below) + "\n\n"
            f"Which section does this line item belong to ({section_above} or {section_below})? Which template line item in that section is the best match? Respond in JSON:\n{{\n  \"section\": \"...\",\n  \"template_line_item\": \"...\"\n}}\n"
        )
        try:
            result = subprocess.run([
                "ollama", "run", model, "--format", "json"],
                input=prompt.encode("utf-8"),
                capture_output=True,
                timeout=60
            )
            output = result.stdout.decode("utf-8")
            # Extract JSON from output
            json_start = output.find('{')
            json_end = output.rfind('}') + 1
            if json_start != -1 and json_end != -1:
                response = json.loads(output[json_start:json_end])
                return response.get("section"), response.get("template_line_item")
            else:
                print(f"[LLM-ERROR] Could not parse JSON from Ollama output: {output}")
                return None, None
        except Exception as e:
            print(f"[LLM-ERROR] Ollama call failed: {e}")
            return None, None

    def group_balance_sheet_items(self, year_data: Dict) -> Dict[str, List[Dict]]:
        """Group balance sheet items by section using two-pass rule-based, context-aware, and LLM disambiguation logic. Now uses sum tracking to detect/exclude total/subtotal rows."""
        sections = {
            'current_assets': [],
            'noncurrent_assets': [],
            'current_liabilities': [],
            'noncurrent_liabilities': [],
            'equity': []
        }
        section_keys = [
            "current_assets",
            "noncurrent_assets",
            "current_liabilities",
            "noncurrent_liabilities",
            "equity"
        ]
        section_names = [
            "Current Assets",
            "Noncurrent Assets",
            "Current Liabilities",
            "Noncurrent Liabilities",
            "Equity"
        ]
        template_line_items = {
            "current_assets": ["Cash and equivalents", "Accounts Receivable", "Prepaid Expenses", "Inventory", "Investments", "Other"],
            "noncurrent_assets": ["Net PPE", "Goodwill", "Intangibles", "Other"],
            "current_liabilities": ["Accounts Payable", "Accrued Interest", "Short term Borrowing", "Current Portion of Long Term Debt", "Other"],
            "noncurrent_liabilities": ["Long Term Debt", "Deferred income taxes", "Other"],
            "equity": ["Common Stock", "Retained Earnings", "Paid in Capital", "Other"]
        }
        def is_parsing_artifact(desc):
            return bool(re.match(r'^and \d{1,3}(,\d{3})* in \d{4} and \d{4}, respectively', desc.lower()))
        lines = list(year_data.items())
        # --- First pass: rule-based assignment ---
        assignments = []  # list of section_key or 'N/A'
        prev_values = []  # running list of previous values for sum tracking
        for desc, value in lines:
            # Exclude total/subtotal rows by sum tracking
            if self.is_total_or_net_row(desc, value, prev_values, year=desc):
                assignments.append(None)
                prev_values.append(0)  # placeholder, won't affect future sums
                continue
            if is_parsing_artifact(desc):
                assignments.append(None)
                prev_values.append(0)
                continue
            desc_lower = desc.lower()
            assigned_section = None
            
            # Check for liability patterns FIRST (before asset patterns)
            if any(k in desc_lower for k in ['accounts payable', 'accrued', 'taxes payable']):
                if any(k in desc_lower for k in ['noncurrent', 'long-term', 'long term']):
                    assigned_section = 'noncurrent_liabilities'
                else:
                    assigned_section = 'current_liabilities'
            elif any(k in desc_lower for k in ['liability', 'obligation', 'contingent']):
                if any(k in desc_lower for k in ['current', 'short term', 'current portion']):
                    assigned_section = 'current_liabilities'
                else:
                    assigned_section = 'noncurrent_liabilities'
            elif any(k in desc_lower for k in ['payable', 'accrued', 'taxes payable']):
                assigned_section = 'current_liabilities'
            elif any(k in desc_lower for k in ['debt', 'deferred', 'long-term', 'long term', 'noncurrent liability']):
                assigned_section = 'noncurrent_liabilities'
            elif any(k in desc_lower for k in ['deferred compensation', 'other noncurrent liabilities']):
                assigned_section = 'noncurrent_liabilities'
            # Then check for asset patterns
            elif any(re.search(r'inventor', desc_lower) or k in desc_lower for k in ['cash', 'accounts receivable', 'prepaid', 'margin deposit', 'derivative asset', 'notes receivable', 'other current asset']):
                assigned_section = 'current_assets'
            elif any(k in desc_lower for k in ['tax deposit', 'income tax deposit']):  # Tax deposits are assets
                assigned_section = 'noncurrent_assets'
            elif any(k in desc_lower for k in ['deferred compensation.*investment', 'deferred compensation.*plan']):  # Deferred comp investments are assets
                assigned_section = 'noncurrent_assets'
            elif any(k in desc_lower for k in ['net ppe', 'property', 'equipment', 'goodwill', 'intangible', 'right of use', 'finance lease', 'deferred compensation', 'other noncurrent']):
                assigned_section = 'noncurrent_assets'
            elif any(k in desc_lower for k in ['stock', 'capital', 'retained', 'equity', 'noncontrolling']):
                assigned_section = 'equity'
            
            assignments.append(assigned_section if assigned_section else 'N/A')
            prev_values.append(value if value is not None else 0)
        # --- Second pass: context-aware assignment for 'N/A' ---
        for idx, section in enumerate(assignments):
            if section == 'N/A':
                prev_section = next_section = None
                prev_idx = next_idx = None
                # Search backwards
                for j in range(idx-1, -1, -1):
                    if assignments[j] not in [None, 'N/A']:
                        prev_section = assignments[j]
                        prev_idx = j
                        break
                # Search forwards
                for j in range(idx+1, len(assignments)):
                    if assignments[j] not in [None, 'N/A']:
                        next_section = assignments[j]
                        next_idx = j
                        break
                # Assign based on context
                if prev_section and next_section and prev_section == next_section:
                    assignments[idx] = prev_section
                    print(f"[GROUP-CONTEXT] '{lines[idx][0]}' -> {prev_section} (neighbors agree)")
                elif prev_section and not next_section:
                    assignments[idx] = prev_section
                    print(f"[GROUP-CONTEXT] '{lines[idx][0]}' -> {prev_section} (prev only)")
                elif next_section and not prev_section:
                    assignments[idx] = next_section
                    print(f"[GROUP-CONTEXT] '{lines[idx][0]}' -> {next_section} (next only)")
                elif prev_section and next_section and prev_section != next_section:
                    # --- LLM disambiguation step ---
                    section_above = section_names[section_keys.index(prev_section)]
                    section_below = section_names[section_keys.index(next_section)]
                    template_above = template_line_items[prev_section]
                    template_below = template_line_items[next_section]
                    line_item = lines[idx][0]
                    llm_section, llm_template = self.call_ollama_disambiguate(
                        line_item, section_above, section_below, template_above, template_below
                    )
                    if llm_section and llm_template:
                        # Map section name back to section_key
                        llm_section_key = None
                        for k, n in zip(section_keys, section_names):
                            if n.lower() == llm_section.lower():
                                llm_section_key = k
                                break
                        if llm_section_key:
                            assignments[idx] = llm_section_key
                            print(f"[GROUP-LLM] '{line_item}' -> {llm_section_key} / {llm_template} (Ollama)")
                        else:
                            assignments[idx] = prev_section
                            print(f"[GROUP-LLM-FAIL] '{line_item}' -> {prev_section} (fallback to prev)")
                    else:
                        assignments[idx] = prev_section
                        print(f"[GROUP-LLM-FAIL] '{line_item}' -> {prev_section} (fallback to prev)")
                else:
                    assignments[idx] = 'current_assets'  # fallback, should be rare
                    print(f"[GROUP-CONTEXT] '{lines[idx][0]}' -> current_assets (fallback)")
        # --- Assign to sections dict ---
        for (desc, value), section in zip(lines, assignments):
            if section and section != 'N/A':
                sections[section].append({'description': desc, 'value': value})
        return sections

    def get_year_columns_from_template(self, bs_sheet) -> Dict[str, str]:
        """Extract year columns from the template"""
        year_cols = {}
        
        print(f"[DEBUG] Template sheet dimensions: {bs_sheet.max_row} rows x {bs_sheet.max_column} columns")
        
        # Try multiple rows for year detection
        for row_num in [6, 5, 7, 4, 8]:
            print(f"[DEBUG] Checking row {row_num} for year headers...")
            for col_idx in range(2, min(10, bs_sheet.max_column + 1)):
                cell_val = bs_sheet.cell(row=row_num, column=col_idx).value
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                
                if isinstance(cell_val, int) and 1990 <= cell_val <= 2050:
                    year_cols[str(cell_val)] = col_letter
                    print(f"[DEBUG] Found year {cell_val} in column {col_letter}")
                elif isinstance(cell_val, str):
                    # Handle Excel formulas
                    if cell_val.startswith('=') and '+1' in cell_val:
                        try:
                            base_cell = cell_val.split('+')[0][1:]
                            base_col = base_cell[0]
                            base_row = int(base_cell[1:])
                            base_year = bs_sheet[f"{base_col}{base_row}"].value
                            if isinstance(base_year, int):
                                col_offset = col_idx - openpyxl.utils.column_index_from_string(base_col)
                                calculated_year = base_year + col_offset
                                if 1990 <= calculated_year <= 2050:
                                    year_cols[str(calculated_year)] = col_letter
                                    print(f"[DEBUG] Found calculated year {calculated_year} in column {col_letter}")
                        except Exception as e:
                            pass
                    # Handle year strings
                    elif cell_val.isdigit() and len(cell_val) == 4 and 1990 <= int(cell_val) <= 2050:
                        year_cols[cell_val] = col_letter
                        print(f"[DEBUG] Found year string {cell_val} in column {col_letter}")
                    # Handle year-like strings with spaces or formatting
                    elif re.match(r'^\s*\d{4}\s*$', cell_val):
                        year_val = cell_val.strip()
                        if 1990 <= int(year_val) <= 2050:
                            year_cols[year_val] = col_letter
                            print(f"[DEBUG] Found formatted year {year_val} in column {col_letter}")
        
        print(f"[DEBUG] Final year columns found: {year_cols}")
        return year_cols

    def get_bs_row_map(self, sheet, start_row, end_row):
        """Get row map for a BS section"""
        row_map = {}
        for row in range(start_row, end_row + 1):
            cell_val = sheet.cell(row=row, column=1).value
            if cell_val:
                row_map[cell_val] = row
        # Add Net PPE if present
        for row in range(start_row, end_row + 1):
            cell_val = sheet.cell(row=row, column=1).value
            if cell_val and cell_val.strip().lower() == 'net ppe':
                row_map['Net PPE'] = row
        return row_map

    def map_balance_sheet(self, extracted_data: Dict, template_path: str) -> str:
        """Main balance sheet mapping function"""
        print("\n🚀 Starting Balance Sheet Mapping")
        print("=" * 60)
        
        # Load the template workbook
        shutil.copy(template_path, "temp_bs_template.xlsx")
        wb = load_workbook("temp_bs_template.xlsx")
        bs_sheet = wb['BS']

        # Get year columns from template
        year_cols = self.get_year_columns_from_template(bs_sheet)
        if not year_cols:
            print("❌ ERROR: Could not determine year columns from template. Aborting.")
            return ""

        print(f"✅ Found year columns: {year_cols}")

        # Create year mapping between extracted and template years
        extracted_years = list(extracted_data.keys())
        template_years = list(year_cols.keys())
        year_mapping = {}
        for i, extracted_year in enumerate(extracted_years):
            if i < len(template_years):
                template_year = template_years[i]
                year_mapping[extracted_year] = template_year
        
        print(f"✅ Year mapping: {year_mapping}")

        # Get template row mappings for each section
        row_maps = {
            'current_assets': self.get_bs_row_map(bs_sheet, 7, 12),
            'noncurrent_assets': self.get_bs_row_map(bs_sheet, 15, 18),
            'current_liabilities': self.get_bs_row_map(bs_sheet, 24, 28),
            'noncurrent_liabilities': self.get_bs_row_map(bs_sheet, 31, 33),
            'equity': self.get_bs_row_map(bs_sheet, 39, 42)
        }

        # Track mapping statistics
        stats = {
            'rule_based': 0,
            'semantic': 0,
            'unmapped': 0,
            'total_items': 0
        }

        # Initialize sentence transformer model once
        self.model = SentenceTransformer('all-MiniLM-L6-v2')

        # Process each year
        for extracted_year, mapped_year in year_mapping.items():
            if mapped_year not in year_cols:
                continue
            
            col = year_cols[mapped_year]
            year_data = extracted_data.get(extracted_year, {})
            
            if not isinstance(year_data, dict):
                continue
            
            print(f"\n📅 Processing year {extracted_year} -> {mapped_year}")
            
            # Group items by section
            section_items = self.group_balance_sheet_items(year_data)
            
            # Process each section
            for section, items in section_items.items():
                if not items:
                    continue
                
                row_map = row_maps.get(section)
                if not row_map:
                    continue
                
                template_items = list(row_map.keys())
                print(f"  [TEMPLATE ITEMS] Section: {section} -> {template_items}")
                
                # Map items in this section
                for item in items:
                    desc = item['description']
                    value = item['value']
                    
                    # Debug: Track "Accounts payable" specifically
                    if 'accounts payable' in desc.lower():
                        print(f"[DEBUG-AP] Found 'Accounts payable': '{desc}' = {value} in section {section}")
                    
                    if self.is_total_or_net_row(desc, value, None, year=extracted_year):
                        if 'accounts payable' in desc.lower():
                            print(f"[DEBUG-AP] 'Accounts payable' skipped as total/subtotal")
                        continue
                    
                    stats['total_items'] += 1
                    
                    # Try rule-based mapping first
                    template_item, confidence = self.apply_rule_based_mapping(desc, section)
                    
                    if 'accounts payable' in desc.lower():
                        print(f"[DEBUG-AP] 'Accounts payable' mapping result: {template_item} (confidence: {confidence})")
                    
                    # Use semantic matching if rule-based confidence is low or no match found
                    if not template_item or confidence < 0.5:
                        semantic_item, semantic_confidence = self.get_semantic_match(desc, template_items)
                        if semantic_item and semantic_confidence > 0.5:
                            template_item = semantic_item
                            confidence = semantic_confidence
                            print(f"[MAP-SEMANTIC] '{desc}' -> '{template_item}' (confidence: {semantic_confidence:.3f})")
                    
                    if template_item and template_item in row_map:
                        row_idx = row_map[template_item]
                        # --- SUM LOGIC: Add to existing value if present ---
                        existing_val = bs_sheet[f"{col}{row_idx}"].value or 0
                        if isinstance(existing_val, str):
                            try:
                                existing_val = float(existing_val.replace(',', ''))
                            except Exception:
                                existing_val = 0
                        bs_sheet[f"{col}{row_idx}"] = existing_val + value
                        if confidence >= 0.5:
                            stats['rule_based'] += 1
                            print(f"[MAP-SUCCESS] '{desc}' -> '{template_item}' [rule_based]")
                        else:
                            stats['semantic'] += 1
                            print(f"[MAP-SUCCESS] '{desc}' -> '{template_item}' [semantic]")
                    else:
                        # Add to "Other" category
                        if 'Other' in row_map:
                            row_idx = row_map['Other']
                            existing_val = bs_sheet[f"{col}{row_idx}"].value or 0
                            if isinstance(existing_val, str):
                                existing_val = 0
                            bs_sheet[f"{col}{row_idx}"] = existing_val + value
                            print(f"[MAP-OTHER] '{desc}' -> 'Other' [Other category]")
                        stats['unmapped'] += 1
                        print(f"[MAP-FAIL] '{desc}' could not be mapped to any template item.")

        # Print final statistics
        print("\n📊 MAPPING STATISTICS")
        print("=" * 60)
        print(f"Total items processed: {stats['total_items']}")
        print(f"Rule-based mappings: {stats['rule_based']} ({stats['rule_based']/max(1, stats['total_items'])*100:.1f}%)")
        print(f"Semantic mappings: {stats['semantic']} ({stats['semantic']/max(1, stats['total_items'])*100:.1f}%)")
        print(f"Unmapped items: {stats['unmapped']} ({stats['unmapped']/max(1, stats['total_items'])*100:.1f}%)")

        # Save the populated template
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        current_dir = Path(__file__).resolve().parent
        project_root = current_dir.parent.parent
        output_dir = project_root / "output_excel"
        output_dir.mkdir(exist_ok=True)
        
        populated_template_path = output_dir / f"balance_sheet_mapped_{timestamp}.xlsx"
        wb.save(populated_template_path)
        
        # Close the workbook and clean up
        wb.close()
        try:
            Path("temp_bs_template.xlsx").unlink()
        except (PermissionError, FileNotFoundError):
            pass
        
        print(f"\n✅ Balance sheet mapped and saved to: {populated_template_path}")
        return str(populated_template_path)

def main():
    # Get project root directory
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent.parent
    
    # Get paths
    template_path = project_root / "templates" / "financial_template.xlsx"
    if not template_path.exists():
        print(f"Template not found at {template_path}")
        return
        
    # Get most recent extracted Excel file or use command-line argument
    output_dir = project_root / "output_excel"
    if not output_dir.exists():
        print("No output directory found")
        return
        
    if len(sys.argv) > 1:
        latest_file = Path(sys.argv[1])
        if not latest_file.exists():
            print(f"Specified file does not exist: {latest_file}")
            return
    else:
        excel_files = [f for f in output_dir.glob("*.xlsx") if not f.name.startswith('~$')]
        if not excel_files:
            print("No valid (non-temporary) Excel files found in output directory")
            return
        # Sort by creation time and get most recent
        latest_file = max(excel_files, key=lambda x: x.stat().st_ctime)
    
    print(f"\nProcessing {latest_file}")
    
    # Read extracted data - focus only on balance sheet
    xls = pd.ExcelFile(latest_file)
    print(f"[DEBUG] Available sheets: {xls.sheet_names}")
    sheet_name = None
    for candidate in ['Balance Sheet', 'balance_sheet', 'Sheet1']:
        if candidate in xls.sheet_names:
            sheet_name = candidate
            break
    if sheet_name is None:
        print(f"[WARN] Could not find a standard balance sheet sheet name. Using first sheet: {xls.sheet_names[0]}")
        sheet_name = xls.sheet_names[0]
    else:
        print(f"[DEBUG] Using sheet: {sheet_name}")
    df = pd.read_excel(latest_file, sheet_name=sheet_name)
    
    # Find all year columns (exclude 'Description')
    year_cols = [col for col in df.columns if col != 'Description']
    print(f"[DEBUG] Found year columns for balance_sheet: {year_cols}")
    
    # Initialize year dictionaries
    extracted_data = {}
    for year in year_cols:
        extracted_data[str(year)] = {}
    
    # Process each row and populate all years
    for _, row in df.iterrows():
        desc = row['Description']
        if pd.notna(desc):
            for year in year_cols:
                if pd.notna(row.get(year)):
                    extracted_data[str(year)][desc] = row[year]
    
    print(f"[DEBUG] Loaded balance_sheet: {len(extracted_data)} years")
    for year, items in extracted_data.items():
        print(f"[DEBUG]   {year}: {len(items)} items")
    
    # Map balance sheet to template
    mapper = BalanceSheetMapper()
    output_path = mapper.map_balance_sheet(extracted_data, str(template_path))
    print(f"\nBalance sheet mapping completed: {output_path}")

if __name__ == "__main__":
    main() 