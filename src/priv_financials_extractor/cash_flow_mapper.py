import pandas as pd
import numpy as np
from pathlib import Path
import re
import shutil
from openpyxl import load_workbook
from collections import defaultdict
import openpyxl
from datetime import datetime
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import sys
from typing import Dict, List, Tuple, Optional

class CashFlowMapper:
    def __init__(self):
        self.model = SentenceTransformer('all-MiniLM-L6-v2')

    def is_excluded_row(self, description: str) -> bool:
        """Exclude subtotal, calculated, and net income rows in the cash flow statement."""
        desc = description.lower().strip()
        # Exclude all 'net cash from(used)' rows and anything below financing
        if re.match(r'net cash (from|used|provided|provided by|used for|from\(used\)).*', desc):
            return True
        if desc in [
            'net change in cash', 'starting cash', 'ending cash',
            'net change in cash and cash equivalents',
            'cash and cash equivalents at beginning of year',
            'cash and cash equivalents at end of year',
            'net income',  # Skip net income as well
        ]:
            return True
        return False

    def get_semantic_match(self, description: str, template_items: list) -> tuple[str, float]:
        if not template_items:
            return None, 0.0
        try:
            desc_embedding = self.model.encode([description])
            template_embeddings = self.model.encode(template_items)
            similarities = cosine_similarity(desc_embedding, template_embeddings)[0]
            best_idx = np.argmax(similarities)
            best_score = similarities[best_idx]
            if best_score > 0.5:
                return template_items[best_idx], best_score
            return None, 0.0
        except Exception as e:
            print(f"Error in semantic matching: {e}")
            return None, 0.0

    def apply_rule_based_mapping(self, description: str, section: str) -> tuple[str, float]:
        desc = description.lower()
        # Operating
        if section == 'operating':
            if 'net income' in desc:
                return 'Net Income', 0.95
            elif 'depreciation' in desc or 'amortization' in desc:
                return 'Changes in noncash items', 0.95
            elif 'change' in desc and ('asset' in desc or 'liabilit' in desc):
                return 'Changes in Asses and Liabilities', 0.95
            elif 'cash from operating' in desc:
                return 'Net Cash from(used) Operating Activities', 0.95
            else:
                return 'Others', 0.3
        # Investing
        elif section == 'investing':
            if 'capex' in desc or 'purchase of long-lived assets' in desc or 'capital expenditure' in desc:
                return 'CapEx', 0.95
            elif 'proceeds from asset sales' in desc or 'sale of property' in desc:
                return 'Proceeds from asset sales', 0.95
            elif 'cash from investing' in desc:
                return 'Net cash from(used) for investing', 0.95
            else:
                return 'Others', 0.3
        # Financing
        elif section == 'financing':
            if 'issuance of debt' in desc:
                return 'Issuance of Debt (long+short term)', 0.95
            elif 'retirement of debt' in desc or 'payments of long-term debt' in desc:
                return 'Retirement of Debt (long+short term)', 0.95
            elif 'issuance of stock' in desc:
                return 'Issuance of Stock', 0.95
            elif 'dividends paid' in desc:
                return 'Dividends Paid', 0.95
            elif 'cash from financing' in desc:
                return 'Net cash from(used) for financing', 0.95
            else:
                return 'Other', 0.3
        return None, 0.0

    def get_template_row_map(self, sheet, start_row, end_row):
        row_map = {}
        for row in range(start_row, end_row + 1):
            cell_val = sheet.cell(row=row, column=1).value
            if cell_val:
                row_map[cell_val] = row
        return row_map

    def get_year_columns_from_template(self, cf_sheet) -> Dict[str, str]:
        # Hard code year columns for 2023 and 2024
        return {'2023': 'B', '2024': 'C'}

    def map_cash_flow(self, extracted_data: Dict, template_path: str) -> str:
        print("\n🚀 Starting Cash Flow Statement Mapping\n" + "="*60)
        shutil.copy(template_path, "temp_cf_template.xlsx")
        wb = load_workbook("temp_cf_template.xlsx")
        cf_sheet = wb['IS.CF']
        year_cols = self.get_year_columns_from_template(cf_sheet)
        if not year_cols:
            print("❌ ERROR: Could not determine year columns from template. Aborting.")
            return ""
        print(f"✅ Found year columns: {year_cols}")
        extracted_years = list(extracted_data.keys())
        template_years = list(year_cols.keys())
        year_mapping = {}
        for i, extracted_year in enumerate(extracted_years):
            if i < len(template_years):
                template_year = template_years[i]
                year_mapping[extracted_year] = template_year
        print(f"✅ Year mapping: {year_mapping}")
        # Template row mappings
        row_maps = {
            'operating': self.get_template_row_map(cf_sheet, 25, 26),
            'investing': self.get_template_row_map(cf_sheet, 30, 32),
            'financing': self.get_template_row_map(cf_sheet, 36, 40)
        }
        # Section boundaries for grouping
        section_headers = {
            'operating': re.compile(r'^operating activities', re.I),
            'investing': re.compile(r'^investing activities', re.I),
            'financing': re.compile(r'^financing activities', re.I)
        }
        net_cash_patterns = [
            re.compile(r'^net cash (from|used|provided|provided by|used for|from\(used\)).*', re.I)
        ]
        # Process each year
        for extracted_year, mapped_year in year_mapping.items():
            if mapped_year not in year_cols:
                continue
            col = year_cols[mapped_year]
            year_data = extracted_data.get(extracted_year, {})
            if not isinstance(year_data, dict):
                continue
            print(f"\n📅 Processing year {extracted_year} -> {mapped_year}")
            # --- Group by section ---
            section = None
            section_items = {'operating': [], 'investing': [], 'financing': []}
            for desc, value in year_data.items():
                desc_norm = desc.lower().strip()
                if section_headers['operating'].match(desc_norm):
                    section = 'operating'
                    continue
                elif section_headers['investing'].match(desc_norm):
                    section = 'investing'
                    continue
                elif section_headers['financing'].match(desc_norm):
                    section = 'financing'
                    continue
                # Stop at first net cash after financing
                if section == 'financing' and any(p.match(desc_norm) for p in net_cash_patterns):
                    break
                if section and not self.is_excluded_row(desc):
                    print(f"[SECTION-ASSIGN] '{desc}' -> {section}")
                    section_items[section].append({'description': desc, 'value': value})
            # --- Map items in each section ---
            for section, items in section_items.items():
                if not items:
                    continue
                row_map = row_maps.get(section)
                if not row_map:
                    continue
                template_items = list(row_map.keys())
                print(f"  [TEMPLATE ITEMS] Section: {section} -> {template_items}")
                for item in items:
                    desc = item['description']
                    value = item['value']
                    if section == 'operating':
                        # Always classify as either noncash or assets/liabilities using semantic similarity
                        candidates = ['Changes in noncash items', 'Changes in Asses and Liabilities']
                        desc_embedding = self.model.encode([desc])
                        candidate_embeddings = self.model.encode(candidates)
                        similarities = cosine_similarity(desc_embedding, candidate_embeddings)[0]
                        best_idx = int(np.argmax(similarities))
                        template_item = candidates[best_idx]
                        print(f"[OPERATING-CLASSIFY] '{desc}' -> '{template_item}' (score: {similarities[best_idx]:.3f})")
                        confidence = similarities[best_idx]
                    elif section == 'investing':
                        # Always classify as CapEx, Proceeds from asset sales, or Others using semantic similarity
                        candidates = ['CapEx', 'Proceeds from asset sales', 'Others']
                        desc_embedding = self.model.encode([desc])
                        candidate_embeddings = self.model.encode(candidates)
                        similarities = cosine_similarity(desc_embedding, candidate_embeddings)[0]
                        best_idx = int(np.argmax(similarities))
                        template_item = candidates[best_idx]
                        print(f"[INVESTING-CLASSIFY] '{desc}' -> '{template_item}' (score: {similarities[best_idx]:.3f})")
                        confidence = similarities[best_idx]
                    else:
                        # Rule-based mapping
                        template_item, confidence = self.apply_rule_based_mapping(desc, section)
                        # Use semantic matching if rule-based confidence is low
                        if not template_item or confidence < 0.5:
                            semantic_item, semantic_confidence = self.get_semantic_match(desc, template_items)
                            if semantic_item and semantic_confidence > 0.5:
                                template_item = semantic_item
                                confidence = semantic_confidence
                                print(f"[MAP-SEMANTIC] '{desc}' -> '{template_item}' (confidence: {semantic_confidence:.3f})")
                    if template_item and template_item in row_map:
                        row_idx = row_map[template_item]
                        existing_val = cf_sheet[f"{col}{row_idx}"].value or 0
                        if isinstance(existing_val, str):
                            try:
                                existing_val = float(existing_val.replace(',', ''))
                            except Exception:
                                existing_val = 0
                        cf_sheet[f"{col}{row_idx}"] = existing_val + value
                        print(f"[MAP-SUCCESS] '{desc}' -> '{template_item}'")
                    else:
                        # Add to 'Other' if available
                        if 'Other' in row_map:
                            row_idx = row_map['Other']
                            existing_val = cf_sheet[f"{col}{row_idx}"].value or 0
                            if isinstance(existing_val, str):
                                existing_val = 0
                            cf_sheet[f"{col}{row_idx}"] = existing_val + value
                            print(f"[MAP-OTHER] '{desc}' -> 'Other'")
                        else:
                            print(f"[MAP-FAIL] '{desc}' could not be mapped.")
        # Save
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        current_dir = Path(__file__).resolve().parent
        project_root = current_dir.parent.parent
        output_dir = project_root / "output_excel"
        output_dir.mkdir(exist_ok=True)
        populated_template_path = output_dir / f"cash_flow_mapped_{timestamp}.xlsx"
        wb.save(populated_template_path)
        wb.close()
        try:
            Path("temp_cf_template.xlsx").unlink()
        except (PermissionError, FileNotFoundError):
            pass
        print(f"\n✅ Cash flow mapped and saved to: {populated_template_path}")
        return str(populated_template_path)

def main():
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent.parent
    template_path = project_root / "templates" / "financial_template.xlsx"
    if not template_path.exists():
        print(f"Template not found at {template_path}")
        return
    output_dir = project_root / "output_excel"
    if not output_dir.exists():
        print("No output directory found")
        return
    if len(sys.argv) > 1:
        latest_file = Path(sys.argv[1])
        if not latest_file.exists():
            print(f"Specified file does not exist: {latest_file}")
            return
    else:
        excel_files = [f for f in output_dir.glob("*.xlsx") if not f.name.startswith('~$')]
        if not excel_files:
            print("No valid (non-temporary) Excel files found in output directory")
            return
        latest_file = max(excel_files, key=lambda x: x.stat().st_ctime)
    print(f"\nProcessing {latest_file}")
    xls = pd.ExcelFile(latest_file)
    print(f"[DEBUG] Available sheets: {xls.sheet_names}")
    sheet_name = None
    for candidate in ['Cash Flow', 'cash_flow', 'Sheet1']:
        if candidate in xls.sheet_names:
            sheet_name = candidate
            break
    if sheet_name is None:
        print(f"[WARN] Could not find a standard cash flow sheet name. Using first sheet: {xls.sheet_names[0]}")
        sheet_name = xls.sheet_names[0]
    else:
        print(f"[DEBUG] Using sheet: {sheet_name}")
    df = pd.read_excel(latest_file, sheet_name=sheet_name)
    year_cols = [col for col in df.columns if col != 'Description']
    print(f"[DEBUG] Found year columns for cash_flow: {year_cols}")
    extracted_data = {}
    for year in year_cols:
        extracted_data[str(year)] = {}
    for _, row in df.iterrows():
        desc = row['Description']
        if pd.notna(desc):
            for year in year_cols:
                if pd.notna(row.get(year)):
                    extracted_data[str(year)][desc] = row[year]
    print(f"[DEBUG] Loaded cash_flow: {len(extracted_data)} years")
    for year, items in extracted_data.items():
        print(f"[DEBUG]   {year}: {len(items)} items")
    mapper = CashFlowMapper()
    output_path = mapper.map_cash_flow(extracted_data, str(template_path))
    print(f"\nCash flow mapping completed: {output_path}")

if __name__ == "__main__":
    main() 