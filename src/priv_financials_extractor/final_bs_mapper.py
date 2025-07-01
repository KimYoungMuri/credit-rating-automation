#!/usr/bin/env python3
"""
Final Knowledge Graph Mapper - Integrated with Original Template
- Uses the original financial_template.xlsx from templates/ directory
- Enhanced pattern coverage and section assignment
- Multi-item consolidation and template mapping
- Uses shutil for template management
- Populates the actual original template structure
"""

import json
import re
import shutil
import os
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from final_extractor_adaptive import TextExtractor
from final_find_fs import FinancialStatementFinder

@dataclass
class MappedValue:
    """Represents a mapped financial value with section context"""
    original_description: str
    template_field: str
    section: str
    value_2023: Optional[float] = None
    value_2024: Optional[float] = None
    confidence: float = 1.0
    mapping_method: str = ""
    source_data: dict = None

class FinalKGMapper:
    """Final mapper integrated with original financial template"""
    
    def __init__(self):
        self.extractor = TextExtractor()
        self.finder = FinancialStatementFinder()
        
        # Set up template paths
        self.original_template_path = Path("../../templates/financial_template.xlsx")
        self.working_template_path = Path("./working_financial_template.xlsx")
        
        # Template field mappings - based on the original template structure
        self.template_mappings = {
            # ASSETS section
            'Cash and equivalents': ('B', 7),  # Row 7, Column B for 2023, C for 2024
            'Accounts Receivable': ('B', 8),
            'Prepaid Expenses': ('B', 9),
            'Inventory': ('B', 10),
            'Investments': ('B', 11),
            'Other_current_assets': ('B', 12),  # Other current assets
            # Total Current Assets calculated
            
            'Net PPE': ('B', 15),
            'Goodwill': ('B', 16),
            'Intangibles': ('B', 17),
            'Other_noncurrent_assets': ('B', 18),  # Other non-current assets
            # Total Non Current Assets calculated
            # Total Assets calculated
            
            # LIABILITIES section  
            'Accounts Payable': ('B', 24),
            'Accrued Interest': ('B', 25),
            'Short term Borrowing': ('B', 26),
            'Current Portion of Long Term Debt': ('B', 27),
            'Other_current_liabilities': ('B', 28),  # Other current liabilities
            # Total Current Liabilities calculated
            
            'Long Term Debt': ('B', 31),
            'Deferred income taxes': ('B', 32),
            'Other_noncurrent_liabilities': ('B', 33),  # Other non-current liabilities
            # Total Non Current Liabilities calculated
            # Total Liabilities calculated
            
            # EQUITY section
            'Common Stock': ('B', 39),
            'Retained Earnings': ('B', 40),
            'Paid in Capital': ('B', 41),
            'Other_equity': ('B', 42),  # Other equity
            # Total Equity calculated
        }
        
        # Enhanced rule-based patterns with complete coverage
        self.bs_rules = {
            # === CURRENT ASSETS ===
            r'cash\s+(?:and\s+)?(?:cash\s+)?equivalents?': ('Cash and equivalents', 'current_assets'),
            r'accounts?\s+receivable(?:[—-]net)?': ('Accounts Receivable', 'current_assets'),
            r'notes?\s+receivable[—-]current\s+portion': ('Other', 'current_assets'),
            r'prepaid\s+expenses?': ('Prepaid Expenses', 'current_assets'),
            r'prepaid\s+insurance': ('Prepaid Expenses', 'current_assets'),
            r'prepaid\s+assets?': ('Prepaid Expenses', 'current_assets'),
            r'advances?\s+(?:to|and)\s+': ('Prepaid Expenses', 'current_assets'),
            r'inventor(?:y|ies)(?:[—-]net)?': ('Inventory', 'current_assets'),
            r'margin\s+deposits?': ('Investments', 'current_assets'),
            r'derivative\s+assets?': ('Investments', 'current_assets'),
            r'other\s+current\s+assets?': ('Other', 'current_assets'),
            
            # === NON-CURRENT ASSETS ===
            r'notes?\s+receivable(?!\s*[—-]current)': ('Other', 'noncurrent_assets'),  # Non-current notes receivable
            r'property\s+(?:and\s+)?equipment(?:[—-]net)?': ('Net PPE', 'noncurrent_assets'),
            r'right\s+of\s+use\s+assets?': ('Net PPE', 'noncurrent_assets'),
            r'finance\s+lease\s+assets?': ('Net PPE', 'noncurrent_assets'),
            r'goodwill(?:[—-]net)?': ('Goodwill', 'noncurrent_assets'),
            r'(?:other\s+)?intangible\s+assets?(?:[—-]net)?': ('Intangibles', 'noncurrent_assets'),
            r'deferred\s+compensation\s+plan': ('Other', 'noncurrent_assets'),
            r'other\s+noncurrent\s+assets?': ('Other', 'noncurrent_assets'),
            
            # === CURRENT LIABILITIES ===
            r'accounts?\s+payable': ('Accounts Payable', 'current_liabilities'),
            r'accrued\s+(?:liabilities?|interest)': ('Accrued Interest', 'current_liabilities'),
            r'sales?\s*,?\s*excise\s+.*?taxes?\s+payable': ('Accrued Interest', 'current_liabilities'),
            r'revolving\s+lines?\s+of\s+credit': ('Long Term Debt', 'noncurrent_liabilities'),
            r'current\s+portion\s+of\s+long[- ]term\s+debt': ('Current Portion of Long Term Debt', 'current_liabilities'),
            r'long[- ]term\s+debt[—-]current\s+portion': ('Current Portion of Long Term Debt', 'current_liabilities'),
            r'finance\s+lease\s+liability[—-]current': ('Current Portion of Long Term Debt', 'current_liabilities'),
            r'operating\s+lease\s+liability[—-]current': ('Current Portion of Long Term Debt', 'current_liabilities'),
            r'derivative\s+liabilities?': ('Other', 'current_liabilities'),
            r'contingent\s+consideration': ('Other', 'current_liabilities'),
            r'long[- ]term\s+incentive[—-]current': ('Other', 'current_liabilities'),
            
            # === NON-CURRENT LIABILITIES ===
            r'long[- ]term\s+debt(?!.*current)': ('Long Term Debt', 'noncurrent_liabilities'),
            r'finance\s+lease\s+liability(?!.*current)': ('Long Term Debt', 'noncurrent_liabilities'),
            r'operating\s+lease\s+liability(?!.*current)': ('Long Term Debt', 'noncurrent_liabilities'),
            r'deferred\s+income\s+taxes?': ('Deferred income taxes', 'noncurrent_liabilities'),
            r'deferred\s+compensation(?!.*plan)': ('Other', 'noncurrent_liabilities'),
            r'long[- ]term\s+incentive(?!.*current)': ('Other', 'noncurrent_liabilities'),
            r'other\s+noncurrent\s+liabilities?': ('Other', 'noncurrent_liabilities'),
            
            # === EQUITY ===
            r'common\s+stock': ('Common Stock', 'equity'),
            r'retained\s+earnings?': ('Retained Earnings', 'equity'),
            r'paid[- ]in\s+capital': ('Paid in Capital', 'equity'),
            r'total\s+common\s+shareholders?\s*[\'\u2019]?\s*equity': ('Common Stock', 'equity'),
            r'noncontrolling\s+interests?': ('Other', 'equity'),
            
            # === MISC/OTHER ITEMS ===
            r'subchapter\s+s\s+income\s+tax.*obligation': ('Other', 'current_liabilities'),  # Tax obligation (liability)
        }
        
        # Template structure mapping sections to their "Other" fields
        self.section_other_mapping = {
            'current_assets': 'Other',           
            'noncurrent_assets': 'Other',        
            'current_liabilities': 'Other',      
            'noncurrent_liabilities': 'Other',   
            'equity': 'Other'                    
        }
    
    def setup_template(self) -> bool:
        """Copy original template to working directory using shutil"""
        try:
            print(f"📋 Setting up template...")
            print(f"   Source: {self.original_template_path}")
            print(f"   Working copy: {self.working_template_path}")
            
            if not self.original_template_path.exists():
                print(f"❌ Original template not found: {self.original_template_path}")
                return False
            
            # Use shutil to copy the template
            shutil.copy2(self.original_template_path, self.working_template_path)
            
            if self.working_template_path.exists():
                print(f"✅ Template copied successfully")
                return True
            else:
                print(f"❌ Failed to copy template")
                return False
                
        except Exception as e:
            print(f"❌ Error setting up template: {e}")
            return False
    
    def is_total_or_net_row(self, description: str) -> bool:
        """Check if description is a total or net row (from final_template_mapper.py)"""
        desc_lower = description.lower().strip()
        
        # Special exception: TOTAL COMMON SHAREHOLDERS' EQUITY maps to Common Stock
        if 'total' in desc_lower and 'shareholders' in desc_lower and 'equity' in desc_lower:
            return False  # Don't filter this one out
            
        # Expanded patterns for totals/subtotals
        total_patterns = [
            r'^total(\s|$)',
            r'(\s|^)total\s',
            r'(\s|^)sum(\s|$)',
            r'(\s|^)subtotal(\s|$)',
            r'(\s|^)net(\s|$)',
            r'(\s|^)aggregate(\s|$)',
            r'(\s|^)grand total(\s|$)',
            r'(\s|^)overall(\s|$)',
            r'(\s|^)balance(\s|$)',
        ]
        for pat in total_patterns:
            if re.search(pat, desc_lower):
                return True
        
        # Filter out header/formatting rows
        header_patterns = [
            r'^\s*\d{4}\s+\d{4}\s*$',  # Year headers like "2024 2023"
            r'^\s*assets?\s*$',         # Section headers
            r'^\s*liabilities?\s*$',
            r'^\s*equity\s*$',
            r'continued|concluded',     # Page markers
            r'^\s*-\s*\d+\s*-\s*$',    # Page numbers like "- 5 -"
            r'amounts\s+in\s+thousands',
            r'see\s+notes\s+to',
            r'consolidated\s+balance\s+sheets',
        ]
        for pat in header_patterns:
            if re.search(pat, desc_lower):
                return True
        
        return False
    
    def apply_enhanced_mapping(self, description: str) -> Tuple[Optional[str], Optional[str], float]:
        """Apply enhanced rule-based mapping with section assignment"""
        desc_lower = description.lower().strip()
        
        for pattern, (template_field, section) in self.bs_rules.items():
            if re.search(pattern, desc_lower):
                return template_field, section, 0.9
        
        return None, None, 0.0
    
    def consolidate_multi_mappings(self, mapped_items: Dict[str, MappedValue]) -> Dict[str, MappedValue]:
        """Consolidate multiple items that map to the same template field"""
        consolidated = {}
        
        # Group by template field AND section for "Other" items, just template field for others
        field_groups = {}
        for key, mapped_value in mapped_items.items():
            field = mapped_value.template_field
            section = mapped_value.section
            
            # For "Other" items, group by both field and section
            if field == 'Other':
                group_key = f"{field}_{section}"
            else:
                group_key = field
                
            if group_key not in field_groups:
                field_groups[group_key] = []
            field_groups[group_key].append((key, mapped_value))
        
        # Consolidate each field group
        for group_key, items in field_groups.items():
            if len(items) == 1:
                # Single item - keep as is
                key, mapped_value = items[0]
                consolidated[key] = mapped_value
            else:
                # Multiple items - consolidate
                total_2023 = sum(mv.value_2023 for k, mv in items if mv.value_2023)
                total_2024 = sum(mv.value_2024 for k, mv in items if mv.value_2024)
                
                descriptions = [mv.original_description for k, mv in items]
                sections = [mv.section for k, mv in items]
                primary_section = max(set(sections), key=sections.count)  # Most common section
                
                # Get field name from first item
                field = items[0][1].template_field
                
                consolidated_value = MappedValue(
                    original_description=f"Consolidated: {'; '.join(descriptions[:2])}{'...' if len(descriptions) > 2 else ''}",
                    template_field=field,
                    section=primary_section,
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.85,
                    mapping_method="multi_item_consolidation",
                    source_data={"consolidated_count": len(items), "items": descriptions}
                )
                
                consolidated[group_key] = consolidated_value
                
                print(f"🔗 Consolidated {len(items)} items → {field} ({primary_section})")
                print(f"   Items: {', '.join(descriptions)}")
                v23 = f"${total_2023:,.0f}" if total_2023 else "-"
                v24 = f"${total_2024:,.0f}" if total_2024 else "-"
                print(f"   Total: 2023={v23}, 2024={v24}")
                print()
        
        return consolidated

    def extract_and_process(self, pdf_path: str) -> Dict[str, MappedValue]:
        """Main processing function with enhanced logic"""
        print("FINAL KNOWLEDGE GRAPH MAPPER - ORIGINAL TEMPLATE INTEGRATION")
        print("=" * 70)
        print("Features:")
        print("1. Uses original financial_template.xlsx from templates/")
        print("2. Enhanced pattern coverage and section assignment")  
        print("3. Multi-item consolidation")
        print("4. Template management with shutil")
        print("5. Populates actual template structure")
        print()
        
        # Setup template
        if not self.setup_template():
            print("❌ Failed to setup template")
            return {}
        
        # Extract data using final_extractor
        confirmed_pages = {
            'balance_sheet': [7, 8],
            'income_statement': [9], 
            'cash_flow': [11, 12]
        }
        
        statement_pages_dict = {}
        for stmt_type, pages in confirmed_pages.items():
            if pages:
                statement_pages_dict[stmt_type] = pages
        
        excel_path, extracted_data = self.extractor.extract_text(
            pdf_path, 
            process_numbers=True, 
            statement_pages=statement_pages_dict
        )
        
        if not extracted_data or 'balance_sheet' not in extracted_data:
            print("❌ No balance sheet data found")
            return {}
        
        balance_sheet_data = extracted_data['balance_sheet']
        print(f"✅ Extracted {len(balance_sheet_data)} balance sheet items")
        
        # Step 1: Filter out totals/subtotals and headers
        non_total_items = []
        total_items = []
        
        for item in balance_sheet_data:
            description = item.get('description', '').strip()
            if not description:
                continue
                
            if self.is_total_or_net_row(description):
                total_items.append(item)
            else:
                non_total_items.append(item)
        
        print(f"📊 After filtering: {len(non_total_items)} items to map, {len(total_items)} totals/headers filtered")
        print()
        
        # Step 2: Filter items that have numerical values only
        items_with_numbers = []
        for item in non_total_items:
            description = item.get('description', '').strip()
            numbers = item.get('numbers', {})
            
            # Parse values and check if we have any valid numbers
            has_numbers = False
            value_2023 = None
            value_2024 = None
            for year, value_str in numbers.items():
                if value_str is not None:
                    try:
                        value = float(str(value_str).replace(',', ''))
                        has_numbers = True
                        if year == '2023':
                            value_2023 = value
                        elif year == '2024':
                            value_2024 = value
                    except (ValueError, TypeError):
                        continue
            
            # Only process items that have numerical values
            if has_numbers:
                items_with_numbers.append((item, description, value_2023, value_2024))
        
        print(f"📊 Items with numerical values: {len(items_with_numbers)}")
        
        # Step 3: Process items with numerical values
        mapped_items = {}
        unmapped_by_section = {}
        
        print("🔄 Enhanced Processing:")
        print("-" * 50)
        
        for item, description, value_2023, value_2024 in items_with_numbers:
            
            # Enhanced mapping with section assignment
            template_field, section, confidence = self.apply_enhanced_mapping(description)
            
            if template_field and section and confidence >= 0.8:
                # Successfully mapped with section
                mapped_value = MappedValue(
                    original_description=description,
                    template_field=template_field,
                    section=section,
                    value_2023=value_2023,
                    value_2024=value_2024,
                    confidence=confidence,
                    mapping_method="enhanced_rule_based",
                    source_data=item
                )
                
                # Create unique key for consolidation later
                key = f"{template_field}_{section}_{len(mapped_items)}"
                mapped_items[key] = mapped_value
                
                print(f"✅ {description[:45]}...")
                print(f"   → {template_field} (section: {section})")
                v23 = f"${value_2023:,.0f}" if value_2023 else "-"
                v24 = f"${value_2024:,.0f}" if value_2024 else "-"
                print(f"   Values: 2023={v23}, 2024={v24}")
                print()
            else:
                # Unmapped - determine section heuristically
                inferred_section = self.infer_section_from_context(description)
                
                if inferred_section not in unmapped_by_section:
                    unmapped_by_section[inferred_section] = []
                unmapped_by_section[inferred_section].append((description, value_2023, value_2024, item))
                
                print(f"❓ {description[:45]}...")
                print(f"   → Will map to 'Other' in {inferred_section}")
        
        # Step 3: Consolidate multi-mappings
        print(f"\n🔗 Consolidating multi-item mappings:")
        print("-" * 50)
        consolidated_mapped = self.consolidate_multi_mappings(mapped_items)
        
        # Step 4: Handle unmapped items - consolidate into "Other" fields by section
        print(f"\n🔧 Consolidating unmapped items into 'Other' fields:")
        print("-" * 50)
        
        for section, unmapped_items in unmapped_by_section.items():
            if section == 'unknown':
                # FALLBACK: Put unknown items in Other_noncurrent_assets
                # This is typically the safest section for unclear items
                if unmapped_items:
                    section = 'noncurrent_assets'  # Override to noncurrent_assets
                    print(f"🔄 unknown → Other_noncurrent_assets (fallback)")
                    print(f"   Moving {len(unmapped_items)} unknown items to noncurrent assets as fallback")
                else:
                    continue  # Skip if no items
                
            other_field = self.section_other_mapping.get(section)
            if not other_field:
                continue
                
            # Consolidate all unmapped items in this section
            total_2023 = sum(item[1] for item in unmapped_items if item[1] is not None)
            total_2024 = sum(item[2] for item in unmapped_items if item[2] is not None)
            
            descriptions = [item[0] for item in unmapped_items]
            
            # Create consolidated "Other" entry
            other_key = f"{other_field}_{section}"
            consolidated_mapped[other_key] = MappedValue(
                original_description=f"Consolidated {len(unmapped_items)} unmapped items",
                template_field=other_field,
                section=section,
                value_2023=total_2023 if total_2023 != 0 else None,
                value_2024=total_2024 if total_2024 != 0 else None,
                confidence=0.75,
                mapping_method="section_other_consolidation",
                source_data={"consolidated_items": descriptions}
            )
            
            print(f"🔄 {section} → {other_field}")
            print(f"   Consolidated {len(unmapped_items)} unmapped items")
            v23 = f"${total_2023:,.0f}" if total_2023 and total_2023 != 0 else "-"
            v24 = f"${total_2024:,.0f}" if total_2024 and total_2024 != 0 else "-"
            print(f"   Total values: 2023={v23}, 2024={v24}")
            print()
        
        return consolidated_mapped
    
    def infer_section_from_context(self, description: str) -> str:
        """Infer section from description context with enhanced fallbacks"""
        desc_lower = description.lower()
        
        # TIER 1: Keyword-based inference (enhanced coverage)
        # Current assets indicators
        if any(word in desc_lower for word in ['current', 'receivable', 'inventory', 'prepaid', 'cash', 'advances', 'deposits']):
            return 'current_assets'
        
        # Current liabilities indicators  
        if any(word in desc_lower for word in ['payable', 'accrued', 'current portion', 'short', 'taxes payable']):
            return 'current_liabilities'
        
        # Non-current assets indicators
        if any(word in desc_lower for word in ['property', 'equipment', 'goodwill', 'intangible', 'investment', 'lease assets']):
            return 'noncurrent_assets'
        
        # Non-current liabilities indicators
        if any(word in desc_lower for word in ['long-term', 'lease liability', 'deferred', 'compensation', 'incentive']):
            return 'noncurrent_liabilities'
        
        # Equity indicators
        if any(word in desc_lower for word in ['stock', 'equity', 'earnings', 'capital', 'shareholders']):
            return 'equity'
        
        # TIER 2: Position context fallback
        position_section = self.infer_from_position_context(description)
        if position_section != 'unknown':
            print(f"   📍 Position context: {description[:30]}... → {position_section}")
            return position_section
        
        # TIER 3: LLM fallback for truly unknown items
        llm_section = self.ask_ollama_for_section(description)
        if llm_section != 'unknown':
            print(f"   🤖 Ollama inference: {description[:30]}... → {llm_section}")
            return llm_section
        
        # TIER 4: Smart keyword fallback when Ollama fails
        smart_section = self.smart_section_fallback(description)
        if smart_section != 'unknown':
            print(f"   🧠 Smart fallback: {description[:30]}... → {smart_section}")
            return smart_section
        
        return 'unknown'
    
    def infer_from_position_context(self, description: str) -> str:
        """Infer section based on position context in the balance sheet"""
        # This requires access to the full balance sheet data and current position
        # For now, we'll implement basic heuristics based on common patterns
        
        desc_lower = description.lower()
        
        # Look for position clues in the description itself
        # Items that appear with asset-like context
        if any(phrase in desc_lower for phrase in [
            'net of', 'allowance', 'accumulated', 'amortization', 'depreciation',
            'cost basis', 'fair value', 'carrying amount'
        ]):
            return 'noncurrent_assets'  # Usually asset-related
        
        # Items that appear with liability-like context  
        if any(phrase in desc_lower for phrase in [
            'obligation', 'payable to', 'due to', 'liability for',
            'reserve for', 'provision for'
        ]):
            return 'noncurrent_liabilities'  # Usually liability-related
        
        # Items with current timing indicators
        if any(phrase in desc_lower for phrase in [
            'within one year', 'next 12 months', 'due within',
            'maturing', 'short-term'
        ]):
            # Need to determine if asset or liability based on other context
            if any(word in desc_lower for word in ['due', 'payable', 'owed']):
                return 'current_liabilities'
            else:
                return 'current_assets'
        
        return 'unknown'  # Still couldn't determine
    
    def ask_ollama_for_section(self, description: str) -> str:
        """Ask Ollama LLM to classify the line item section"""
        try:
            import requests
            import json
            
            # Simple prompt for fast processing
            
            prompt = f"""Classify this balance sheet item into ONE section:

"{description}"

Choose EXACTLY one:
- current_assets
- noncurrent_assets 
- current_liabilities
- noncurrent_liabilities
- equity

Answer with just the section name."""

            # Call Ollama API
            response = requests.post(
                'http://localhost:11434/api/generate',
                json={
                    'model': 'phi3:mini',
                    'prompt': prompt,
                    'stream': False,
                    'options': {
                        'temperature': 0.1,  # Low temperature for consistent classification
                        'num_predict': 20    # Short response
                    }
                },
                timeout=25  # Optimized timeout for phi3:mini
            )
            
            if response.status_code == 200:
                result = response.json()
                ollama_response = result.get('response', '').strip().lower()
                
                # Extract section from response
                valid_sections = ['current_assets', 'noncurrent_assets', 'current_liabilities', 
                                'noncurrent_liabilities', 'equity']
                for section in valid_sections:
                    if section in ollama_response:
                        return section
                        
            print(f"   ⚠️ Ollama API call failed or unclear response")
            return 'unknown'
            
        except Exception as e:
            print(f"   ⚠️ Ollama fallback failed: {e}")
            return 'unknown'
    
    def smart_section_fallback(self, description: str) -> str:
        """Intelligent section assignment fallback using position context and keywords"""
        desc_lower = description.lower()
        
        # Enhanced keyword-based section detection
        section_keywords = {
            'current_assets': [
                'cash', 'receivable', 'inventory', 'prepaid', 'margin', 'derivative assets',
                'current portion', 'short-term', 'deposit.*current'
            ],
            'noncurrent_assets': [
                'property', 'equipment', 'goodwill', 'intangible', 'investment', 'ppe',
                'right of use', 'finance lease assets', 'deferred compensation plan',
                'deposit(?!.*current)', 'long-term assets'
            ],
            'current_liabilities': [
                'payable', 'accrued', 'taxes payable', 'current portion', 'short-term debt',
                'obligation.*current', 'liability.*current'
            ],
            'noncurrent_liabilities': [
                'long-term debt', 'long-term incentive', 'deferred', 'revolving',
                'finance lease liability(?!.*current)', 'operating lease liability(?!.*current)',
                'long-term', 'noncurrent liabilities'
            ],
            'equity': [
                'shareholders', 'common stock', 'retained earnings', 'paid.*capital',
                'equity', 'noncontrolling'
            ]
        }
        
        # Score each section based on keyword matches
        section_scores = {}
        for section, keywords in section_keywords.items():
            score = 0
            for keyword in keywords:
                if re.search(keyword, desc_lower):
                    score += 1
            section_scores[section] = score
        
        # Return section with highest score
        if section_scores:
            best_section = max(section_scores, key=section_scores.get)
            if section_scores[best_section] > 0:
                return best_section
        
        # No special case handling - let positional inference work
            
        # Default fallback based on common patterns
        if any(word in desc_lower for word in ['asset', 'receivable', 'cash', 'inventory']):
            return 'current_assets'
        elif any(word in desc_lower for word in ['payable', 'liability', 'debt', 'obligation']):
            return 'current_liabilities'
        elif any(word in desc_lower for word in ['equity', 'stock', 'earnings']):
            return 'equity'
        
        return 'unknown'
    
    def populate_original_template(self, mapped_items: Dict[str, MappedValue]) -> str:
        """Populate the original template with mapped values"""
        if not self.working_template_path.exists():
            print("❌ Working template not found")
            return ""
        
        try:
            print("\n📝 Populating original template structure...")
            
            # Load the Excel workbook
            wb = load_workbook(self.working_template_path)
            
            # Get the first worksheet (assuming financial data is in first sheet)
            ws = wb.active
            print(f"   Working with sheet: {ws.title}")
            
            # Track populated fields and totals for verification
            populated_fields = []
            section_totals = {
                'current_assets': {'2023': 0, '2024': 0},
                'noncurrent_assets': {'2023': 0, '2024': 0}, 
                'current_liabilities': {'2023': 0, '2024': 0},
                'noncurrent_liabilities': {'2023': 0, '2024': 0},
                'equity': {'2023': 0, '2024': 0}
            }
            
            # Process each KG mapping
            for key, mapping in mapped_items.items():
                template_field = mapping.template_field
                value_2023 = mapping.value_2023
                value_2024 = mapping.value_2024
                section = mapping.section
                
                # Handle "Other" fields by section
                if template_field == 'Other':
                    template_key = f"Other_{section}"
                else:
                    template_key = template_field
                
                # Check if we have a mapping for this template field
                if template_key in self.template_mappings:
                    col_letter, row_num = self.template_mappings[template_key]
                    
                    # Populate 2023 value (column B)
                    if value_2023 is not None:
                        cell_2023 = f"B{row_num}"
                        ws[cell_2023] = value_2023
                        section_totals[section]['2023'] += value_2023
                        print(f"   ✅ {template_field} ({section}) 2023: {cell_2023} = {value_2023:,.0f}")
                    
                    # Populate 2024 value (column C) 
                    if value_2024 is not None:
                        cell_2024 = f"C{row_num}"
                        ws[cell_2024] = value_2024
                        section_totals[section]['2024'] += value_2024
                        print(f"   ✅ {template_field} ({section}) 2024: {cell_2024} = {value_2024:,.0f}")
                    
                    populated_fields.append(f"{template_field} ({section})")
                else:
                    print(f"   ⚠️ No template mapping for: {template_key}")
            
            # Print section totals for verification
            print(f"\n📊 Section Totals Verification:")
            for section, totals in section_totals.items():
                if totals['2023'] > 0 or totals['2024'] > 0:
                    print(f"   {section}: 2023=${totals['2023']:,.0f}, 2024=${totals['2024']:,.0f}")
            
            # Save the populated template
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"populated_original_template_{timestamp}.xlsx"
            wb.save(output_file)
            
            print(f"\n✅ Original template populated successfully!")
            print(f"   Output file: {output_file}")
            print(f"   Fields populated: {len(populated_fields)}")
            print(f"   Fields: {', '.join(populated_fields)}")
            
            return output_file
            
        except Exception as e:
            print(f"❌ Error populating original template: {e}")
            return ""
    
    def analyze_coverage(self, mapped_items: Dict[str, MappedValue]):
        """Analyze mapping coverage"""
        print("\n📊 FINAL MAPPING ANALYSIS:")
        print("=" * 60)
        
        # Count by section
        section_counts = {}
        template_fields = set()
        
        for mapped_value in mapped_items.values():
            section = mapped_value.section
            section_counts[section] = section_counts.get(section, 0) + 1
            template_fields.add(mapped_value.template_field)
        
        print("Mapped items by section:")
        for section, count in section_counts.items():
            print(f"  {section}: {count} fields")
        
        print(f"\nTotal unique template fields mapped: {len(template_fields)}")
        print(f"Template fields: {sorted(template_fields)}")
        
        # Check coverage of key template requirements
        required_fields = {
            'Cash and equivalents', 'Accounts Receivable', 'Prepaid Expenses', 'Inventory', 'Investments',
            'Net PPE', 'Goodwill', 'Intangibles', 
            'Accounts Payable', 'Accrued Interest', 'Short term Borrowing', 'Current Portion of Long Term Debt',
            'Long Term Debt', 'Deferred income taxes',
            'Common Stock', 'Retained Earnings', 'Paid in Capital'
        }
        
        mapped_required = template_fields.intersection(required_fields)
        missing_required = required_fields - template_fields
        
        print(f"\nRequired field coverage: {len(mapped_required)}/{len(required_fields)} ({100*len(mapped_required)/len(required_fields):.1f}%)")
        print(f"✅ Mapped: {sorted(mapped_required)}")
        if missing_required:
            print(f"❌ Missing: {sorted(missing_required)}")
    
    def cleanup_template(self):
        """Clean up working template file"""
        try:
            if self.working_template_path.exists():
                os.remove(self.working_template_path)
                print(f"🧹 Cleaned up working template: {self.working_template_path}")
        except Exception as e:
            print(f"⚠️ Warning: Could not clean up template: {e}")

def main():
    mapper = FinalKGMapper()
    pdf_path = "../../input_pdfs/US_Venture_2024.pdf"
    
    try:
        # Extract and process
        mapped_items = mapper.extract_and_process(pdf_path)
        
        # Analyze results
        mapper.analyze_coverage(mapped_items)
        
        # Populate the ORIGINAL template structure
        original_template_file = mapper.populate_original_template(mapped_items)
        
        # Save JSON results
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_output_file = f"final_kg_us_venture_bs_{timestamp}.json"
        
        # Convert to JSON-serializable format
        json_output = {}
        for key, mapped_value in mapped_items.items():
            json_output[key] = {
                "template_field": mapped_value.template_field,
                "section": mapped_value.section,
                "original_description": mapped_value.original_description,
                "value_2023": mapped_value.value_2023,
                "value_2024": mapped_value.value_2024,
                "confidence": mapped_value.confidence,
                "mapping_method": mapped_value.mapping_method
            }
        
        with open(json_output_file, 'w') as f:
            json.dump(json_output, f, indent=2)
        
        print(f"\n💾 Results saved:")
        if original_template_file:
            # Copy to main output directory
            main_output = f"../../output_excel/{original_template_file}"
            shutil.copy2(original_template_file, main_output)
            print(f"   Original Template: {main_output}")
        print(f"   JSON: {json_output_file}")
        
    finally:
        # Clean up
        mapper.cleanup_template()

if __name__ == "__main__":
    main() 