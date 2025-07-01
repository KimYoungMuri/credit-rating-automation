#!/usr/bin/env python3
"""
Final Cash Flow Statement Mapper - Following the proven structure
- Uses the original financial_template.xlsx from templates/ directory  
- Enhanced pattern coverage and three-section approach (Operating, Investing, Financing)
- Net change in cash verification and stopping logic
- Uses shutil for template management
- Populates the actual IS.CF sheet structure (rows 21-44)
"""

import sys
from pathlib import Path
import json
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, List, Optional, Any, Set, Union, Tuple
import re
from collections import defaultdict
from final_extractor_adaptive import TextExtractor
from final_find_fs import FinancialStatementFinder
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import logging
import os
import requests

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class CFSMappedValue:
    """Represents a mapped cash flow statement value"""
    original_description: str
    template_field: str
    section: str  # operating, investing, financing
    value_2023: Optional[float] = None
    value_2024: Optional[float] = None
    confidence: float = 1.0
    mapping_method: str = ""
    source_data: dict = None

class FinalCFSMapper:
    """Final Cash Flow Statement mapper integrated with original financial template"""
    
    def __init__(self):
        self.extractor = TextExtractor()
        self.finder = FinancialStatementFinder()
        
        # Set up template paths
        self.original_template_path = Path("../../templates/financial_template.xlsx")
        self.working_template_path = Path("./working_financial_template.xlsx")
        
        # Cash Flow Statement template field mappings - based on IS.CF sheet (rows 21-44)
        self.template_mappings = {
            # OPERATING ACTIVITIES section (rows 23-26)
            # 'Net Income': Row 24 - AUTO-FILLED FROM IS, DON'T MAP
            'Changes in noncash items': ('B', 25),             # Row 25 (depreciation, etc.)
            'Changes in Assets and Liabilities': ('B', 26),    # Row 26 (working capital changes)
            # 'Net Cash from Operating': Row 23 - CALCULATED, DON'T MAP
            
            # INVESTING ACTIVITIES section (rows 29-32)
            'CapEx': ('B', 29),                                # Row 29 (negative - cash outflow)
            'Proceeds from asset sales': ('B', 30),           # Row 30 (positive - cash inflow)
            'Others_investing': ('B', 31),                     # Row 31 (other investing activities)
            'Net Cash from Investing': ('B', 32),              # Row 32 (calculated total)
            
            # FINANCING ACTIVITIES section (rows 35-40)
            'Issuance of Debt': ('B', 35),                    # Row 35 (positive - cash inflow)
            'Retirement of Debt': ('B', 36),                  # Row 36 (negative - cash outflow)
            'Issuance of Stock': ('B', 37),                   # Row 37 (positive - cash inflow)
            'Dividends Paid': ('B', 38),                      # Row 38 (negative - cash outflow)
            'Other_financing': ('B', 39),                     # Row 39 (other financing activities)
            'Net Cash from Financing': ('B', 40),             # Row 40 (calculated total)
            
            # SUMMARY section (rows 42-44)
            'Net change in Cash': ('B', 42),                  # Row 42 (calculated total)
            'Starting Cash': ('B', 43),                       # Row 43 (beginning cash balance)
            'Ending Cash': ('B', 44),                         # Row 44 (ending cash balance)
        }
        
        # Enhanced rule-based patterns for Cash Flow Statement items with 3-section approach
        self.cfs_rules = {
            # === NET CHANGE IN CASH DETECTION (STOPPING CONDITION) ===
            r'net\s+(?:increase|decrease|change)\s+in\s+cash(?:\s+and\s+(?:cash\s+)?equivalents?)?(?:\s|$)': ('_net_change_found', 'net_change_stop'),
            r'(?:increase|decrease)\s+in\s+cash(?:\s+and\s+(?:cash\s+)?equivalents?)?': ('_net_change_found', 'net_change_stop'),
            
            # === EXCLUDE CALCULATED TOTALS FIRST ===
            # DO NOT MAP these - they are calculated in the template
            r'net\s+cash\s+(?:provided\s+by|used\s+in)\s+operating\s+activities': ('_exclude_calculated_total', 'exclude'),
            r'net\s+cash\s+(?:provided\s+by|used\s+in)\s+investing\s+activities': ('_exclude_calculated_total', 'exclude'),
            r'net\s+cash\s+(?:provided\s+by|used\s+in)\s+financing\s+activities': ('_exclude_calculated_total', 'exclude'),
            r'total\s+(?:operating|investing|financing)\s+activities': ('_exclude_calculated_total', 'exclude'),
            r'cash\s+flows?\s+from\s+operating\s+activities': ('_exclude_section_header', 'exclude'),
            r'cash\s+flows?\s+from\s+investing\s+activities': ('_exclude_section_header', 'exclude'),
            r'cash\s+flows?\s+from\s+financing\s+activities': ('_exclude_section_header', 'exclude'),
            
            # === SECTION 1: OPERATING ACTIVITIES ===
            # Net Income (starting point)
            r'net\s+income(?:\s+attributable\s+to\s+common\s+shareholders)?(?:\s|$)': ('Net Income', 'operating'),
            r'net\s+earnings?(?:\s+attributable\s+to\s+common\s+shareholders)?': ('Net Income', 'operating'),
            
            # Non-cash items (positive adjustments)
            r'depreciation(?:\s+and\s+amortization)?(?:\s+expense)?': ('Changes in noncash items', 'operating'),
            r'amortization(?:\s+and\s+depreciation)?(?:\s+expense)?': ('Changes in noncash items', 'operating'),
            r'amortization\s+of\s+intangibles?': ('Changes in noncash items', 'operating'),
            r'stock[- ]based\s+compensation': ('Changes in noncash items', 'operating'),
            r'share[- ]based\s+compensation': ('Changes in noncash items', 'operating'),
            r'provision\s+for\s+(?:bad\s+)?debt': ('Changes in noncash items', 'operating'),
            r'impairment\s+(?:of|losses?)': ('Changes in noncash items', 'operating'),
            r'deferred\s+(?:income\s+)?taxes?': ('Changes in noncash items', 'operating'),
            r'(?:gain|loss)\s+on\s+(?:sale\s+of\s+)?(?:assets?|equipment)': ('Changes in noncash items', 'operating'),
            r'unrealized\s+(?:gain|loss)': ('Changes in noncash items', 'operating'),
            
            # Working capital changes (can be positive or negative)
            r'(?:increase|decrease)\s+in\s+accounts?\s+receivable': ('Changes in Assets and Liabilities', 'operating'),
            r'(?:increase|decrease)\s+in\s+inventor(?:y|ies)': ('Changes in Assets and Liabilities', 'operating'),
            r'(?:increase|decrease)\s+in\s+prepaid': ('Changes in Assets and Liabilities', 'operating'),
            r'(?:increase|decrease)\s+in\s+accounts?\s+payable': ('Changes in Assets and Liabilities', 'operating'),
            r'(?:increase|decrease)\s+in\s+accrued': ('Changes in Assets and Liabilities', 'operating'),
            r'changes?\s+in\s+operating\s+assets?\s+and\s+liabilities?': ('Changes in Assets and Liabilities', 'operating'),
            r'changes?\s+in\s+working\s+capital': ('Changes in Assets and Liabilities', 'operating'),
            r'other\s+operating\s+activities': ('Changes in Assets and Liabilities', 'operating'),
            
            # === SECTION 2: INVESTING ACTIVITIES ===
            # Capital expenditures (negative - cash outflow)
            r'capital\s+expenditures?': ('CapEx', 'investing'),
            r'capex': ('CapEx', 'investing'),
            r'purchases?\s+of\s+property(?:\s*,?\s*plant)?(?:\s+and\s+equipment)?': ('CapEx', 'investing'),
            r'acquisition\s+of\s+(?:property|equipment|assets?)': ('CapEx', 'investing'),
            r'investments?\s+in\s+property': ('CapEx', 'investing'),
            
            # Asset sales (positive - cash inflow)
            r'proceeds?\s+from\s+(?:sale\s+of\s+)?(?:assets?|equipment|property)': ('Proceeds from asset sales', 'investing'),
            r'disposal\s+of\s+(?:assets?|equipment)': ('Proceeds from asset sales', 'investing'),
            r'sales?\s+of\s+(?:assets?|equipment|property)': ('Proceeds from asset sales', 'investing'),
            
            # Other investing activities
            r'acquisition\s+of\s+(?:business|companies?)': ('Others_investing', 'investing'),
            r'purchases?\s+of\s+investments?': ('Others_investing', 'investing'),
            r'sales?\s+of\s+investments?': ('Others_investing', 'investing'),
            r'other\s+investing\s+activities': ('Others_investing', 'investing'),
            
            # === SECTION 3: FINANCING ACTIVITIES ===
            # Debt activities
            r'proceeds?\s+from\s+(?:issuance\s+of\s+)?(?:long[- ]term\s+)?debt': ('Issuance of Debt', 'financing'),
            r'borrowings?\s+(?:under|from)': ('Issuance of Debt', 'financing'),
            r'issuance\s+of\s+(?:long[- ]term\s+)?debt': ('Issuance of Debt', 'financing'),
            r'repayments?\s+of\s+(?:long[- ]term\s+)?debt': ('Retirement of Debt', 'financing'),
            r'payments?\s+on\s+(?:long[- ]term\s+)?debt': ('Retirement of Debt', 'financing'),
            r'retirement\s+of\s+debt': ('Retirement of Debt', 'financing'),
            
            # Equity activities
            r'proceeds?\s+from\s+(?:issuance\s+of\s+)?(?:common\s+)?stock': ('Issuance of Stock', 'financing'),
            r'issuance\s+of\s+(?:common\s+)?stock': ('Issuance of Stock', 'financing'),
            r'stock\s+issuance': ('Issuance of Stock', 'financing'),
            
            # Dividend payments (negative - cash outflow)
            r'dividends?\s+paid': ('Dividends Paid', 'financing'),
            r'cash\s+dividends?': ('Dividends Paid', 'financing'),
            r'distributions?\s+to\s+(?:shareholders?|owners?)': ('Dividends Paid', 'financing'),
            
            # Other financing activities
            r'repurchases?\s+of\s+(?:common\s+)?stock': ('Other_financing', 'financing'),
            r'treasury\s+stock\s+(?:purchases?|repurchases?)': ('Other_financing', 'financing'),
            r'payments?\s+of\s+financing\s+costs?': ('Other_financing', 'financing'),
            r'other\s+financing\s+activities': ('Other_financing', 'financing'),
        }
        
        # Section mapping for three-section consolidation
        self.section_consolidation_mapping = {
            'operating': 'Changes in Assets and Liabilities',  # Operating activities consolidate here
            'investing': 'Others_investing',                   # Investing activities consolidate here  
            'financing': 'Other_financing'                     # Financing activities consolidate here
        }
        
        # Three-section calculation logic
        self.section_calculations = {
            'operating': {
                'starting_item': 'Net Income',
                'positive_items': ['Net Income', 'Changes in noncash items'],  # Add-backs to net income
                'negative_items': [],  # Typically working capital changes can be positive or negative
                'result_field': 'Net Cash from Operating',
                'description': 'Net Income plus non-cash adjustments plus working capital changes'
            },
            'investing': {
                'starting_item': None,
                'positive_items': ['Proceeds from asset sales'],  # Cash inflows
                'negative_items': ['CapEx'],  # Cash outflows (capex is negative)
                'result_field': 'Net Cash from Investing',
                'description': 'Cash inflows minus cash outflows from investing activities'
            },
            'financing': {
                'starting_item': None,
                'positive_items': ['Issuance of Debt', 'Issuance of Stock'],  # Cash inflows
                'negative_items': ['Retirement of Debt', 'Dividends Paid'],  # Cash outflows
                'result_field': 'Net Cash from Financing',
                'description': 'Cash inflows minus cash outflows from financing activities'
            }
        }
        
        # Track extracted values for verification
        self.extracted_net_change_2023 = None
        self.extracted_net_change_2024 = None
        self.extracted_starting_cash_2023 = None
        self.extracted_starting_cash_2024 = None
        self.extracted_ending_cash_2023 = None
        self.extracted_ending_cash_2024 = None
    
    def setup_template(self) -> bool:
        """Copy original template to working directory"""
        try:
            print(f"📋 Setting up CFS template...")
            
            if not self.original_template_path.exists():
                print(f"❌ Original template not found: {self.original_template_path}")
                return False
            
            shutil.copy2(self.original_template_path, self.working_template_path)
            
            if self.working_template_path.exists():
                print(f"✅ Template copied successfully")
                return True
            else:
                print(f"❌ Failed to copy template")
                return False
                
        except Exception as e:
            print(f"❌ Error setting up template: {e}")
            return False
    
    def is_total_or_net_row(self, description: str) -> bool:
        """Check if description is a total or calculated row"""
        desc_lower = description.lower().strip()
        
        # === CRITICAL: EXCLUDE NET INCOME LINES (COMES FROM IS) ===
        # Net Income should come from Income Statement, not Cash Flow Statement
        if re.search(r'(?:operating\s+)?activities:\s*net\s+income', desc_lower):
            print(f"🚫 NET INCOME LINE EXCLUDED (from IS): {description}")
            return True  # Filter it out - use IS value instead
        
        if re.search(r'\bnet\s+income(?:\s+attributable)?(?:\s+to\s+common)?(?:\s+shareholders)?(?:\s|$)', desc_lower):
            print(f"🚫 NET INCOME LINE EXCLUDED (from IS): {description}")
            return True  # Filter it out - use IS value instead
        
        # === CRITICAL: CHECK FOR NET CHANGE IN CASH FIRST ===
        # This is our stopping condition
        if re.search(r'net\s+(?:increase|decrease|change)\s+in\s+cash(?:\s+and\s+(?:cash\s+)?equivalents?)?(?:\s|$)', desc_lower):
            print(f"🛑 NET CHANGE IN CASH FOUND: {description}")
            return False  # Don't filter it out, we need to extract the value first
        
        # Check for starting/ending cash balances (don't filter these)
        if re.search(r'(?:beginning|starting)\s+cash(?:\s+and\s+(?:cash\s+)?equivalents?)?', desc_lower):
            print(f"💰 STARTING CASH FOUND: {description}")
            return False
        
        if re.search(r'(?:ending|final)\s+cash(?:\s+and\s+(?:cash\s+)?equivalents?)?', desc_lower):
            print(f"💰 ENDING CASH FOUND: {description}")
            return False
        
        # Debug specific problematic lines
        if "net cash" in desc_lower or "total" in desc_lower:
            print(f"🔍 DEBUG is_total_or_net_row: Testing '{description}'")
            print(f"   desc_lower: '{desc_lower}'")
        
        # Specific calculated/total rows to filter out (these are calculated in template)
        calculated_rows = [
            'net cash provided by operating activities',
            'net cash used in operating activities',
            'net cash from operating activities',
            'net cash provided by investing activities',
            'net cash used in investing activities',
            'net cash from investing activities',
            'net cash provided by financing activities',
            'net cash used in financing activities',
            'net cash from financing activities',
            'total operating activities',
            'total investing activities',
            'total financing activities'
        ]
        
        for calc_row in calculated_rows:
            if calc_row in desc_lower:
                if "net cash" in desc_lower:
                    print(f"   ✅ MATCHED: '{calc_row}' - FILTERING OUT!")
                return True
        
        # Filter out general totals and subtotals (but keep specific line items)
        total_patterns = [
            r'^total(\s|$)',
            r'(\s|^)sum(\s|$)',
            r'(\s|^)subtotal(\s|$)',
            r'(\s|^)aggregate(\s|$)',
            r'(\s|^)grand total(\s|$)',
        ]
        
        # Check if it's a total pattern but NOT a specific line item we want
        for pat in total_patterns:
            if re.search(pat, desc_lower):
                if "total" in desc_lower:
                    print(f"   ✅ REGEX MATCHED: '{pat}' - FILTERING OUT!")
                return True
        
        # Filter out header/formatting rows
        header_patterns = [
            r'^\s*\d{4}\s+\d{4}\s*$',  # Year headers
            r'^\s*cash\s+flow\s+statement\s*$',
            r'^\s*statement\s+of.*cash.*flow\s*$',
            r'^\s*(?:consolidated\s+)?statements?\s+of\s+cash\s+flows?\s*$',
            r'^\s*operating\s+activities\s*$',
            r'^\s*investing\s+activities\s*$',
            r'^\s*financing\s+activities\s*$'
        ]
        
        for pat in header_patterns:
            if re.search(pat, desc_lower):
                return True
                
        return False
    
    def apply_enhanced_mapping(self, description: str) -> Tuple[Optional[str], Optional[str], float]:
        """Apply enhanced rule-based mapping for cash flow statement items"""
        desc_lower = description.lower().strip()
        
        # Debug specific problematic items
        if "net change" in desc_lower or "starting cash" in desc_lower or "ending cash" in desc_lower:
            print(f"🔍 DEBUG apply_enhanced_mapping: Testing '{description}'")
            print(f"   desc_lower: '{desc_lower}'")
        
        for pattern, (template_field, section) in self.cfs_rules.items():
            if re.search(pattern, desc_lower):
                if "net change" in desc_lower or "starting cash" in desc_lower or "ending cash" in desc_lower:
                    print(f"   ✅ PATTERN MATCHED: '{pattern}' → {template_field} | {section}")
                
                # Skip calculated fields and excluded totals - we don't want to map these
                if template_field.startswith('_calculated_') or template_field.startswith('_exclude_'):
                    return None, None, 0.0
                return template_field, section, 0.9
        
        if "net change" in desc_lower or "starting cash" in desc_lower or "ending cash" in desc_lower:
            print(f"   ❌ NO PATTERN MATCHED - falling back to other tiers")
        
        return None, None, 0.0 

    def ask_ollama_for_classification(self, description: str) -> Tuple[Optional[str], Optional[str]]:
        """Ask Ollama LLM to classify cash flow statement line items"""
        try:
            # Import ollama only when needed to avoid dependency issues
            import ollama
            import time
            
            # Extract the key part of the description for faster processing
            key_desc = description[:50].lower()
            
            # Special system prompt for cash flow statement classification
            system_prompt = """You are a financial statement analyzer. Classify cash flow items into:
FIELDS: Net Income, Changes in noncash items, Changes in Assets and Liabilities, CapEx, Proceeds from asset sales, Issuance of Debt, Retirement of Debt, Dividends Paid
SECTIONS: operating, investing, financing

Respond ONLY with: FIELD|SECTION
Example: "depreciation" → "Changes in noncash items|operating"
"""
            
            user_prompt = f"Classify this cash flow item: '{key_desc}'"
            
            # Call Ollama API with phi3:mini
            try:
                result = ollama.chat(
                    model='phi3:mini',
                    messages=[
                        {'role': 'system', 'content': system_prompt},
                        {'role': 'user', 'content': user_prompt}
                    ],
                    options={'num_ctx': 512, 'temperature': 0.1}
                )
                
                ollama_response = result.get('response', '').strip()
                
                # Parse the response
                if '|' in ollama_response:
                    field, section = ollama_response.split('|', 1)
                    field = field.strip()
                    section = section.strip()
                    
                    # Validate the response
                    valid_fields = ['Net Income', 'Changes in noncash items', 'Changes in Assets and Liabilities', 
                                  'CapEx', 'Proceeds from asset sales', 'Issuance of Debt', 'Retirement of Debt', 'Dividends Paid']
                    valid_sections = ['operating', 'investing', 'financing']
                    
                    if field in valid_fields and section in valid_sections:
                        return field, section
                        
            except Exception as e:
                print(f"   ⚠️ LLM classification failed: {e}")
                
        except ImportError:
            # Ollama not available, skip LLM classification
            pass
        except Exception as e:
            print(f"   ⚠️ LLM classification failed: {e}")
        
        return None, None
    
    def apply_multi_tier_fallback(self, description: str) -> Tuple[Optional[str], Optional[str], float, str]:
        """Apply 5-tier fallback system for cash flow statement classification"""
        
        # TIER 1: Enhanced regex patterns (already tried above, so skip)
        
        # TIER 2: Fuzzy matching against template fields (70% confidence)
        template_field, section, confidence = self.apply_fuzzy_matching(description)
        if template_field and confidence >= 0.7:
            print(f"   🔄 Fuzzy match: {description[:30]}... → {template_field}")
            return template_field, section, confidence, 'fuzzy_matching'
        
        # TIER 3: Keyword-based analysis (60% confidence)
        template_field, section = self.apply_keyword_analysis(description)
        if template_field and section:
            print(f"   🔍 Keyword analysis: {description[:30]}... → {template_field}")
            return template_field, section, 0.6, 'keyword_analysis'
        
        # TIER 4: LLM fallback with phi3:mini (70% confidence)
        llm_field, llm_section = self.ask_ollama_for_classification(description)
        if llm_field and llm_section:
            print(f"   🤖 Ollama inference: {description[:30]}... → {llm_field}")
            return llm_field, llm_section, 0.7, 'llm_fallback'
        
        # TIER 5: Smart cash flow fallback (50% confidence)
        smart_field, smart_section = self.smart_cash_flow_fallback(description)
        if smart_field and smart_section:
            print(f"   🧠 Smart fallback: {description[:30]}... → {smart_field}")
            return smart_field, smart_section, 0.5, 'smart_fallback'
        
        return None, None, 0.0, 'no_match'
    
    def apply_fuzzy_matching(self, description: str) -> Tuple[Optional[str], Optional[str], float]:
        """Apply fuzzy string matching against known template fields"""
        from difflib import SequenceMatcher
        
        desc_lower = description.lower().strip()
        
        # Template field patterns for fuzzy matching
        template_patterns = {
            'Net Income': ['net income', 'net earnings', 'income'],
            'Changes in noncash items': ['depreciation', 'amortization', 'stock compensation', 'impairment'],
            'Changes in Assets and Liabilities': ['working capital', 'receivables', 'payables', 'inventory'],
            'CapEx': ['capital expenditures', 'capex', 'property plant equipment', 'ppe'],
            'Proceeds from asset sales': ['asset sales', 'disposal', 'proceeds from sales'],
            'Issuance of Debt': ['debt issuance', 'borrowings', 'proceeds from debt'],
            'Retirement of Debt': ['debt repayment', 'debt retirement', 'payments on debt'],
            'Issuance of Stock': ['stock issuance', 'equity issuance', 'proceeds from stock'],
            'Dividends Paid': ['dividends paid', 'dividend payments', 'distributions']
        }
        
        best_match = None
        best_confidence = 0.0
        
        for template_field, patterns in template_patterns.items():
            for pattern in patterns:
                similarity = SequenceMatcher(None, desc_lower, pattern).ratio()
                if similarity > best_confidence and similarity >= 0.6:  # 60% similarity threshold
                    best_match = template_field
                    best_confidence = similarity
        
        if best_match:
            # Determine section based on template field
            section_mapping = {
                'Net Income': 'operating',
                'Changes in noncash items': 'operating',
                'Changes in Assets and Liabilities': 'operating',
                'CapEx': 'investing',
                'Proceeds from asset sales': 'investing',
                'Issuance of Debt': 'financing',
                'Retirement of Debt': 'financing',
                'Issuance of Stock': 'financing',
                'Dividends Paid': 'financing'
            }
            section = section_mapping.get(best_match, 'operating')
            return best_match, section, best_confidence
        
        return None, None, 0.0
    
    def apply_keyword_analysis(self, description: str) -> Tuple[Optional[str], Optional[str]]:
        """Apply keyword-based classification for cash flow statement items"""
        desc_lower = description.lower().strip()
        
        # Cash flow specific keyword classifications
        keyword_mappings = {
            # Operating activities
            ('Net Income', 'operating'): [
                'net income', 'net earnings', 'earnings'
            ],
            
            ('Changes in noncash items', 'operating'): [
                'depreciation', 'amortization', 'stock', 'compensation',
                'impairment', 'provision', 'deferred', 'unrealized'
            ],
            
            ('Changes in Assets and Liabilities', 'operating'): [
                'receivable', 'inventory', 'payable', 'accrued',
                'working', 'capital', 'prepaid', 'operating'
            ],
            
            # Investing activities
            ('CapEx', 'investing'): [
                'capital', 'expenditure', 'capex', 'property',
                'plant', 'equipment', 'acquisition', 'purchase'
            ],
            
            ('Proceeds from asset sales', 'investing'): [
                'proceeds', 'disposal', 'sale', 'assets'
            ],
            
            # Financing activities
            ('Issuance of Debt', 'financing'): [
                'issuance', 'proceeds', 'borrowing', 'debt'
            ],
            
            ('Retirement of Debt', 'financing'): [
                'repayment', 'retirement', 'payment', 'debt'
            ],
            
            ('Issuance of Stock', 'financing'): [
                'stock', 'equity', 'issuance', 'shares'
            ],
            
            ('Dividends Paid', 'financing'): [
                'dividend', 'distribution', 'payment'
            ]
        }
        
        # Score each mapping based on keyword matches
        best_score = 0
        best_mapping = None
        
        for (template_field, section), keywords in keyword_mappings.items():
            score = sum(1 for keyword in keywords if keyword in desc_lower)
            if score > best_score and score >= 2:  # Require at least 2 keyword matches
                best_score = score
                best_mapping = (template_field, section)
        
        if best_mapping:
            return best_mapping[0], best_mapping[1]
        
        return None, None
    
    def smart_cash_flow_fallback(self, description: str) -> Tuple[Optional[str], Optional[str]]:
        """Smart fallback classification for cash flow statement items"""
        desc_lower = description.lower().strip()
        
        # Pattern-based smart classification
        if any(word in desc_lower for word in ['increase', 'decrease']) and 'cash' in desc_lower:
            return 'Changes in Assets and Liabilities', 'operating'
        
        if any(word in desc_lower for word in ['purchase', 'acquisition', 'invest']) and not 'stock' in desc_lower:
            return 'CapEx', 'investing'
        
        if any(word in desc_lower for word in ['borrow', 'loan', 'credit']) and 'proceeds' in desc_lower:
            return 'Issuance of Debt', 'financing'
        
        if any(word in desc_lower for word in ['repay', 'retire']) and 'debt' in desc_lower:
            return 'Retirement of Debt', 'financing'
        
        if 'stock' in desc_lower and any(word in desc_lower for word in ['issue', 'proceeds']):
            return 'Issuance of Stock', 'financing'
        
        if any(word in desc_lower for word in ['dividend', 'distribution']):
            return 'Dividends Paid', 'financing'
        
        # Default to operating activities for unclear items
        return 'Changes in Assets and Liabilities', 'operating'
    
    def consolidate_multi_mappings_improved(self, mapped_items: Dict[str, CFSMappedValue]) -> Dict[str, CFSMappedValue]:
        """Three-section consolidation approach: Operating, Investing, and Financing activities"""
        consolidated = {}
        
        print("🔄 THREE-SECTION CFS CONSOLIDATION:")
        print("-" * 50)
        
        # Separate items by section
        operating_items = []
        investing_items = []
        financing_items = []
        
        for key, mapped_value in mapped_items.items():
            section = mapped_value.section
            
            if section == "operating":
                operating_items.append((key, mapped_value))
            elif section == "investing":
                investing_items.append((key, mapped_value))
            elif section == "financing":
                financing_items.append((key, mapped_value))
        
        # === SECTION 1: OPERATING ACTIVITIES ===
        if operating_items:
            print("📊 SECTION 1 - Operating Activities:")
            
            # Group by template field
            field_groups = {}
            for key, mapped_value in operating_items:
                field = mapped_value.template_field
                if field not in field_groups:
                    field_groups[field] = []
                field_groups[field].append((key, mapped_value))
            
            # Consolidate each field group
            for field, items in field_groups.items():
                total_2023 = sum(mv.value_2023 for k, mv in items if mv.value_2023)
                total_2024 = sum(mv.value_2024 for k, mv in items if mv.value_2024)
                
                consolidated[f"{field.lower().replace(' ', '_')}_consolidated"] = CFSMappedValue(
                    original_description=f"Total {field} (All Operating)",
                    template_field=field,
                    section="operating",
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.95,
                    mapping_method="section1_operating_consolidation",
                    source_data={"item_count": len(items), "section": "operating"}
                )
                print(f"   ✅ {field}: {len(items)} items → ${total_2024:,.0f} (2024)")
        
        # === SECTION 2: INVESTING ACTIVITIES ===
        if investing_items:
            print("\n📊 SECTION 2 - Investing Activities:")
            
            # Group by template field
            field_groups = {}
            for key, mapped_value in investing_items:
                field = mapped_value.template_field
                if field not in field_groups:
                    field_groups[field] = []
                field_groups[field].append((key, mapped_value))
            
            # Consolidate each field group
            for field, items in field_groups.items():
                total_2023 = sum(mv.value_2023 for k, mv in items if mv.value_2023)
                total_2024 = sum(mv.value_2024 for k, mv in items if mv.value_2024)
                
                consolidated[f"{field.lower().replace(' ', '_')}_consolidated"] = CFSMappedValue(
                    original_description=f"Total {field} (All Investing)",
                    template_field=field,
                    section="investing",
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.9,
                    mapping_method="section2_investing_consolidation",
                    source_data={"item_count": len(items), "section": "investing"}
                )
                print(f"   ✅ {field}: {len(items)} items → ${total_2024:,.0f} (2024)")
        
        # === SECTION 3: FINANCING ACTIVITIES ===
        if financing_items:
            print("\n📊 SECTION 3 - Financing Activities:")
            
            # Group by template field
            field_groups = {}
            for key, mapped_value in financing_items:
                field = mapped_value.template_field
                if field not in field_groups:
                    field_groups[field] = []
                field_groups[field].append((key, mapped_value))
            
            # Consolidate each field group
            for field, items in field_groups.items():
                total_2023 = sum(mv.value_2023 for k, mv in items if mv.value_2023)
                total_2024 = sum(mv.value_2024 for k, mv in items if mv.value_2024)
                
                consolidated[f"{field.lower().replace(' ', '_')}_consolidated"] = CFSMappedValue(
                    original_description=f"Total {field} (All Financing)",
                    template_field=field,
                    section="financing",
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.9,
                    mapping_method="section3_financing_consolidation",
                    source_data={"item_count": len(items), "section": "financing"}
                )
                print(f"   ✅ {field}: {len(items)} items → ${total_2024:,.0f} (2024)")
        
        print(f"\n📊 Three-Section CFS Consolidation Complete: {len(consolidated)} consolidated items")
        return consolidated
    
    def verify_net_change_in_cash(self, mapped_items: Dict[str, CFSMappedValue]):
        """Verify extracted Net Change in Cash against calculated values using three-section approach"""
        print(f"\n🔍 THREE-SECTION NET CHANGE IN CASH VERIFICATION:")
        print("=" * 70)
        
        if not self.extracted_net_change_2023 and not self.extracted_net_change_2024:
            print("❌ No extracted Net Change in Cash values found")
            return
        
        # Three-section calculation
        operating_2023 = operating_2024 = 0
        investing_2023 = investing_2024 = 0 
        financing_2023 = financing_2024 = 0
        
        print("📊 SECTION-BY-SECTION CALCULATION:")
        print("-" * 40)
        
        # Section 1: Operating Activities
        print("💼 SECTION 1 - Operating Cash Flow:")
        for key, mapped_value in mapped_items.items():
            if mapped_value.section == "operating":
                v23 = mapped_value.value_2023 or 0
                v24 = mapped_value.value_2024 or 0
                
                operating_2023 += v23
                operating_2024 += v24
                print(f"   + {mapped_value.template_field}: 2023=${v23:,.0f}, 2024=${v24:,.0f}")
        
        print(f"   = Operating Cash Flow: 2023=${operating_2023:,.0f}, 2024=${operating_2024:,.0f}")
        
        # Section 2: Investing Activities
        print("\n🏭 SECTION 2 - Investing Cash Flow:")
        for key, mapped_value in mapped_items.items():
            if mapped_value.section == "investing":
                v23 = mapped_value.value_2023 or 0
                v24 = mapped_value.value_2024 or 0
                
                investing_2023 += v23
                investing_2024 += v24
                print(f"   + {mapped_value.template_field}: 2023=${v23:,.0f}, 2024=${v24:,.0f}")
        
        print(f"   = Investing Cash Flow: 2023=${investing_2023:,.0f}, 2024=${investing_2024:,.0f}")
        
        # Section 3: Financing Activities
        print("\n💰 SECTION 3 - Financing Cash Flow:")
        for key, mapped_value in mapped_items.items():
            if mapped_value.section == "financing":
                v23 = mapped_value.value_2023 or 0
                v24 = mapped_value.value_2024 or 0
                
                financing_2023 += v23
                financing_2024 += v24
                print(f"   + {mapped_value.template_field}: 2023=${v23:,.0f}, 2024=${v24:,.0f}")
        
        print(f"   = Financing Cash Flow: 2023=${financing_2023:,.0f}, 2024=${financing_2024:,.0f}")
        
        # Calculate total net change in cash
        calculated_2023 = operating_2023 + investing_2023 + financing_2023
        calculated_2024 = operating_2024 + investing_2024 + financing_2024
        
        print(f"\n🧮 FINAL VERIFICATION:")
        print("=" * 40)
        print(f"Calculated Net Change 2023: ${calculated_2023:,.0f}")
        print(f"Calculated Net Change 2024: ${calculated_2024:,.0f}")
        print(f"Extracted Net Change 2023:  ${self.extracted_net_change_2023 or 0:,.0f}")
        print(f"Extracted Net Change 2024:  ${self.extracted_net_change_2024 or 0:,.0f}")
        
        print(f"\n📊 INTERMEDIATE TOTALS:")
        print(f"Operating Cash Flow:  2023=${operating_2023:,.0f}, 2024=${operating_2024:,.0f}")
        print(f"Investing Cash Flow:  2023=${investing_2023:,.0f}, 2024=${investing_2024:,.0f}")
        print(f"Financing Cash Flow:  2023=${financing_2023:,.0f}, 2024=${financing_2024:,.0f}")
        
        # Verify accuracy
        tolerance = 0.05  # 5% tolerance
        
        if self.extracted_net_change_2023:
            diff_2023 = abs(calculated_2023 - self.extracted_net_change_2023)
            pct_diff_2023 = diff_2023 / abs(self.extracted_net_change_2023) if self.extracted_net_change_2023 != 0 else 0
            
            if pct_diff_2023 <= tolerance:
                print(f"✅ 2023 Match: {pct_diff_2023:.1%} difference (within {tolerance:.1%} tolerance)")
            else:
                print(f"❌ 2023 Mismatch: {pct_diff_2023:.1%} difference (exceeds {tolerance:.1%} tolerance)")
                print(f"   Difference: ${diff_2023:,.0f}")
        
        if self.extracted_net_change_2024:
            diff_2024 = abs(calculated_2024 - self.extracted_net_change_2024)
            pct_diff_2024 = diff_2024 / abs(self.extracted_net_change_2024) if self.extracted_net_change_2024 != 0 else 0
            
            if pct_diff_2024 <= tolerance:
                print(f"✅ 2024 Match: {pct_diff_2024:.1%} difference (within {tolerance:.1%} tolerance)")
            else:
                print(f"❌ 2024 Mismatch: {pct_diff_2024:.1%} difference (exceeds {tolerance:.1%} tolerance)")
    
    def extract_and_process(self, pdf_path: str) -> Dict[str, CFSMappedValue]:
        """Main processing function for cash flow statement with three-section approach"""
        print("ENHANCED CASH FLOW STATEMENT KNOWLEDGE GRAPH MAPPER")
        print("=" * 70)
        print("NEW FEATURES:")
        print("1. Three-section processing: Operating → Investing → Financing")
        print("2. Stops after 'Net cash from financing activities'")
        print("3. Proper section-based classification")
        print("4. Enhanced pattern matching for US_Venture CFS")
        print()
        
        # Extract cash flow statement using the existing extractor
        try:
            # Use simple confirmed pages (like other mappers)
            confirmed_pages = {
                'cash_flow': [11, 12]  # Pages 11-12 contain the actual cash flow statement
            }
            
            excel_path, extracted_data = self.extractor.extract_text(
                pdf_path, 
                process_numbers=True,
                statement_pages=confirmed_pages
            )
            
            if 'cash_flow' not in extracted_data:
                print("❌ No cash flow statement found in PDF")
                return {}
            
            # Get the cash flow data - it's directly a list, not nested by year
            year_data = extracted_data['cash_flow']
            
            if not isinstance(year_data, list):
                print(f"❌ Invalid cash flow data format: {type(year_data)}")
                return {}
            
            print(f"✅ Extracted {len(year_data)} cash flow items")
            
        except Exception as e:
            print(f"❌ Error extracting cash flow data: {e}")
            return {}
        
        # Filter out headers and setup section tracking
        filtered_items = []
        total_filtered = 0
        current_section = None
        financing_activities_ended = False
        
        for item in year_data:
            description = item.get('description', '').strip()
            if not description:
                continue
            
            desc_lower = description.lower()
            
            # DEBUG: Track the main depreciation line
            if 'depreciation and amortization' in desc_lower:
                print(f"🔍 TRACKING DEPRECIATION LINE: {description}")
                print(f"   Numbers: {item.get('numbers', {})}")
                
            # Skip obvious header/formatting rows
            if (not any(item.get('numbers', {}).values()) and 
                ('venture' in desc_lower or 'consolidated' in desc_lower or 
                 'for the years' in desc_lower or 'amounts in thousands' in desc_lower or
                 desc_lower.strip() in ['2024 2023', '- 9 -', '- 10 -', '(continued)', '(concluded)'])):
                if 'depreciation and amortization' in desc_lower:
                    print(f"   🚫 FILTERED: Header/formatting row")
                total_filtered += 1
                continue
            
            # CRITICAL: Stop processing after "Net cash from financing activities"
            if re.search(r'net\s+cash\s+(?:used|provided).*financing\s+activities', desc_lower):
                if 'depreciation and amortization' in desc_lower:
                    print(f"   🚫 FILTERED: Financing section end")
                print(f"🛑 FINANCING SECTION END: {description}")
                print(f"   STOPPING - Items after this are cash balances and should be ignored")
                financing_activities_ended = True
                break
            
            # Track sections - make patterns more specific to avoid matching line items
            if re.search(r'^(operating|investing|financing)\s+activities:', desc_lower):
                if 'operating' in desc_lower:
                    current_section = 'operating'
                    print(f"📍 SECTION: Operating Activities")
                elif 'investing' in desc_lower:
                    current_section = 'investing'
                    print(f"📍 SECTION: Investing Activities")
                elif 'financing' in desc_lower:
                    current_section = 'financing'
                    print(f"📍 SECTION: Financing Activities")
                    
                if 'depreciation and amortization' in desc_lower:
                    print(f"   🚫 FILTERED: Section header row")
                total_filtered += 1
                continue
            
            # Filter out calculated totals within sections
            if re.search(r'net\s+cash\s+(?:provided|used).*(?:operating|investing)\s+activities', desc_lower):
                if 'depreciation and amortization' in desc_lower:
                    print(f"   🚫 FILTERED: Calculated total")
                print(f"🚫 CALCULATED TOTAL: {description}")
                total_filtered += 1
                continue
            
            # Add section context to item for processing
            item['_section_context'] = current_section
            filtered_items.append(item)
            
            # DEBUG: Track if depreciation line gets added
            if 'depreciation and amortization' in desc_lower:
                print(f"   ✅ ADDED TO FILTERED ITEMS: {description}")
                print(f"   Section context: {current_section}")
        
        print(f"📊 After filtering: {len(filtered_items)} items to map, {total_filtered} headers/totals filtered")
        print(f"🛑 Financing activities end detected: {financing_activities_ended}")
        
        # Enhanced pattern matching for actual US_Venture items (OCR-resistant)
        enhanced_patterns = {
            # SPECIFIC DEPRECIATION PATTERNS FIRST (before generic operating patterns)
            r'depreci?\s*a?\s*tion\s+a?\s*nd\s+amortization': ('Changes in noncash items', 'operating'),
            r'by\s+operating\s+activities:\s+depreciation\s+and\s+amortization': ('Changes in noncash items', 'operating'),
            
            # OPERATING ACTIVITIES - exact patterns from actual CFS (with OCR tolerance)
            r'operating\s+activities:\s*net\s+income': ('Net Income', 'operating'),
            r'adjustments\s+to\s+reconcile\s+net\s+income': ('Net Income', 'operating'),
            r'(?:by\s+)?operati?\s*ng\s+activi?\s*ties:\s*depreci?\s*a?\s*nd\s+amortization': ('Changes in noncash items', 'operating'),
            r'impai?\s*rment\s+l?\s*os?\s*ses?\s+on\s+long[- ]li?\s*ved\s+a?\s*nd\s+intangible\s+a?\s*ss?ets?': ('Changes in noncash items', 'operating'),
            r'amortization\s+of\s+debt\s+fi?\s*nancing\s+costs?': ('Changes in noncash items', 'operating'),
            r'ga?\s*i?\s*n\s+on\s+s?\s*ale\s+of\s+property\s+and\s+equipment': ('Changes in noncash items', 'operating'),
            r'bad\s+debt\s+and\s+investment\s+write\s+off': ('Changes in noncash items', 'operating'),
            r'change\s+in\s+fi?\s*nance\s+lease\s+l?\s*iabil?\s*ity': ('Changes in Assets and Liabilities', 'operating'),
            r'change\s+in\s+ri?\s*ght\s+of\s+use\s+as?\s*set': ('Changes in Assets and Liabilities', 'operating'),
            r'equity\s+in\s+earnings\s+of\s+unconsol?\s*idated\s+entities': ('Changes in noncash items', 'operating'),
            r'\(income\)\s+expens?\s*e\s+due\s+to\s+uti?\s*lizing\s+lifo\s+inventory\s+method': ('Changes in Assets and Liabilities', 'operating'),
            r'changes?\s+in\s+operating\s+a?\s*ss?\s*ets?\s+a?\s*nd\s+liabilities': ('Changes in Assets and Liabilities', 'operating'),
            r'margin\s+deposits': ('Changes in Assets and Liabilities', 'operating'),
            r'deri?\s*vative\s+a?\s*ss?ets?\s+a?\s*nd\s+l?\s*iabilities[—-]net': ('Changes in Assets and Liabilities', 'operating'),
            r'accounts\s+recei?\s*vabl?\s*e[—-]net': ('Changes in Assets and Liabilities', 'operating'),
            r'inventori?\s*es?[—-]net': ('Changes in Assets and Liabilities', 'operating'),
            r'other\s+as?\s*sets': ('Changes in Assets and Liabilities', 'operating'),
            r'accounts\s+pa?\s*yabl?\s*e': ('Changes in Assets and Liabilities', 'operating'),
            r'accrued\s+l?\s*iabil?\s*ities': ('Changes in Assets and Liabilities', 'operating'),
            r'other\s+lia?\s*bi?\s*liti?\s*es': ('Changes in Assets and Liabilities', 'operating'),
            
            # INVESTING ACTIVITIES - exact patterns (with OCR tolerance)
            r'investing\s+activities:\s*purchase\s+of\s+long[- ]lived\s+as?\s*sets?': ('CapEx', 'investing'),
            r'purchase\s+of\s+long[- ]lived\s+as?\s*sets?': ('CapEx', 'investing'),
            r'acquisiti?\s*ons?[—-]net\s+of\s+cas?\s*h\s+received': ('CapEx', 'investing'),
            r'proceeds\s+on\s+sale\s+of\s+property\s+and\s+equi?\s*pment\s+a?\s*nd\s+other\s+a?\s*ss?ets?': ('Proceeds from asset sales', 'investing'),
            r'proceeds\s+on\s+sale\s+of\s+business\s+net\s+of\s+ca?\s*sh': ('Proceeds from asset sales', 'investing'),
            r'purchase\s+of\s+equity\s+inves?\s*tment': ('CapEx', 'investing'),
            r'di?\s*stributions\s+from\s+equity\s+investments': ('Proceeds from asset sales', 'investing'),
            r'deconsol?\s*idati?\s*on\s+of\s+equity\s+investment': ('CapEx', 'investing'),
            r'purchase\s+of\s+other\s+investments': ('CapEx', 'investing'),
            r'repayments\s+of\s+notes\s+recei?\s*vabl?\s*e': ('Proceeds from asset sales', 'investing'),
            r'iss?\s*uance\s+of\s+notes\s+receivable': ('CapEx', 'investing'),
            
            # FINANCING ACTIVITIES - exact patterns (with OCR tolerance)
            r'financing\s+activities:\s*net\s+\(decrease\)\s+i?\s*ncrease\s+i?\s*n\s+li?\s*ne\s+of\s+credi?\s*t\s+borrowi?\s*ngs': ('Issuance of Debt', 'financing'),
            r'net\s+\(decrease\)\s+i?\s*ncrease\s+i?\s*n\s+li?\s*ne\s+of\s+credi?\s*t\s+borrowi?\s*ngs': ('Issuance of Debt', 'financing'),
            r'pa?\s*yments\s+of\s+fi?\s*nance\s+leas?\s*e\s+li?\s*abili?\s*ty': ('Retirement of Debt', 'financing'),
            r'proceeds\s+from\s+l?\s*ong[- ]term\s+debt': ('Issuance of Debt', 'financing'),
            r'pa?\s*yments\s+of\s+long[- ]term\s+debt': ('Retirement of Debt', 'financing'),
            r'pa?\s*yments\s+of\s+debt\s+fina?\s*ncing\s+costs?': ('Retirement of Debt', 'financing'),
            r'pa?\s*yments\s+of\s+contingent\s+considera?\s*tion': ('Retirement of Debt', 'financing'),
            r'contri?\s*butions\s+from\s+non[- ]control?\s*li?\s*ng\s+interests': ('Issuance of Stock', 'financing'),
            r'di?\s*stributions\s+to\s+s?\s*hareholders': ('Dividends Paid', 'financing'),
        }
        
        # Map items using enhanced patterns and section context
        mapped_items = {}
        
        print(f"\n🔄 Three-Section Processing with Enhanced Patterns:")
        print("-" * 60)
        
        for item in filtered_items:
            description = item.get('description', '').strip()
            numbers = item.get('numbers', {})
            section_context = item.get('_section_context')
            
            if not description:
                continue
            
            desc_lower = description.lower().strip()
            
            # Try enhanced patterns first
            template_field = None
            section = None
            confidence = 0.0
            
            for pattern, (field, sect) in enhanced_patterns.items():
                if re.search(pattern, desc_lower):
                    template_field = field
                    section = sect
                    confidence = 0.95
                    print(f"✅ ENHANCED MATCH: {description[:45]}...")
                    print(f"   → {template_field} (section: {section})")
                    break
            
            # If no enhanced pattern matched, use LLM for operating activities
            if not template_field and section_context == 'operating':
                # Use LLM to classify into the two operating categories
                llm_field, llm_section, llm_confidence = self.classify_cfs_item_with_llm(description)
                if llm_field and llm_section:
                    template_field = llm_field
                    section = llm_section
                    confidence = llm_confidence
                    print(f"🤖 LLM CLASSIFICATION: {description[:45]}...")
                    print(f"   → {template_field} (section: {section})")
                else:
                    # Fallback to default operating category
                    template_field = 'Changes in Assets and Liabilities'
                    section = 'operating'
                    confidence = 0.6
                    print(f"🔄 DEFAULT OPERATING: {description[:45]}...")
                    print(f"   → {template_field} (section: {section})")
            
            # For non-operating sections, use section context fallback
            elif not template_field and section_context:
                if section_context == 'investing':
                    template_field = 'CapEx'
                    section = 'investing'
                    confidence = 0.7
                elif section_context == 'financing':
                    template_field = 'Retirement of Debt'
                    section = 'financing'
                    confidence = 0.7
                
                if template_field:
                    print(f"🔄 SECTION FALLBACK: {description[:45]}...")
                    print(f"   → {template_field} (section: {section})")
            
            if template_field and section and confidence > 0:
                # Parse numerical values
                value_2023 = None
                value_2024 = None
                
                if '2023' in numbers and numbers['2023']:
                    try:
                        value_2023 = float(str(numbers['2023']).replace(',', '').replace('$', '').replace('(', '-').replace(')', ''))
                    except:
                        pass
                
                if '2024' in numbers and numbers['2024']:
                    try:
                        value_2024 = float(str(numbers['2024']).replace(',', '').replace('$', '').replace('(', '-').replace(')', ''))
                    except:
                        pass
                
                if value_2023 or value_2024:
                    key = f"{template_field}_{section}_{len(mapped_items)}"
                    mapped_items[key] = CFSMappedValue(
                        original_description=description,
                        template_field=template_field,
                        section=section,
                        value_2023=value_2023,
                        value_2024=value_2024,
                        confidence=confidence,
                        mapping_method="enhanced_patterns",
                        source_data=numbers
                    )
                    
                    v23 = f"${value_2023:,.0f}" if value_2023 else "-"
                    v24 = f"${value_2024:,.0f}" if value_2024 else "-"
                    print(f"   Values: 2023={v23}, 2024={v24}")
            else:
                print(f"❌ UNMAPPED: {description[:45]}...")
        
        print(f"\n🔗 Three-Section Consolidation:")
        print("-" * 50)
        
        # Consolidate mapped items by section and field
        consolidated_items = self.consolidate_multi_mappings_improved(mapped_items)
        
        return consolidated_items
    
    def populate_template(self, mapped_items: Dict[str, CFSMappedValue]) -> str:
        """Populate the cash flow section of the IS.CF sheet with mapped values"""
        try:
            print(f"\n📝 Populating Cash Flow Statement template...")
            
            # Load the working template
            workbook = load_workbook(self.working_template_path)
            
            # Check for IS.CF sheet
            if 'IS.CF' not in workbook.sheetnames:
                print(f"❌ IS.CF sheet not found. Available sheets: {workbook.sheetnames}")
                return ""
            
            worksheet = workbook['IS.CF']
            print(f"   Working with sheet: IS.CF")
            
            # Track populated fields
            populated_fields = []
            
            # *** CORRECT APPROACH: Only fill rows 25-26 ***
            print("   🎯 CORRECT MODE: Only filling rows 25-26 (Operating CF section)")
            print("   📋 Row 23 (Net Cash from Operating): CALCULATED - Skipping")
            print("   📋 Row 24 (Net Income): AUTO-FILLED from IS - Skipping")
            print("   ✅ Row 25 (Changes in noncash items): WE FILL THIS")
            print("   ✅ Row 26 (Changes in Assets and Liabilities): WE FILL THIS")
            
            # *** HARDCODE NET INCOME ROW 20 (as requested by user) ***
            print("   🧪 HARDCODING Net Income row 20 (user requested)")
            worksheet['B20'] = 62866   # 2023 Net Income (user requested)
            worksheet['C20'] = 119074  # 2024 Net Income (user requested)
            print(f"   ✅ Net Income 2023: B20 = 62,866 (hardcoded)")
            print(f"   ✅ Net Income 2024: C20 = 119,074 (hardcoded)")
            populated_fields.append('Net Income (hardcoded)')
            
            # Populate each mapped item
            for key, mapped_value in mapped_items.items():
                template_field = mapped_value.template_field
                
                if template_field in self.template_mappings:
                    col_base, row = self.template_mappings[template_field]
                    
                    # Populate 2023 (column B) and 2024 (column C)
                    if mapped_value.value_2023 is not None:
                        cell_2023 = f"{col_base}{row}"
                        worksheet[cell_2023] = mapped_value.value_2023
                        print(f"   ✅ {template_field} 2023: {cell_2023} = {mapped_value.value_2023:,.0f}")
                    
                    if mapped_value.value_2024 is not None:
                        cell_2024 = f"C{row}"  # Column C for 2024
                        worksheet[cell_2024] = mapped_value.value_2024
                        print(f"   ✅ {template_field} 2024: {cell_2024} = {mapped_value.value_2024:,.0f}")
                    
                    populated_fields.append(template_field)
            
            # Add starting and ending cash if available
            if self.extracted_starting_cash_2023 or self.extracted_starting_cash_2024:
                if 'Starting Cash' in self.template_mappings:
                    col_base, row = self.template_mappings['Starting Cash']
                    
                    if self.extracted_starting_cash_2023:
                        worksheet[f"{col_base}{row}"] = self.extracted_starting_cash_2023
                        print(f"   ✅ Starting Cash 2023: {col_base}{row} = {self.extracted_starting_cash_2023:,.0f}")
                    
                    if self.extracted_starting_cash_2024:
                        worksheet[f"C{row}"] = self.extracted_starting_cash_2024
                        print(f"   ✅ Starting Cash 2024: C{row} = {self.extracted_starting_cash_2024:,.0f}")
                    
                    populated_fields.append('Starting Cash')
            
            if self.extracted_ending_cash_2023 or self.extracted_ending_cash_2024:
                if 'Ending Cash' in self.template_mappings:
                    col_base, row = self.template_mappings['Ending Cash']
                    
                    if self.extracted_ending_cash_2023:
                        worksheet[f"{col_base}{row}"] = self.extracted_ending_cash_2023
                        print(f"   ✅ Ending Cash 2023: {col_base}{row} = {self.extracted_ending_cash_2023:,.0f}")
                    
                    if self.extracted_ending_cash_2024:
                        worksheet[f"C{row}"] = self.extracted_ending_cash_2024
                        print(f"   ✅ Ending Cash 2024: C{row} = {self.extracted_ending_cash_2024:,.0f}")
                    
                    populated_fields.append('Ending Cash')
            
            # Save the populated template
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"populated_cfs_template_{timestamp}.xlsx"
            output_path = Path(output_filename)
            
            workbook.save(output_path)
            workbook.close()
            
            print(f"\n✅ Cash Flow Statement template populated successfully!")
            print(f"   Output file: {output_filename}")
            print(f"   Fields populated: {len(populated_fields)}")
            print(f"   Fields: {', '.join(populated_fields)}")
            
            return str(output_path)
            
        except Exception as e:
            print(f"❌ Error populating template: {e}")
            return ""
    
    def analyze_coverage(self, mapped_items: Dict[str, CFSMappedValue]):
        """Analyze mapping coverage and identify gaps"""
        print(f"\n📊 CASH FLOW STATEMENT MAPPING ANALYSIS:")
        print("=" * 60)
        
        # Count mapped items by section
        section_counts = {}
        for key, mapped_value in mapped_items.items():
            section = mapped_value.section
            section_counts[section] = section_counts.get(section, 0) + 1
        
        print("Mapped items by section:")
        for section, count in section_counts.items():
            print(f"  {section}: {count} fields")
        
        # Analyze template field coverage
        template_fields = set(self.template_mappings.keys())
        mapped_fields = set(mv.template_field for mv in mapped_items.values())
        
        print(f"\nTotal unique template fields mapped: {len(mapped_fields)}")
        print(f"Template fields: {sorted(mapped_fields)}")
        
        # Define required fields for cash flow
        required_fields = {
            'Net Income', 'Changes in noncash items', 'Changes in Assets and Liabilities',
            'CapEx', 'Net Cash from Operating'
        }
        
        mapped_required = required_fields.intersection(mapped_fields)
        missing_required = required_fields - mapped_fields
        
        coverage_pct = len(mapped_required) / len(required_fields) * 100
        print(f"\nRequired field coverage: {len(mapped_required)}/{len(required_fields)} ({coverage_pct:.1f}%)")
        print(f"✅ Mapped: {sorted(mapped_required)}")
        if missing_required:
            print(f"❌ Missing: {sorted(missing_required)}")
        
        # Optional field analysis
        optional_fields = template_fields - required_fields
        mapped_optional = optional_fields.intersection(mapped_fields)
        
        if mapped_optional:
            print(f"\n🎯 Optional fields mapped: {sorted(mapped_optional)}")
    
    def cleanup_template(self):
        """Clean up working template file"""
        try:
            if self.working_template_path.exists():
                os.remove(self.working_template_path)
                print(f"🧹 Cleaned up working template: {self.working_template_path.name}")
        except Exception as e:
            print(f"⚠️ Could not clean up template: {e}")

    def classify_cfs_item_with_llm(self, description: str) -> Tuple[Optional[str], Optional[str], float]:
        """
        Use LLM to classify cash flow statement items into one of two operating categories:
        1. Changes in noncash items (depreciation, amortization, non-cash adjustments)
        2. Changes in Assets and Liabilities (working capital changes)
        """
        try:
            prompt = f"""You are a financial expert classifying cash flow statement line items.

TASK: Classify this cash flow statement line item into exactly ONE of these two categories:

CATEGORY 1: "Changes in noncash items"
- Depreciation and amortization
- Impairment losses
- Stock-based compensation  
- Gain/loss on asset sales
- Bad debt provisions
- Deferred taxes
- Non-cash adjustments to reconcile net income

CATEGORY 2: "Changes in Assets and Liabilities" 
- Changes in accounts receivable
- Changes in inventory
- Changes in prepaid expenses
- Changes in accounts payable
- Changes in accrued liabilities
- Working capital changes
- Changes in operating assets and liabilities

LINE ITEM TO CLASSIFY: "{description}"

RESPOND WITH EXACTLY ONE OF:
- "Changes in noncash items"
- "Changes in Assets and Liabilities" 
- "UNCLEAR" (if you cannot determine)

RESPONSE:"""

            response = requests.post(
                'http://localhost:11434/api/generate',
                json={
                    'model': 'llama3.2:3b',
                    'prompt': prompt,
                    'stream': False,
                    'options': {
                        'temperature': 0.1,
                        'top_p': 0.9,
                        'num_predict': 10
                    }
                },
                timeout=10
            )
            
            if response.status_code == 200:
                result = response.json()
                llm_response = result['response'].strip().lower()
                
                # Parse LLM response
                if 'changes in noncash items' in llm_response:
                    return 'Changes in noncash items', 'operating', 0.9
                elif 'changes in assets and liabilities' in llm_response:
                    return 'Changes in Assets and Liabilities', 'operating', 0.9
                else:
                    print(f"   🤖 LLM unclear response: {result['response'][:50]}...")
                    return None, None, 0.0
            else:
                print(f"   ❌ LLM request failed: {response.status_code}")
                return None, None, 0.0
                
        except Exception as e:
            print(f"   ❌ LLM classification error: {e}")
            return None, None, 0.0

def main():
    """Main execution function"""
    print("Starting Enhanced Cash Flow Statement Mapper...")
    print("=" * 60)
    print("ENHANCED CASH FLOW STATEMENT KNOWLEDGE GRAPH MAPPER")
    print("=" * 70)
    print("NEW FEATURES:")
    print("1. Three-section processing: Operating → Investing → Financing")
    print("2. Stops after 'Net cash from financing activities'")
    print("3. Proper section-based classification")
    print("4. Enhanced pattern matching for US_Venture CFS")
    print()
    
    mapper = FinalCFSMapper()
    
    # Setup template
    if not mapper.setup_template():
        print("❌ Failed to setup template")
        return
    
    # Process PDF
    pdf_path = "../../input_pdfs/US_Venture_2024.pdf"
    print(f"Processing {pdf_path}")
    
    try:
        # Extract and process cash flow statement
        mapped_items = mapper.extract_and_process(pdf_path)
        
        if mapped_items:
            # Analyze mapping coverage
            mapper.analyze_coverage(mapped_items)
            
            # Populate template
            output_path = mapper.populate_template(mapped_items)
            
            if output_path:
                # Save mapping data
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                json_filename = f"final_cfs_us_venture_{timestamp}.json"
                
                # Convert mapped items to JSON-serializable format
                json_data = {}
                for key, mapped_value in mapped_items.items():
                    json_data[key] = {
                        'original_description': mapped_value.original_description,
                        'template_field': mapped_value.template_field,
                        'section': mapped_value.section,
                        'value_2023': mapped_value.value_2023,
                        'value_2024': mapped_value.value_2024,
                        'confidence': mapped_value.confidence,
                        'mapping_method': mapped_value.mapping_method,
                        'source_data': mapped_value.source_data
                    }
                
                with open(json_filename, 'w') as f:
                    json.dump(json_data, f, indent=2)
                
                print(f"💾 Mapping data saved: {json_filename}")
                
                # Copy to output directory
                output_dir = Path("../../output_excel")
                if output_dir.exists():
                    output_copy = output_dir / output_path
                    shutil.copy2(output_path, output_copy)
                    print(f"📁 Output copied to: {output_copy}")
            
            print(f"\n✅ Enhanced Cash Flow Statement Mapping Complete!")
            print(f"   Net change stopping: {'YES' if mapper.extracted_net_change_2023 or mapper.extracted_net_change_2024 else 'NO'}")
            if mapper.extracted_net_change_2023:
                print(f"   Extracted Net Change 2023: ${mapper.extracted_net_change_2023:,.0f}")
            if mapper.extracted_net_change_2024:
                print(f"   Extracted Net Change 2024: ${mapper.extracted_net_change_2024:,.0f}")
        
        else:
            print("❌ No cash flow items were successfully mapped")
    
    except Exception as e:
        print(f"❌ Error processing cash flow statement: {e}")
    
    finally:
        # Cleanup
        mapper.cleanup_template()

if __name__ == "__main__":
    main() 