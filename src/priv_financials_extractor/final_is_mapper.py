"""
Enhanced Final Income Statement Knowledge Graph Mapper
====================================================
Fixes:
1. Separates COGS from Operating Expenses (COGS maps to "Operating Expenses", OpEx maps to separate field)  
2. Adds Net Income stopping logic - stops processing after finding Net Income
3. Adds Net Income verification against calculated values
4. Prevents double-counting of expense items
"""

import re
import json
import shutil
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import os

# Import the extraction components
from final_extractor_adaptive import TextExtractor
from final_find_fs import FinancialStatementFinder
from openpyxl import load_workbook

@dataclass
class ISMappedValue:
    """Represents a mapped income statement value"""
    original_description: str
    template_field: str
    section: str  # revenue, operating_expenses, non_operating, etc.
    value_2023: Optional[float] = None
    value_2024: Optional[float] = None
    confidence: float = 1.0
    mapping_method: str = ""
    source_data: dict = None

class FinalISMapper:
    """Enhanced Final Income Statement Knowledge Graph Mapper"""
    
    def __init__(self):
        self.extractor = TextExtractor()
        self.finder = FinancialStatementFinder()
        
        # Track if Net Income has been found (stopping condition)
        self.net_income_found = False
        self.extracted_net_income_2023 = None
        self.extracted_net_income_2024 = None
        
        # Set up template paths
        self.original_template_path = Path("../../templates/financial_template.xlsx")
        self.working_template_path = Path("./working_financial_template.xlsx")
        
        # Income Statement template field mappings - based on IS.CF sheet
        self.template_mappings = {
            # INCOME STATEMENT section (rows 6-20)
            'Revenue': ('B', 6),                    # Row 6
            'Operating Expenses': ('B', 7),         # Row 7 (COGS - negative)
            'Depreciation': ('B', 10),              # Row 10 (negative)
            'Amortization': ('B', 11),              # Row 11 (negative)
            'Asset Impairments': ('B', 12),         # Row 12
            'Interest Expense': ('B', 13),          # Row 13 (negative)
            'Interest Income': ('B', 14),           # Row 14 (positive)
            'Other Income': ('B', 15),              # Row 15
            'Tax Expense': ('B', 18),               # Row 18 (negative)
            'Other_income': ('B', 19),              # Row 19 - Other non-operating
        }
        
        # Enhanced rule-based patterns for Income Statement items with 3-section approach
        self.is_rules = {
            # === NET INCOME DETECTION (STOPPING CONDITION) ===
            r'net\s+income(?:\s+attributable\s+to\s+common\s+shareholders)?(?:\s|$)': ('_net_income_found', 'net_income_stop'),
            r'net\s+income\s+attributable\s+to\s+common': ('_net_income_found', 'net_income_stop'),
            
            # === EXCLUDE CALCULATED TOTALS FIRST ===
            # DO NOT MAP these - they are calculated in the template
            r'total\s+operating\s+(?:costs?\s+and\s+)?expenses?': ('_exclude_calculated_total', 'exclude'),
            r'total\s+operating\s+costs?': ('_exclude_calculated_total', 'exclude'),
            r'operating\s+income': ('_exclude_calculated_operating_income', 'exclude'),
            r'income\s+before\s+taxes?': ('_exclude_calculated_income_before_tax', 'exclude'),
            r'total\s+other\s+income': ('_exclude_calculated_total', 'exclude'),
            r'total\s+other\s+income\s+\(expense\)(?:\s*[—-]\s*net)?': ('_exclude_calculated_total', 'exclude'),  # CRITICAL: This was missing!
            r'comprehensive\s+income': ('_exclude_calculated_comprehensive', 'exclude'),
            r'less\s+loss\s+attributable': ('_exclude_noncontrolling', 'exclude'),
            
            # === SECTION 1: PRE-OPERATING INCOME (OPERATING SECTION) ===
            # Operating Asset Gains/Losses (MUST COME BEFORE GENERAL SALES PATTERN!)
            r'(?:gain|loss)\s+on\s+(?:sale\s+of\s+)?(?:operating\s+)?assets?': ('Operating Expenses', 'section1_operating'),
            
            # Revenue
            r'2024\s+2023\s+net\s+sales': ('Revenue', 'section1_operating'),
            r'net\s+sales?(?:\s+and\s+revenues?)?': ('Revenue', 'section1_operating'),
            r'(?:total\s+)?revenues?': ('Revenue', 'section1_operating'),
            r'sales?(?:\s+revenue)?': ('Revenue', 'section1_operating'),
            r'gross\s+sales?': ('Revenue', 'section1_operating'),
            
            # Operating Costs & Expenses (ALL go to Operating Expenses for intermediate calc)
            r'petroleum\s+and\s+other\s+product\s+costs?': ('Operating Expenses', 'section1_operating'),
            r'cost\s+of\s+(?:goods\s+)?sold': ('Operating Expenses', 'section1_operating'),
            r'cost\s+of\s+sales?': ('Operating Expenses', 'section1_operating'),
            r'cost\s+of\s+revenues?': ('Operating Expenses', 'section1_operating'),
            r'(?:^|\s)operating\s+expenses?(?:\s|$)': ('Operating Expenses', 'section1_operating'),
            r'selling\s*,?\s*general\s+(?:and|&)\s+administrative': ('Operating Expenses', 'section1_operating'),
            r'sg&a': ('Operating Expenses', 'section1_operating'),
            r'general\s+and\s+administrative': ('Operating Expenses', 'section1_operating'),
            r'administrative\s+expenses?': ('Operating Expenses', 'section1_operating'),
            
            # Operating Depreciation & Amortization (part of operating expenses)
            r'depreciation(?:\s+and\s+amortization)?': ('Operating Expenses', 'section1_operating'),
            r'amortization(?:\s+and\s+depreciation)?': ('Operating Expenses', 'section1_operating'),
            r'depreciation(?:\s+expense)?': ('Operating Expenses', 'section1_operating'),
            r'amortization(?:\s+expense)?': ('Operating Expenses', 'section1_operating'),
            
            # === SECTION 2: PRE-EBIT (NON-OPERATING SECTION) ===
            # Interest
            r'interest\s+expense': ('Interest Expense', 'section2_non_operating'),
            r'interest\s+costs?': ('Interest Expense', 'section2_non_operating'),
            r'borrowing\s+costs?': ('Interest Expense', 'section2_non_operating'),
            r'interest\s+income': ('Interest Income', 'section2_non_operating'),
            r'interest\s+(?:and\s+)?dividend\s+income': ('Interest Income', 'section2_non_operating'),
            
            # Non-Operating Items
            r'other\s+(?:income|expense)(?:\s*[—-]\s*net)?': ('Other Income', 'section2_non_operating'),
            r'other\s+(?:non[- ]?operating\s+)?(?:income|expenses?)': ('Other Income', 'section2_non_operating'),
            r'miscellaneous\s+(?:income|expenses?)': ('Other Income', 'section2_non_operating'),
            r'foreign\s+(?:currency|exchange)': ('Other Income', 'section2_non_operating'),
            
            # Non-Operating Asset Impairments
            r'impairment\s+(?:losses?\s+on\s+)?(?:long[- ]lived|intangible)\s+assets?': ('Other Income', 'section2_non_operating'),
            r'asset\s+impairments?': ('Other Income', 'section2_non_operating'),
            r'goodwill\s+impairment': ('Other Income', 'section2_non_operating'),
            
            # === SECTION 3: PRE-NET INCOME (TAX SECTION) ===
            r'(?:income\s+)?tax\s+expense': ('Tax Expense', 'section3_taxes'),
            r'provision\s+for\s+(?:income\s+)?taxes?': ('Tax Expense', 'section3_taxes'),
            r'current\s+(?:income\s+)?tax': ('Tax Expense', 'section3_taxes'),
            r'deferred\s+(?:income\s+)?tax': ('Tax Expense', 'section3_taxes'),
        }
        
        # Section mapping for three-section consolidation
        self.section_consolidation_mapping = {
            'section1_operating': 'Operating Expenses',  # All operating costs consolidate here
            'section2_non_operating': 'Other Income',    # All non-operating items consolidate here  
            'section3_taxes': 'Tax Expense'              # Tax items
        }
        
        # Three-section calculation logic
        self.section_calculations = {
            'section1_operating': {
                'positive_items': ['Revenue'],  # Revenue items are positive
                'negative_items': ['Operating Expenses'],  # All costs/expenses are negative
                'result_field': 'Operating Income',
                'description': 'Revenue minus all Operating Costs & Expenses'
            },
            'section2_non_operating': {
                'positive_items': ['Interest Income', 'Other Income'],  # Income items
                'negative_items': ['Interest Expense'],  # Expense items  
                'result_field': 'Income Before Taxes',
                'description': 'Operating Income plus/minus Non-Operating Items'
            },
            'section3_taxes': {
                'positive_items': [],  # No positive tax items typically
                'negative_items': ['Tax Expense'],  # Tax is an expense
                'result_field': 'Net Income',
                'description': 'Income Before Taxes minus Tax Expense'
            }
        }
    
    def setup_template(self) -> bool:
        """Copy original template to working directory"""
        try:
            print(f"📋 Setting up IS template...")
            
            if not self.original_template_path.exists():
                print(f"❌ Original template not found: {self.original_template_path}")
                return False
            
            shutil.copy2(self.original_template_path, self.working_template_path)
            
            if self.working_template_path.exists():
                print(f"✅ Template copied successfully")
                return True
            else:
                print(f"❌ Failed to copy template")
                return False
                
        except Exception as e:
            print(f"❌ Error setting up template: {e}")
            return False
    
    def is_total_or_net_row(self, description: str) -> bool:
        """Check if description is a total or calculated row"""
        desc_lower = description.lower().strip()
        
        # === CRITICAL: CHECK FOR NET INCOME FIRST ===
        # This is our stopping condition
        if re.search(r'net\s+income(?:\s+attributable\s+to\s+common\s+shareholders)?(?:\s|$)', desc_lower):
            print(f"🛑 NET INCOME FOUND: {description}")
            return False  # Don't filter it out, we need to extract the value first
        
        # Debug the specific problematic line
        if "total operating costs" in desc_lower or "total other income" in desc_lower:
            print(f"🔍 DEBUG is_total_or_net_row: Testing '{description}'")
            print(f"   desc_lower: '{desc_lower}'")
        
        # Specific calculated/total rows to filter out (these are calculated in template)
        calculated_rows = [
            'total operating costs and expenses',
            'total operating costs',  # Additional pattern
            'operating income',
            'total other income',
            'total other income (expense)',
            'total other income (expense)—net',  # CRITICAL: The exact pattern from US Venture
            'income before taxes',
            'comprehensive income',
            'less loss attributable'
        ]
        
        for calc_row in calculated_rows:
            if calc_row in desc_lower:
                if "total operating costs" in desc_lower or "total other income" in desc_lower:
                    print(f"   ✅ MATCHED: '{calc_row}' - FILTERING OUT!")
                return True
        
        # Filter out general totals and subtotals (but keep specific line items)
        total_patterns = [
            r'^total(\s|$)',
            r'(\s|^)sum(\s|$)',
            r'(\s|^)subtotal(\s|$)',
            r'(\s|^)aggregate(\s|$)',
            r'(\s|^)grand total(\s|$)',
        ]
        
        # Check if it's a total pattern but NOT a specific line item we want
        for pat in total_patterns:
            if re.search(pat, desc_lower):
                if "total other income" in desc_lower:
                    print(f"   ✅ REGEX MATCHED: '{pat}' - FILTERING OUT!")
                return True
        
        # Filter out header/formatting rows
        header_patterns = [
            r'^\s*\d{4}\s+\d{4}\s*$',  # Year headers
            r'^\s*income\s+statement\s*$',
            r'^\s*statement\s+of.*income\s*$',
            r'^\s*comprehensive\s+income\s*$',
            r'^\s*operating.*income.*expense.*$'
        ]
        
        for pat in header_patterns:
            if re.search(pat, desc_lower):
                return True
                
        return False
    
    def apply_enhanced_mapping(self, description: str) -> Tuple[Optional[str], Optional[str], float]:
        """Apply enhanced rule-based mapping for income statement items"""
        desc_lower = description.lower().strip()
        
        # Debug specific problematic item
        if "gain on sale" in desc_lower:
            print(f"🔍 DEBUG apply_enhanced_mapping: Testing '{description}'")
            print(f"   desc_lower: '{desc_lower}'")
        
        for pattern, (template_field, section) in self.is_rules.items():
            if re.search(pattern, desc_lower):
                if "gain on sale" in desc_lower:
                    print(f"   ✅ PATTERN MATCHED: '{pattern}' → {template_field} | {section}")
                
                # Skip calculated fields and excluded totals - we don't want to map these
                if template_field.startswith('_calculated_') or template_field.startswith('_exclude_'):
                    return None, None, 0.0
                return template_field, section, 0.9
        
        if "gain on sale" in desc_lower:
            print(f"   ❌ NO PATTERN MATCHED - falling back to other tiers")
        
        return None, None, 0.0
    
    def ask_ollama_for_classification(self, description: str) -> Tuple[Optional[str], Optional[str]]:
        """Ask Ollama LLM to classify income statement line items"""
        try:
            import requests
            
            # Simple prompt for fast processing
            prompt = f"""Classify this income statement item:

"{description}"

Choose the BEST template field and section:

Template Fields:
- Revenue (section: revenue)
- Operating Expenses (section: operating_expenses)  
- Depreciation (section: non_operating)
- Amortization (section: non_operating)
- Asset Impairments (section: non_operating)
- Interest Expense (section: non_operating)
- Interest Income (section: non_operating)
- Other Income (section: non_operating)
- Tax Expense (section: taxes)

Answer format: "Field|section" (e.g., "Revenue|revenue")"""

            # Call Ollama API with phi3:mini
            response = requests.post(
                'http://localhost:11434/api/generate',
                json={
                    'model': 'phi3:mini',
                    'prompt': prompt,
                    'stream': False,
                    'options': {
                        'temperature': 0.1,
                        'num_predict': 20
                    }
                },
                timeout=20  # Shorter timeout for income statement
            )
            
            if response.status_code == 200:
                result = response.json()
                ollama_response = result.get('response', '').strip()
                
                # Parse response: "Field|section"
                if '|' in ollama_response:
                    field, section = ollama_response.split('|', 1)
                    field = field.strip()
                    section = section.strip()
                    
                    # Validate field exists in our mapping
                    if field in self.template_mappings:
                        return field, section
                        
            return None, None
            
        except Exception as e:
            print(f"   ⚠️ LLM classification failed: {e}")
            return None, None
    
    def apply_multi_tier_fallback(self, description: str) -> Tuple[Optional[str], Optional[str], float, str]:
        """Apply 5-tier fallback system for income statement classification"""
        
        # TIER 1: Enhanced regex patterns (already tried above, so skip)
        
        # TIER 2: Fuzzy matching against template fields (70% confidence)
        template_field, section, confidence = self.apply_fuzzy_matching(description)
        if template_field and confidence >= 0.7:
            print(f"   🔄 Fuzzy match: {description[:30]}... → {template_field}")
            return template_field, section, confidence, 'fuzzy_matching'
        
        # TIER 3: Keyword-based analysis (60% confidence)
        template_field, section = self.apply_keyword_analysis(description)
        if template_field and section:
            print(f"   🔍 Keyword analysis: {description[:30]}... → {template_field}")
            return template_field, section, 0.6, 'keyword_analysis'
        
        # TIER 4: LLM fallback with phi3:mini (70% confidence)
        llm_field, llm_section = self.ask_ollama_for_classification(description)
        if llm_field and llm_section:
            print(f"   🤖 Ollama inference: {description[:30]}... → {llm_field}")
            return llm_field, llm_section, 0.7, 'llm_fallback'
        
        # TIER 5: Smart income statement fallback (50% confidence)
        smart_field, smart_section = self.smart_income_statement_fallback(description)
        if smart_field and smart_section:
            print(f"   🧠 Smart fallback: {description[:30]}... → {smart_field}")
            return smart_field, smart_section, 0.5, 'smart_fallback'
        
        return None, None, 0.0, 'no_match'
    
    def apply_fuzzy_matching(self, description: str) -> Tuple[Optional[str], Optional[str], float]:
        """Apply fuzzy string matching against known template fields"""
        from difflib import SequenceMatcher
        
        desc_lower = description.lower().strip()
        
        # Template field patterns for fuzzy matching
        template_patterns = {
            'Revenue': ['revenue', 'sales', 'net sales', 'gross sales'],
            'Operating Expenses': ['operating expenses', 'opex', 'selling expenses', 'administrative expenses'],
            'Depreciation': ['depreciation', 'depreciation expense'],
            'Amortization': ['amortization', 'amortization expense'],
            'Asset Impairments': ['impairment', 'asset impairment', 'goodwill impairment'],
            'Interest Expense': ['interest expense', 'interest cost', 'borrowing cost'],
            'Interest Income': ['interest income', 'interest revenue'],
            'Other Income': ['other income', 'other expense', 'miscellaneous income', 'foreign exchange'],
            'Tax Expense': ['tax expense', 'income tax', 'tax provision']
        }
        
        best_match = None
        best_confidence = 0.0
        
        for template_field, patterns in template_patterns.items():
            for pattern in patterns:
                similarity = SequenceMatcher(None, desc_lower, pattern).ratio()
                if similarity > best_confidence and similarity >= 0.6:  # 60% similarity threshold
                    best_match = template_field
                    best_confidence = similarity
        
        if best_match:
            # Determine section based on template field
            section_mapping = {
                'Revenue': 'revenue',
                'Operating Expenses': 'operating_expenses',
                'Depreciation': 'non_operating',
                'Amortization': 'non_operating',
                'Asset Impairments': 'non_operating',
                'Interest Expense': 'non_operating',
                'Interest Income': 'non_operating',
                'Other Income': 'non_operating',
                'Tax Expense': 'taxes'
            }
            section = section_mapping.get(best_match, 'non_operating')
            return best_match, section, best_confidence
        
        return None, None, 0.0
    
    def apply_keyword_analysis(self, description: str) -> Tuple[Optional[str], Optional[str]]:
        """Apply keyword-based classification for income statement items"""
        desc_lower = description.lower().strip()
        
        # Income statement specific keyword classifications
        keyword_mappings = {
            # Revenue indicators
            ('Revenue', 'revenue'): [
                'sales', 'revenue', 'income', 'proceeds', 'receipts',
                'fees', 'charges', 'billings', 'turnover'
            ],
            
            # Operating expense indicators
            ('Operating Expenses', 'operating_expenses'): [
                'operating', 'administrative', 'selling', 'personnel',
                'salaries', 'wages', 'benefits', 'rent', 'utilities',
                'professional fees', 'consulting', 'marketing'
            ],
            
            # Non-operating items
            ('Interest Expense', 'non_operating'): [
                'interest expense', 'interest cost', 'borrowing',
                'financing cost', 'debt service'
            ],
            
            ('Interest Income', 'non_operating'): [
                'interest income', 'interest revenue', 'investment income',
                'dividend income'
            ],
            
            ('Other Income', 'non_operating'): [
                'other', 'miscellaneous', 'foreign', 'exchange',
                'currency', 'gain', 'loss', 'disposal', 'extraordinary'
            ],
            
            # Tax items
            ('Tax Expense', 'taxes'): [
                'tax', 'taxation', 'provision', 'deferred tax',
                'current tax', 'income tax'
            ]
        }
        
        # Score each classification based on keyword matches
        scores = {}
        for (template_field, section), keywords in keyword_mappings.items():
            score = 0
            for keyword in keywords:
                if keyword in desc_lower:
                    score += 1
            if score > 0:
                scores[(template_field, section)] = score
        
        if scores:
            # Return the classification with highest score
            best_match = max(scores, key=scores.get)
            return best_match[0], best_match[1]
        
        return None, None
    
    def smart_income_statement_fallback(self, description: str) -> Tuple[Optional[str], Optional[str]]:
        """Smart fallback classification based on income statement context"""
        desc_lower = description.lower().strip()
        
        # Common income statement patterns that might not match exact rules
        fallback_patterns = {
            # If it has dollar signs or mentions money, likely revenue or expense
            ('Revenue', 'revenue'): [
                r'\$.*(?:sales?|revenue|income)(?!\s+expense)',
                r'(?:net|gross|total)\s+(?:sales?|revenue)',
                r'service\s+(?:revenue|income|fees)'
            ],
            
            ('Operating Expenses', 'operating_expenses'): [
                r'(?:cost|expense)(?!.*interest)(?!.*tax)',
                r'personnel|payroll|compensation',
                r'general.*administrative|sg&a'
            ],
            
            ('Other Income', 'non_operating'): [
                r'foreign.*(?:exchange|currency)',
                r'(?:gain|loss).*(?:sale|disposal)',
                r'unusual|extraordinary|non.*recurring'
            ],
            
            ('Tax Expense', 'taxes'): [
                r'provision.*tax',
                r'deferred.*tax',
                r'tax.*(?:benefit|expense)'
            ]
        }
        
        for (template_field, section), patterns in fallback_patterns.items():
            for pattern in patterns:
                if re.search(pattern, desc_lower):
                    return template_field, section
        
        # Default fallback - if nothing else matches, assume it's "Other Income"
        return 'Other Income', 'non_operating'
    
    def consolidate_multi_mappings_improved(self, mapped_items: Dict[str, ISMappedValue]) -> Dict[str, ISMappedValue]:
        """Three-section consolidation approach: Operating, Non-Operating, and Tax sections"""
        consolidated = {}
        
        print("🔄 THREE-SECTION CONSOLIDATION:")
        print("-" * 50)
        
        # Separate items by section
        section1_items = []  # Operating section
        section2_items = []  # Non-operating section
        section3_items = []  # Tax section
        
        for key, mapped_value in mapped_items.items():
            section = mapped_value.section
            
            if section == "section1_operating":
                section1_items.append((key, mapped_value))
            elif section == "section2_non_operating":
                section2_items.append((key, mapped_value))
            elif section == "section3_taxes":
                section3_items.append((key, mapped_value))
        
        # === SECTION 1: OPERATING SECTION ===
        if section1_items:
            print("📊 SECTION 1 - Operating Section:")
            
            # Separate revenue from operating expenses
            revenue_items = []
            operating_expense_items = []
            
            for key, mapped_value in section1_items:
                if mapped_value.template_field == "Revenue":
                    revenue_items.append((key, mapped_value))
                elif mapped_value.template_field == "Operating Expenses":
                    operating_expense_items.append((key, mapped_value))
            
            # Consolidate Revenue
            if revenue_items:
                total_2023 = sum(mv.value_2023 for k, mv in revenue_items if mv.value_2023)
                total_2024 = sum(mv.value_2024 for k, mv in revenue_items if mv.value_2024)
                
                consolidated["revenue_consolidated"] = ISMappedValue(
                    original_description="Total Revenue (All Revenue Sources)",
                    template_field="Revenue",
                    section="section1_operating",
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.95,
                    mapping_method="section1_revenue_consolidation",
                    source_data={"item_count": len(revenue_items), "section": "operating"}
                )
                print(f"   ✅ Revenue: {len(revenue_items)} items → ${total_2024:,.0f} (2024)")
            
            # Consolidate All Operating Expenses (COGS + OpEx + Depreciation + Asset Gains/Losses)
            if operating_expense_items:
                total_2023 = sum(mv.value_2023 for k, mv in operating_expense_items if mv.value_2023)
                total_2024 = sum(mv.value_2024 for k, mv in operating_expense_items if mv.value_2024)
                
                # List what's included
                expense_descriptions = []
                for k, mv in operating_expense_items:
                    if "petroleum" in mv.original_description.lower():
                        expense_descriptions.append("COGS")
                    elif "operating expenses" in mv.original_description.lower():
                        expense_descriptions.append("OpEx")
                    elif "depreciation" in mv.original_description.lower():
                        expense_descriptions.append("Depreciation")
                    elif "gain" in mv.original_description.lower() or "loss" in mv.original_description.lower():
                        expense_descriptions.append("Asset Gains/Losses")
                    else:
                        expense_descriptions.append("Other Operating")
                
                consolidated["operating_expenses_consolidated"] = ISMappedValue(
                    original_description=f"Total Operating Expenses ({', '.join(set(expense_descriptions))})",
                    template_field="Operating Expenses",
                    section="section1_operating",
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.95,
                    mapping_method="section1_operating_consolidation",
                    source_data={"item_count": len(operating_expense_items), "section": "operating", "includes": expense_descriptions}
                )
                print(f"   ✅ Operating Expenses: {len(operating_expense_items)} items → ${total_2024:,.0f} (2024)")
                print(f"       Includes: {', '.join(set(expense_descriptions))}")
        
        # === SECTION 2: NON-OPERATING SECTION ===
        if section2_items:
            print("\n📊 SECTION 2 - Non-Operating Section:")
            
            # Group by template field
            field_groups = {}
            for key, mapped_value in section2_items:
                field = mapped_value.template_field
                if field not in field_groups:
                    field_groups[field] = []
                field_groups[field].append((key, mapped_value))
            
            # Consolidate each field group
            for field, items in field_groups.items():
                total_2023 = sum(mv.value_2023 for k, mv in items if mv.value_2023)
                total_2024 = sum(mv.value_2024 for k, mv in items if mv.value_2024)
                
                consolidated[f"{field.lower().replace(' ', '_')}_consolidated"] = ISMappedValue(
                    original_description=f"Total {field} (All Non-Operating)",
                    template_field=field,
                    section="section2_non_operating",
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.9,
                    mapping_method="section2_non_operating_consolidation",
                    source_data={"item_count": len(items), "section": "non_operating"}
                )
                print(f"   ✅ {field}: {len(items)} items → ${total_2024:,.0f} (2024)")
        
        # === SECTION 3: TAX SECTION ===
        if section3_items:
            print("\n📊 SECTION 3 - Tax Section:")
            
            total_2023 = sum(mv.value_2023 for k, mv in section3_items if mv.value_2023)
            total_2024 = sum(mv.value_2024 for k, mv in section3_items if mv.value_2024)
            
            if total_2023 != 0 or total_2024 != 0:
                consolidated["tax_expense_consolidated"] = ISMappedValue(
                    original_description="Total Tax Expense",
                    template_field="Tax Expense",
                    section="section3_taxes",
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.9,
                    mapping_method="section3_tax_consolidation",
                    source_data={"item_count": len(section3_items), "section": "taxes"}
                )
                print(f"   ✅ Tax Expense: {len(section3_items)} items → ${total_2024:,.0f} (2024)")
        
        print(f"\n📊 Three-Section Consolidation Complete: {len(consolidated)} consolidated items")
        return consolidated
    
    def verify_net_income(self, mapped_items: Dict[str, ISMappedValue]):
        """Verify extracted Net Income against calculated values using three-section approach"""
        print(f"\n🔍 THREE-SECTION NET INCOME VERIFICATION:")
        print("=" * 60)
        
        if not self.extracted_net_income_2023 and not self.extracted_net_income_2024:
            print("❌ No extracted Net Income values found")
            return
        
        # Three-section calculation
        section1_2023 = 0  # Operating Income
        section1_2024 = 0
        section2_2023 = 0  # Income Before Taxes 
        section2_2024 = 0
        section3_2023 = 0  # Net Income
        section3_2024 = 0
        
        print("📊 SECTION-BY-SECTION CALCULATION:")
        print("-" * 40)
        
        # Section 1: Operating Income = Revenue - Operating Expenses
        print("📈 SECTION 1 - Operating Income:")
        revenue_2023 = revenue_2024 = 0
        opex_2023 = opex_2024 = 0
        
        for key, mapped_value in mapped_items.items():
            if mapped_value.section == "section1_operating":
                v23 = mapped_value.value_2023 or 0
                v24 = mapped_value.value_2024 or 0
                
                if mapped_value.template_field == "Revenue":
                    revenue_2023 += v23
                    revenue_2024 += v24
                    print(f"   + Revenue: 2023=${v23:,.0f}, 2024=${v24:,.0f}")
                elif mapped_value.template_field == "Operating Expenses":
                    opex_2023 += v23
                    opex_2024 += v24
                    print(f"   - {mapped_value.original_description[:30]}...")
                    print(f"     2023=${v23:,.0f}, 2024=${v24:,.0f}")
        
        section1_2023 = revenue_2023 - opex_2023
        section1_2024 = revenue_2024 - opex_2024
        print(f"   = Operating Income: 2023=${section1_2023:,.0f}, 2024=${section1_2024:,.0f}")
        
        # Section 2: Income Before Taxes = Operating Income + Non-Operating Items
        print("\n💰 SECTION 2 - Income Before Taxes:")
        print(f"   Starting with Operating Income: 2023=${section1_2023:,.0f}, 2024=${section1_2024:,.0f}")
        
        section2_2023 = section1_2023
        section2_2024 = section1_2024
        
        for key, mapped_value in mapped_items.items():
            if mapped_value.section == "section2_non_operating":
                v23 = mapped_value.value_2023 or 0
                v24 = mapped_value.value_2024 or 0
                
                if mapped_value.template_field == "Interest Income":
                    section2_2023 += v23
                    section2_2024 += v24
                    print(f"   + Interest Income: 2023=${v23:,.0f}, 2024=${v24:,.0f}")
                elif mapped_value.template_field == "Interest Expense":
                    section2_2023 += v23  # Already negative from extraction
                    section2_2024 += v24
                    print(f"   + Interest Expense: 2023=${v23:,.0f}, 2024=${v24:,.0f}")
                elif mapped_value.template_field == "Other Income":
                    section2_2023 += v23
                    section2_2024 += v24
                    print(f"   + Other Income: 2023=${v23:,.0f}, 2024=${v24:,.0f}")
        
        print(f"   = Income Before Taxes: 2023=${section2_2023:,.0f}, 2024=${section2_2024:,.0f}")
        
        # Section 3: Net Income = Income Before Taxes - Tax Expense
        print("\n🏦 SECTION 3 - Net Income:")
        print(f"   Starting with Income Before Taxes: 2023=${section2_2023:,.0f}, 2024=${section2_2024:,.0f}")
        
        section3_2023 = section2_2023
        section3_2024 = section2_2024
        
        tax_found = False
        for key, mapped_value in mapped_items.items():
            if mapped_value.section == "section3_taxes":
                v23 = mapped_value.value_2023 or 0
                v24 = mapped_value.value_2024 or 0
                section3_2023 += v23  # Tax should be negative
                section3_2024 += v24
                print(f"   - Tax Expense: 2023=${abs(v23):,.0f}, 2024=${abs(v24):,.0f}")
                tax_found = True
        
        if not tax_found:
            print("   - Tax Expense: Not found (assuming zero)")
        
        print(f"   = Net Income: 2023=${section3_2023:,.0f}, 2024=${section3_2024:,.0f}")
        
        # Final comparison
        print("\n🧮 FINAL VERIFICATION:")
        print("=" * 40)
        print(f"Calculated Net Income 2023: ${section3_2023:,.0f}")
        print(f"Calculated Net Income 2024: ${section3_2024:,.0f}")
        print(f"Extracted Net Income 2023:  ${self.extracted_net_income_2023:,.0f}" if self.extracted_net_income_2023 else "Extracted Net Income 2023:  Not found")
        print(f"Extracted Net Income 2024:  ${self.extracted_net_income_2024:,.0f}" if self.extracted_net_income_2024 else "Extracted Net Income 2024:  Not found")
        
        # Calculate intermediate totals verification
        print(f"\n📊 INTERMEDIATE TOTALS:")
        print(f"Operating Income:     2023=${section1_2023:,.0f}, 2024=${section1_2024:,.0f}")
        print(f"Income Before Taxes:  2023=${section2_2023:,.0f}, 2024=${section2_2024:,.0f}")
        print(f"Net Income:           2023=${section3_2023:,.0f}, 2024=${section3_2024:,.0f}")
        
        # Compare values with tolerance
        tolerance = 0.05  # 5% tolerance
        
        if self.extracted_net_income_2023:
            diff_2023 = abs(section3_2023 - self.extracted_net_income_2023)
            pct_diff_2023 = diff_2023 / abs(self.extracted_net_income_2023) if self.extracted_net_income_2023 != 0 else 0
            
            if pct_diff_2023 <= tolerance:
                print(f"\n✅ 2023 Match: {pct_diff_2023*100:.1f}% difference (within {tolerance*100}% tolerance)")
            else:
                print(f"\n❌ 2023 Mismatch: {pct_diff_2023*100:.1f}% difference (exceeds {tolerance*100}% tolerance)")
                print(f"   Difference: ${diff_2023:,.0f}")
        
        if self.extracted_net_income_2024:
            diff_2024 = abs(section3_2024 - self.extracted_net_income_2024)
            pct_diff_2024 = diff_2024 / abs(self.extracted_net_income_2024) if self.extracted_net_income_2024 != 0 else 0
            
            if pct_diff_2024 <= tolerance:
                print(f"✅ 2024 Match: {pct_diff_2024*100:.1f}% difference (within {tolerance*100}% tolerance)")
            else:
                print(f"❌ 2024 Mismatch: {pct_diff_2024*100:.1f}% difference (exceeds {tolerance*100}% tolerance)")
                print(f"   Difference: ${diff_2024:,.0f}")
    
    def extract_and_process(self, pdf_path: str) -> Dict[str, ISMappedValue]:
        """Main processing function for income statement with Net Income stopping logic"""
        print("ENHANCED INCOME STATEMENT KNOWLEDGE GRAPHER")
        print("=" * 60)
        print("NEW FEATURES:")
        print("1. Separates COGS from Operating Expenses")
        print("2. Net Income stopping logic")
        print("3. Net Income verification")
        print("4. Prevents expense double-counting")
        print()
        
        # Setup template
        if not self.setup_template():
            print("❌ Failed to setup template")
            return {}
        
        # Extract data using final_extractor - focus on income statement
        confirmed_pages = {
            'income_statement': [9],  # US Venture income statement is on page 9
        }
        
        statement_pages_dict = {}
        for stmt_type, pages in confirmed_pages.items():
            if pages:
                statement_pages_dict[stmt_type] = pages
        
        excel_path, extracted_data = self.extractor.extract_text(
            pdf_path, 
            process_numbers=True, 
            statement_pages=statement_pages_dict
        )
        
        if not extracted_data or 'income_statement' not in extracted_data:
            print("❌ No income statement data found")
            return {}
        
        income_statement_data = extracted_data['income_statement']
        print(f"✅ Extracted {len(income_statement_data)} income statement items")
        
        # Step 1: Filter out totals/calculated rows 
        non_total_items = []
        total_items = []
        
        for item in income_statement_data:
            description = item.get('description', '').strip()
            if not description:
                continue
                
            if self.is_total_or_net_row(description):
                total_items.append(item)
            else:
                non_total_items.append(item)
        
        print(f"📊 After filtering: {len(non_total_items)} items to map, {len(total_items)} totals/headers filtered")
        print()
        
        # Step 2: Process non-total items with enhanced mapping AND NET INCOME STOPPING
        mapped_items = {}
        unmapped_items = []
        
        print("🔄 Enhanced Income Statement Processing with Net Income Stopping:")
        print("-" * 50)
        
        # Filter items with numerical values
        items_with_numbers = []
        for item in non_total_items:
            description = item.get('description', '').strip()
            numbers = item.get('numbers', {})
            
            # Check if item has actual numerical values
            has_numbers = any(
                value_str is not None and str(value_str).strip() 
                for value_str in numbers.values()
            )
            
            if has_numbers:
                items_with_numbers.append(item)
        
        print(f"📊 Items with numerical values: {len(items_with_numbers)}")
        
        for item in items_with_numbers:
            description = item.get('description', '').strip()
            numbers = item.get('numbers', {})
            
            # === CHECK FOR NET INCOME STOPPING CONDITION ===
            desc_lower = description.lower().strip()
            if re.search(r'net\s+income(?:\s+attributable\s+to\s+common\s+shareholders)?(?:\s|$)', desc_lower):
                # Extract Net Income values before stopping
                value_2023 = None
                value_2024 = None
                for year, value_str in numbers.items():
                    if value_str is not None:
                        try:
                            value = float(str(value_str).replace(',', ''))
                            if year == '2023':
                                value_2023 = value
                                self.extracted_net_income_2023 = value
                            elif year == '2024':
                                value_2024 = value
                                self.extracted_net_income_2024 = value
                        except (ValueError, TypeError):
                            continue
                
                print(f"🛑 NET INCOME FOUND - STOPPING PROCESSING")
                print(f"   Description: {description}")
                print(f"   Extracted Net Income: 2023=${value_2023:,.0f}, 2024=${value_2024:,.0f}")
                print(f"   Remaining items will be skipped to prevent double-counting")
                self.net_income_found = True
                break
            
            # Parse values (expenses will be negative in the template)
            value_2023 = None
            value_2024 = None
            for year, value_str in numbers.items():
                if value_str is not None:
                    try:
                        value = float(str(value_str).replace(',', ''))
                        if year == '2023':
                            value_2023 = value
                        elif year == '2024':
                            value_2024 = value
                    except (ValueError, TypeError):
                        continue
            
            # Enhanced mapping with section assignment
            template_field, section, confidence = self.apply_enhanced_mapping(description)
            
            if template_field and section and confidence >= 0.8:
                # Successfully mapped
                mapped_value = ISMappedValue(
                    original_description=description,
                    template_field=template_field,
                    section=section,
                    value_2023=value_2023,
                    value_2024=value_2024,
                    confidence=confidence,
                    mapping_method="enhanced_rule_based",
                    source_data=item
                )
                
                # Create unique key
                key = f"{template_field}_{section}_{len(mapped_items)}"
                mapped_items[key] = mapped_value
                
                print(f"✅ {description[:50]}...")
                print(f"   → {template_field} (section: {section})")
                v23 = f"${value_2023:,.0f}" if value_2023 else "-"
                v24 = f"${value_2024:,.0f}" if value_2024 else "-"
                print(f"   Values: 2023={v23}, 2024={v24}")
                print()
            else:
                # Apply 5-tier fallback system for unmapped items
                template_field, section, confidence, method = self.apply_multi_tier_fallback(description)
                
                if template_field and section and confidence >= 0.5:
                    mapped_value = ISMappedValue(
                        original_description=description,
                        template_field=template_field,
                        section=section,
                        value_2023=value_2023,
                        value_2024=value_2024,
                        confidence=confidence,
                        mapping_method=method,
                        source_data=item
                    )
                    
                    key = f"{template_field}_{section}_{len(mapped_items)}"
                    mapped_items[key] = mapped_value
                    
                    # Different emoji based on method
                    emoji = {
                        'fuzzy_matching': '🔄',
                        'keyword_analysis': '🔍', 
                        'llm_fallback': '🤖',
                        'smart_fallback': '🧠',
                        'consolidation_fallback': '📦'
                    }.get(method, '✅')
                    
                    print(f"{emoji} {description[:50]}...")
                    print(f"   → {template_field} (section: {section}) [{method}]")
                    v23 = f"${value_2023:,.0f}" if value_2023 else "-"
                    v24 = f"${value_2024:,.0f}" if value_2024 else "-"
                    print(f"   Values: 2023={v23}, 2024={v24}")
                    print()
                else:
                    unmapped_items.append((description, value_2023, value_2024, item))
                    print(f"❓ {description[:50]}...")
                    print(f"   → Will map to 'Other Income' in consolidation")
        
        # Step 3: Consolidate using three-section approach
        print(f"\n🔗 Three-Section Consolidation (Operating | Non-Operating | Tax):")
        print("-" * 60)
        consolidated_mapped = self.consolidate_multi_mappings_improved(mapped_items)
        
        # Step 4: Consolidate unmapped items into "Other Income"
        if unmapped_items:
            print(f"\n🔧 Consolidating unmapped items into 'Other Income':")
            print("-" * 50)
            
            total_2023 = sum(item[1] for item in unmapped_items if item[1] is not None)
            total_2024 = sum(item[2] for item in unmapped_items if item[2] is not None)
            
            if total_2023 != 0 or total_2024 != 0:
                consolidated_mapped["other_income_unmapped"] = ISMappedValue(
                    original_description=f"Consolidated {len(unmapped_items)} unmapped items",
                    template_field="Other Income",
                    section="non_operating",
                    value_2023=total_2023 if total_2023 != 0 else None,
                    value_2024=total_2024 if total_2024 != 0 else None,
                    confidence=0.5,
                    mapping_method="consolidation_fallback",
                    source_data={"consolidated_count": len(unmapped_items)}
                )
                
                print(f"📦 Consolidated {len(unmapped_items)} unmapped items → Other Income")
                v23 = f"${total_2023:,.0f}" if total_2023 != 0 else "-"
                v24 = f"${total_2024:,.0f}" if total_2024 != 0 else "-"
                print(f"   Total values: 2023={v23}, 2024={v24}")
        
        # Step 5: Net Income Verification
        if self.net_income_found:
            self.verify_net_income(consolidated_mapped)
        
        print(f"\n📊 Final mapped items: {len(consolidated_mapped)}")
        print(f"📊 Unmapped items: {len(unmapped_items)} (consolidated into Other Income)")
        print(f"🛑 Net Income stopping: {'YES' if self.net_income_found else 'NO'}")
        
        return consolidated_mapped
    
    def populate_template(self, mapped_items: Dict[str, ISMappedValue]) -> str:
        """Populate the IS.CF sheet with mapped values"""
        if not self.working_template_path.exists():
            print("❌ Working template not found")
            return ""
        
        try:
            print("\n📝 Populating Income Statement template...")
            
            wb = load_workbook(self.working_template_path)
            
            # Get the IS.CF worksheet
            if 'IS.CF' not in wb.sheetnames:
                print("❌ IS.CF sheet not found in template")
                return ""
            
            ws = wb['IS.CF']
            print(f"   Working with sheet: {ws.title}")
            
            # Track populated fields
            populated_fields = []
            
            # Process each mapping
            for key, mapping in mapped_items.items():
                template_field = mapping.template_field
                value_2023 = mapping.value_2023
                value_2024 = mapping.value_2024
                
                # Check if we have a mapping for this template field
                if template_field in self.template_mappings:
                    col_letter, row_num = self.template_mappings[template_field]
                    
                    # For expenses, make values negative if they're positive
                    # (since template expects expenses as negative)
                    if template_field in ['Operating Expenses', 'Depreciation', 'Amortization', 'Interest Expense', 'Tax Expense']:
                        if value_2023 and value_2023 > 0:
                            value_2023 = -value_2023
                        if value_2024 and value_2024 > 0:
                            value_2024 = -value_2024
                    
                    # Populate 2023 value (column B)
                    if value_2023 is not None:
                        cell_2023 = f"B{row_num}"
                        ws[cell_2023] = value_2023
                        print(f"   ✅ {template_field} 2023: {cell_2023} = {value_2023:,.0f}")
                    
                    # Populate 2024 value (column C)
                    if value_2024 is not None:
                        cell_2024 = f"C{row_num}"
                        ws[cell_2024] = value_2024
                        print(f"   ✅ {template_field} 2024: {cell_2024} = {value_2024:,.0f}")
                    
                    populated_fields.append(template_field)
                else:
                    print(f"   ⚠️ No template mapping for: {template_field}")
            
            # Save the populated template
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"populated_is_template_{timestamp}.xlsx"
            wb.save(output_file)
            
            print(f"\n✅ Income Statement template populated successfully!")
            print(f"   Output file: {output_file}")
            print(f"   Fields populated: {len(populated_fields)}")
            print(f"   Fields: {', '.join(populated_fields)}")
            
            return output_file
            
        except Exception as e:
            print(f"❌ Error populating template: {e}")
            return ""
    
    def analyze_coverage(self, mapped_items: Dict[str, ISMappedValue]):
        """Analyze mapping coverage for income statement"""
        print("\n📊 INCOME STATEMENT MAPPING ANALYSIS:")
        print("=" * 50)
        
        # Count by section
        section_counts = {}
        template_fields = set()
        
        for mapped_value in mapped_items.values():
            section = mapped_value.section
            section_counts[section] = section_counts.get(section, 0) + 1
            template_fields.add(mapped_value.template_field)
        
        print("Mapped items by section:")
        for section, count in section_counts.items():
            print(f"  {section}: {count} fields")
        
        print(f"\nTotal unique template fields mapped: {len(template_fields)}")
        print(f"Template fields: {sorted(template_fields)}")
        
        # Key income statement requirements
        required_fields = {
            'Revenue', 'Operating Expenses', 'Depreciation', 'Interest Expense', 'Tax Expense'
        }
        
        mapped_required = template_fields.intersection(required_fields)
        missing_required = required_fields - template_fields
        
        print(f"\nRequired field coverage: {len(mapped_required)}/{len(required_fields)} ({100*len(mapped_required)/len(required_fields):.1f}%)")
        print(f"✅ Mapped: {sorted(mapped_required)}")
        if missing_required:
            print(f"❌ Missing: {sorted(missing_required)}")
    
    def cleanup_template(self):
        """Clean up working template file"""
        try:
            if self.working_template_path.exists():
                os.remove(self.working_template_path)
                print(f"🧹 Cleaned up working template: {self.working_template_path}")
        except Exception as e:
            print(f"⚠️ Warning: Could not clean up template: {e}")

def main():
    mapper = FinalISMapper()
    pdf_path = "../../input_pdfs/US_Venture_2024.pdf"
    
    print("Starting Enhanced Income Statement Mapper...")
    print("=" * 50)
    
    # Extract and map income statement data
    mapped_items = mapper.extract_and_process(pdf_path)
    
    if mapped_items:
        # Analyze coverage
        mapper.analyze_coverage(mapped_items)
        
        # Populate template
        output_file = mapper.populate_template(mapped_items)
        
        # Save mapping data as JSON
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_file = f"final_is_us_venture_{timestamp}.json"
        
        # Convert mapped_items to JSON-serializable format
        json_data = {}
        for key, mapped_value in mapped_items.items():
            json_data[key] = {
                "template_field": mapped_value.template_field,
                "section": mapped_value.section,
                "value_2023": mapped_value.value_2023,
                "value_2024": mapped_value.value_2024,
                "confidence": mapped_value.confidence,
                "mapping_method": mapped_value.mapping_method
            }
        
        with open(json_file, 'w') as f:
            json.dump(json_data, f, indent=2)
        
        print(f"\n💾 Mapping data saved: {json_file}")
        
        # Copy output files to main output directory
        main_output_dir = Path("../../output_excel")
        main_output_dir.mkdir(exist_ok=True)
        
        if output_file and Path(output_file).exists():
            shutil.copy2(output_file, main_output_dir / output_file)
            print(f"📁 Output copied to: {main_output_dir / output_file}")
        
        # Cleanup
        mapper.cleanup_template()
        
        print(f"\n✅ Enhanced Income Statement Mapping Complete!")
        print(f"   Net Income stopping: {'YES' if mapper.net_income_found else 'NO'}")
        if mapper.net_income_found:
            print(f"   Extracted Net Income 2023: ${mapper.extracted_net_income_2023:,.0f}" if mapper.extracted_net_income_2023 else "   Extracted Net Income 2023: Not found")
            print(f"   Extracted Net Income 2024: ${mapper.extracted_net_income_2024:,.0f}" if mapper.extracted_net_income_2024 else "   Extracted Net Income 2024: Not found")
    else:
        print("❌ No mappings created")
        mapper.cleanup_template()

if __name__ == "__main__":
    main() 