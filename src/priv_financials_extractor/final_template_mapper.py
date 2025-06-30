'''
Line Items to be filled for 2 years worth of data:

ASSETS
Cash and equivalents
Accounts Receivable
Prepaid Expenses
Inventory
Investments
Other
Total Current Assets

Net PPE
Goodwill
Intangibles
Other

Total Non Current Assets
Total Assets

LIABILITIES
Accounts Payable
Accrued Interest
Short term Borrowing
Current Portion of Long Term Debt
Other
Total Current Liabilities

Long Term Debt
Deferred income taxes
Other

Total Non Current Liabilities
Total Liabilities

EQUITY
Common Stock
Retained Earnings
Paid in Capital
Other
Total Equity

Total Liabilities and Equity

INCOME STATEMENT
Revenue
Operating Expenses
Operating Income

Depreciation (-)
Amortization (-)
Assets gain(loss) impairments
Interest Expense (-)
Interest Income (+)
Other income(expenses)

Income Before Taxes
Tax expense
Other
Net Income

STATEMENT OF CASH FLOW
Operating Activities
Net Income
Changes in noncash items
Changes in Assets and Liabilities
Net Cash from(used) Operating Activities

Investing Activities
CapEx
Proceeds from asset sales
Others
Net cash from(used) for investing 

Financing Activities
Issuance of Debt (long+short term)
Retirement of Debt (long+short term)
Issuance of Stock
Dividends Paid
Other
Net cash from(used) for financing

Net change in Cash
Starting Cash
Ending Cash

'''

import pandas as pd
import numpy as np
from pathlib import Path
import torch
from transformers import BertTokenizer, BertForSequenceClassification
import logging
from typing import Dict, List, Tuple, Optional
import re
import pdfplumber
from sentence_transformers import SentenceTransformer
import shutil
from openpyxl import load_workbook
from collections import defaultdict
import openpyxl
from datetime import datetime
from llm_mapper import LLMMapper  # Import the LLM mapper
from sklearn.metrics.pairwise import cosine_similarity
import sys
import requests

# Place these at the module level, outside of any function
IS_SECTION_TEMPLATE = {
    'revenue': {
        'template': [
            'Revenue', 'Other Revenue'
        ]
    },
    'operating_expenses': {
        'template': [
            'Cost of revenue', 'Station operations costs', 'Payroll and related costs',
            'Depreciation and amortization', 'Impairment and other losses', 'Selling, general and administrative expenses', 'Other Operating Expenses'
        ]
    },
    'other_income_expense': {
        'template': [
            'Interest expense', 'Other (income) and expense, net', 'Other Income/Expense'
        ]
    },
    'tax_net_income': {
        'template': [
            'Income tax benefit', 'Net profit (loss)']
    }
}

CFS_SECTION_TEMPLATE = {
    'operating_activities': {
        'template': [
            'Net profit (loss)', 'Adjustments to reconcile net profit', 'Depreciation', 'Deferred income taxes', 'Impairment and other losses', 'Changes in operating assets and liabilities', 'Net cash provided by (used in) operating activities', 'Other Operating Activities'
        ]
    },
    'investing_activities': {
        'template': [
            'Purchases of property and equipment', 'Proceeds from sale of assets', 'Net cash used in investing activities', 'Other Investing Activities'
        ]
    },
    'financing_activities': {
        'template': [
            'Proceeds from issuance', 'Principal payments', 'Net cash provided by (used in) financing activities', 'Other Financing Activities'
        ]
    },
    'other': {
        'template': ['Other']
    },
    'cash_reconciliation': {
        'template': [
            'Net change in Cash',
            'Starting Cash',
            'Ending Cash'
        ]
    }
}

# Expand manual mapping for IS and CFS
manual_section_map = {}
manual_section_map.update({
    # IS Revenue
    'revenue net': 'revenue',
    'revenue': 'revenue',
    'other revenue': 'revenue',
    # IS Operating Expenses
    'cost of revenue': 'operating_expenses',
    'station operations costs': 'operating_expenses',
    'payroll and related costs': 'operating_expenses',
    'depreciation and amortization': 'operating_expenses',
    'impairment and other losses': 'operating_expenses',
    'selling general and administrative expenses': 'operating_expenses',
    'other operating expenses': 'operating_expenses',
    # IS Other Income/Expense
    'interest expense': 'other_income_expense',
    'other income and expense net': 'other_income_expense',
    'other income expense': 'other_income_expense',
    # IS Tax/Net Income
    'income tax benefit': 'tax_net_income',
    'net profit loss': 'tax_net_income',
    # CFS Operating Activities
    'net profit loss': 'operating_activities',
    'adjustments to reconcile net profit': 'operating_activities',
    'depreciation': 'operating_activities',
    'deferred income taxes': 'operating_activities',
    'impairment and other losses': 'operating_activities',
    'changes in operating assets and liabilities': 'operating_activities',
    'net cash provided by used in operating activities': 'operating_activities',
    'other operating activities': 'operating_activities',
    # CFS Investing Activities
    'purchases of property and equipment': 'investing_activities',
    'proceeds from sale of assets': 'investing_activities',
    'net cash used in investing activities': 'investing_activities',
    'other investing activities': 'investing_activities',
    # CFS Financing Activities
    'proceeds from issuance': 'financing_activities',
    'principal payments': 'financing_activities',
    'net cash provided by used in financing activities': 'financing_activities',
    'other financing activities': 'financing_activities',
})

# Add section boundary keywords for IS and CFS
section_boundary_keywords = {}
section_boundary_keywords['is_section_boundaries'] = [
    'revenue', 'cost of revenue', 'station operations costs', 'payroll and related costs',
    'depreciation and amortization', 'impairment and other losses', 'selling general and administrative expenses',
    'income from operations', 'interest expense', 'other income and expense net', 'income tax benefit', 'net profit loss'
]
section_boundary_keywords['cfs_section_boundaries'] = [
    'cash flows from operating activities', 'cash flows from investing activities', 'cash flows from financing activities',
    'net cash provided by used in operating activities', 'net cash used in investing activities', 'net cash provided by used in financing activities'
]

class TemplateMatcher:
    TOTAL_NET_PATTERNS = [
        r'^total\s+current\s+assets?$',
        r'^total\s+non[- ]?current\s+assets?$',
        r'^total\s+assets?$',
        r'^total\s+current\s+liabilities?$',
        r'^total\s+non[- ]?current\s+liabilities?$',
        r'^total\s+liabilities?$',
        r'^total\s+equity$',
        r'^net\s+cash\s+(?:from|used\s+in|provided\s+by)',
        r'^total\s+(?:current|non[- ]?current)\s+(?:assets?|liabilities?)$',
        r'^total$',
        r'^total\s+shareholders?\s+equity$',
        r'^total\s+stockholders?\s+equity$',
        r'^total\s+debt$',
        r'^total\s+long[- ]term\s+debt$',
        r'^total\s+current\s+debt$',
        r'^total\s+investments?$',
        r'^total\s+property\s+and\s+equipment$',
        r'^total\s+ppe$',
        r'^total\s+intangible\s+assets?$',
        r'^total\s+goodwill$',
        r'^total\s+accounts?\s+receivable$',
        r'^total\s+accounts?\s+payable$',
        r'^total\s+accrued\s+liabilities?$',
        r'^total\s+inventory$',
        r'^total\s+cash$',
        r'^total\s+cash\s+and\s+cash\s+equivalents$',
        r'^total\s+current\s+portion\s+of\s+long[- ]term\s+debt$',
        r'^total\s+long[- ]term\s+incentive$',
        r'^total\s+deferred\s+compensation$',
        r'^total\s+finance\s+lease\s+liability$',
        r'^total\s+operating\s+lease\s+liability$',
        r'^total\s+deferred\s+income\s+taxes?$',
        r'^total\s+other\s+noncurrent\s+liabilities?$',
        r'total\s+other\s+current\s+liabilities?$',
        r'^total\s+other\s+current\s+assets?$',
        r'^total\s+other\s+noncurrent\s+assets?$',
        r'^total\s+noncontrolling\s+interests?$',
        r'^total\s+common\s+stock$',
        r'^total\s+retained\s+earnings?$',
        r'^total\s+paid[- ]in\s+capital$',
    ]

    def __init__(self, model_name="all-MiniLM-L6-v2", use_llm_fallback=True):
        """Initialize with Sentence Transformers model"""
        self.model = SentenceTransformer(model_name)
        self.setup_logging()
        self.used_items = set()  # Track used items globally
        self.extraction_logger = self.setup_extraction_logger()
        
        # Initialize LLM mapper for hybrid approach
        self.llm_mapper = LLMMapper()
        self.use_llm = use_llm_fallback and self.llm_mapper.check_ollama_available()
        if self.use_llm:
            print("✅ LLM mapper available - using hybrid approach")
        else:
            if not use_llm_fallback:
                print("⚡ LLM disabled - using fast rule-based approach only")
            print("   Install Ollama: https://ollama.ai/")
            print("   Run: ollama pull mistral")
        
    def setup_logging(self):
        """Setup logging configuration"""
        logging.basicConfig(
            filename='template_mapping.log',
            level=logging.INFO,
            format='%(message)s',
            filemode='w'
        )

    def setup_extraction_logger(self):
        extraction_logger = logging.getLogger("extraction_logger")
        extraction_logger.setLevel(logging.DEBUG)
        # Remove any existing handlers
        if extraction_logger.hasHandlers():
            extraction_logger.handlers.clear()
        # Set up extraction.log in project root
        import os
        from pathlib import Path
        current_dir = Path(__file__).resolve().parent
        project_root = current_dir.parent.parent
        extraction_log_file = project_root / "extraction.log"
        handler = logging.FileHandler(extraction_log_file, mode='w', encoding='utf-8')
        formatter = logging.Formatter('%(message)s')
        handler.setFormatter(formatter)
        extraction_logger.addHandler(handler)
        return extraction_logger

    def get_similarity(self, text1: str, text2: str) -> float:
        """Get semantic similarity between two texts using sentence-transformers"""
        embeddings = self.model.encode([text1, text2])
        return np.dot(embeddings[0], embeddings[1]) / (np.linalg.norm(embeddings[0]) * np.linalg.norm(embeddings[1]))

    def find_best_match(self, source_item: str, target_items: List[str], section_context: str, threshold: float = 0.5) -> Tuple[Optional[str], float]:
        """Find best matching template item for a given source item within a section context"""
        best_match = None
        best_score = -1
        
        # Normalize source item
        source_item = source_item.lower().strip()
        
        # Direct match patterns for common items
        direct_matches = {
            'Cash and equivalents': [
                r'cash(?:\s+and\s+(?:cash\s+)?equivalents?)?',
                r'cash\s+equivalents?'
            ],
            'Accounts Receivable': [
                r'accounts?\s+receivable',
                r'(?:trade|net)\s+receivables?'
            ],
            'Inventory': [
                r'inventor(?:y|ies)(?:\s*[-—]\s*net)?',
                r'net\s+inventor(?:y|ies)'
            ],
            'Net PPE': [
                r'property(?:\s+and\s+)?equipment(?:\s*[-—]\s*net)?',
                r'(?:net\s+)?(?:ppe|property,?\s+plant\s+and\s+equipment)',
                r'right\s+of\s+use\s+assets?',
                r'finance\s+lease\s+assets?'
            ],
            'Goodwill': [
                r'goodwill(?:\s*[-—]\s*net)?',
                r'(?:net\s+)?goodwill'
            ]
        }
        
        # Check for direct matches first
        for target in target_items:
            if target in direct_matches:
                for pattern in direct_matches[target]:
                    if re.search(pattern, source_item):
                        return target, 1.0
        
        # Section-specific term mappings
        section_terms = {
            'assets': [
                'cash', 'receivable', 'inventory', 'prepaid', 'investment', 
                'property', 'equipment', 'ppe', 'goodwill', 'intangible',
                'lease', 'right of use', 'margin deposit', 'derivative'
            ],
            'liabilities': [
                'payable', 'debt', 'borrowing', 'loan', 'accrued', 'deferred',
                'liability', 'lease', 'obligation', 'tax'
            ],
            'equity': [
                'stock', 'capital', 'earning', 'dividend', 'share', 'equity',
                'retained', 'paid-in', 'comprehensive'
            ]
        }
        
        # Only proceed if item matches section context
        if section_context in section_terms:
            matches_section = any(term in source_item for term in section_terms[section_context])
            if not matches_section:
                return None, 0
        
        for target in target_items:
            target = target.lower().strip()
            base_score = self.get_similarity(source_item, target)
            
            # Add bonus for matching section context
            if section_context in section_terms:
                if any(term in target for term in section_terms[section_context]):
                    base_score += 0.2
            
            # Add bonus for exact word matches
            source_words = set(re.findall(r'\b\w+\b', source_item))
            target_words = set(re.findall(r'\b\w+\b', target))
            common_words = source_words.intersection(target_words)
            if common_words:
                base_score += 0.1 * len(common_words)
            
            # Add bonus for matching parenthetical terms
            source_parens = re.findall(r'\((.*?)\)', source_item)
            target_parens = re.findall(r'\((.*?)\)', target)
            if source_parens and target_parens:
                if any(s.strip() in [t.strip() for t in target_parens] for s in source_parens):
                    base_score += 0.1
            
            # Add bonus for matching numerical indicators
            if ('current' in source_item and 'current' in target) or \
               ('long term' in source_item and 'long term' in target) or \
               ('short term' in source_item and 'short term' in target):
                base_score += 0.15
            
            # Penalize matching to "Other" unless it's really a good match
            if 'other' in target and 'other' not in source_item:
                base_score *= 0.7
            
            if base_score > best_score and base_score >= threshold:
                best_score = base_score
                best_match = target
        
        return best_match, best_score

    def _map_section(self, sheet, data: Dict, col: str, start_row: int, end_row: int, template_items: List[str], section_context: str):
        """Map a section of data to the template"""
        # Track matched items to handle "Other" categories
        matched_items = set()
        other_values = 0.0
        
        # First pass: direct matches with improved semantic matching
        for item, value in data.items():
            # Skip if already used or if it's a total line
            if item in self.used_items or any(total_word in item.lower() for total_word in ['total', 'sum']):
                continue
            
            # Normalize value
            try:
                if isinstance(value, str):
                    value = float(value.replace(',', ''))
                else:
                    value = float(value)
            except (ValueError, TypeError):
                logging.warning(f"Could not convert value {value} to float for {item}")
                continue
            
            best_match, score = self.find_best_match(item, template_items, section_context)
            if best_match:
                # Find row for this item
                for row in range(start_row, end_row + 1):
                    template_item = sheet[f'A{row}'].value
                    if template_item and template_item.strip() == best_match:
                        sheet[f"{col}{row}"] = value
                        matched_items.add(item)
                        self.used_items.add(item)  # Mark as used globally
                        logging.info(f"Matched {item} to {best_match} with score {score}")
                        break
        
        # Second pass: try to match remaining items with more context
        for item, value in data.items():
            if item in self.used_items or item in matched_items:
                continue
            
            # Skip totals
            if any(total_word in item.lower() for total_word in ['total', 'sum']):
                continue
            
            try:
                if isinstance(value, str):
                    value = float(value.replace(',', ''))
                else:
                    value = float(value)
                
                # Try to match based on context
                context_score = 0
                best_match = None
                
                for template_item in template_items:
                    # Check for semantic similarity with section context
                    score = self.get_similarity(item.lower(), template_item.lower())
                    if section_context:
                        score = self.find_best_match(item, [template_item], section_context)[1]
                    
                    if score > context_score:
                        context_score = score
                        best_match = template_item
                
                if best_match and context_score >= 0.5:
                    # Find row for this item
                    for row in range(start_row, end_row + 1):
                        template_item = sheet[f'A{row}'].value
                        if template_item and template_item.strip() == best_match:
                            sheet[f"{col}{row}"] = value
                            matched_items.add(item)
                            self.used_items.add(item)  # Mark as used globally
                            logging.info(f"Context matched {item} to {best_match} with score {context_score}")
                            break
                else:
                    # Only add to Other if it matches section context
                    if self.find_best_match(item, ['Other'], section_context)[1] >= 0.3:
                        other_values += value
                        self.used_items.add(item)  # Mark as used globally
                        logging.info(f"Adding {item} ({value}) to Other category")
            except (ValueError, TypeError):
                continue
        
        # Find and populate "Other" row if we have unmatched items
        if other_values != 0:
            found_other = False
            for row in range(start_row, end_row + 1):
                template_item = sheet[f'A{row}'].value
                print(f"[DEBUG] Checking row {row} for 'Other': '{template_item}'")
                if template_item and template_item.strip() == 'Other':
                    print(f"[DEBUG] Writing {other_values} to {section_context}::Other at cell {col}{row}")
                    sheet[f"{col}{row}"] = other_values
                    found_other = True
                    break
            if not found_other:
                print(f"[WARNING] Could not find 'Other' row in {section_context} section (rows {start_row}-{end_row}) to write value {other_values}")
        
        # Calculate totals for sections that have them
        if end_row in [13, 20, 29, 35, 43]:  # Total rows
            total = 0
            for row in range(start_row, end_row):
                val = sheet[f"{col}{row}"].value
                try:
                    total += float(val) if val not in [None, ''] else 0
                except (ValueError, TypeError):
                    continue
            sheet[f"{col}{end_row}"] = total

    def analyze_subsections(self, item: str) -> str:
        """Determine which subsection an item belongs to based on its content"""
        item = item.lower().strip()
        
        # Define subsection patterns
        subsections = {
            'Current Assets': [
                r'cash', r'equivalent', r'margin\s+deposit', r'derivative\s+asset',
                r'(?:current\s+)?(?:account|trade)\s+receivable', r'current\s+portion',
                r'inventor(?:y|ies)', r'prepaid', r'current\s+asset'
            ],
            'Non-Current Assets': [
                r'(?:property|equipment|ppe)(?:\s*[-—]\s*net)?', r'right\s+of\s+use',
                r'finance\s+lease', r'goodwill', r'intangible', r'non-?current\s+asset',
                r'deferred\s+(?:tax|compensation)', r'long-term', r'investment'
            ],
            'Current Liabilities': [
                r'current\s+liabilit(?:y|ies)', r'accounts?\s+payable',
                r'accrued', r'short[- ]term', r'current\s+portion'
            ],
            'Non-Current Liabilities': [
                r'long[- ]term\s+(?:debt|lease|liability)',
                r'deferred\s+(?:tax|revenue|income)',
                r'non-?current\s+liabilit(?:y|ies)'
            ],
            'Equity': [
                r'(?:common|preferred)\s+stock', r'retained\s+earnings?',
                r'paid[- ]in\s+capital', r'shareholder', r'equity'
            ],
            'Revenue': [
                r'revenue', r'sales', r'income\s+from\s+operations'
            ],
            'Operating Expenses': [
                r'cost\s+of\s+(?:goods\s+sold|revenue|sales)',
                r'operating\s+expense', r'selling', r'administrative',
                r'depreciation', r'amortization'
            ],
            'Other Income/Expense': [
                r'interest\s+(?:income|expense)',
                r'other\s+(?:income|expense)',
                r'gain|loss'
            ],
            'Operating Activities': [
                r'operating\s+activit(?:y|ies)',
                r'cash\s+from\s+operations?',
                r'working\s+capital'
            ],
            'Investing Activities': [
                r'investing\s+activit(?:y|ies)',
                r'capital\s+expenditure',
                r'acquisition',
                r'purchase\s+of'
            ],
            'Financing Activities': [
                r'financing\s+activit(?:y|ies)',
                r'dividend',
                r'stock\s+(?:issue|repurchase)',
                r'debt\s+(?:issue|repayment)'
            ]
        }
        
        # Check each subsection's patterns
        for subsection, patterns in subsections.items():
            for pattern in patterns:
                if re.search(pattern, item):
                    return subsection
        
        return "Uncategorized"

    def print_categorization(self, extracted_data: Dict):
        """Print categorization of all extracted values"""
        print("\nCategorization of Extracted Values:")
        print("=" * 80)
        
        for statement_type, years in extracted_data.items():
            print(f"\n{statement_type.upper()}")
            print("-" * 80)
            
            for year, items in years.items():
                print(f"\n{year}:")
                
                # Group items by subsection
                categorized = defaultdict(list)
                for item, value in items.items():
                    if not any(total_word in item.lower() for total_word in ['total', 'sum']):
                        subsection = self.analyze_subsections(item)
                        categorized[subsection].append((item, value))
                
                # Print categorized items
                for subsection in sorted(categorized.keys()):
                    print(f"\n  {subsection}:")
                    for item, value in sorted(categorized[subsection]):
                        print(f"    {item}: {value}")

    def assign_sections_by_context(self, extracted_lines: list) -> list:
        """
        Assigns BS sections to each line item based on content and explicit headers, using analyze_subsections for fallback.
        """
        assigned = []
        current_section = None
        
        # Keywords to identify major section headers
        section_headers = {
            'current_assets': ['current assets'],
            'noncurrent_assets': ['noncurrent assets', 'non-current assets'],
            'current_liabilities': ['current liabilities'],
            'noncurrent_liabilities': ['noncurrent liabilities', 'non-current liabilities'],
            'equity': ['equity', "stockholders' equity", "shareholders' equity"]
        }

        # Keywords for specific items that can also define a section
        manual_map = {
            'cash and cash equivalents': 'current_assets',
            'accounts receivable': 'current_assets',
            'inventories': 'current_assets',
            'property, plant and equipment': 'noncurrent_assets',
            'accounts payable': 'current_liabilities',
            'retained earnings': 'equity',
            'additional paid in capital': 'equity',
            'common stock': 'equity'
        }
        
        for line in extracted_lines:
            desc = line['description']
            desc_lower = desc.lower()
            
            # 1. Check for major section headers first
            found_header = False
            for section, keywords in section_headers.items():
                if any(keyword in desc_lower for keyword in keywords):
                    current_section = section
                    found_header = True
                    break
            
            # 2. If no header, check for specific item keywords that imply a section change
            if not found_header:
                for keyword, section in manual_map.items():
                    if keyword in desc_lower:
                        current_section = section
                        found_header = True
                                break
            
            # 3. If still not found, use analyze_subsections for fallback
            if not found_header or not current_section:
                subsection = self.analyze_subsections(desc)
                # Map subsection to section
                subsection_map = {
                    'Current Assets': 'current_assets',
                    'Non-Current Assets': 'noncurrent_assets',
                    'Current Liabilities': 'current_liabilities',
                    'Non-Current Liabilities': 'noncurrent_liabilities',
                    'Equity': 'equity'
                }
                if subsection in subsection_map:
                    current_section = subsection_map[subsection]
            
            assigned.append({
                'description': desc,
                'value': line.get('numbers', []),
                'section': current_section
            })
        
        assigned_lines = self.smooth_section_assignments(assigned)
        return assigned_lines

    def smooth_section_assignments(self, assigned_lines):
        """
        Post-process assigned_lines to correct 'sandwiched' outliers:
        If a line's neighbors (i-1, i+1) have the same section and it is different, reassign it and log.
        """
        for i in range(1, len(assigned_lines) - 1):
            prev_section = assigned_lines[i-1].get('section')
            next_section = assigned_lines[i+1].get('section')
            curr_section = assigned_lines[i].get('section')
            if prev_section == next_section and curr_section != prev_section:
                old_section = assigned_lines[i]['section']
                assigned_lines[i]['section'] = prev_section
                desc = assigned_lines[i].get('description', '')
                print(f"[SMOOTH] Corrected section for line {i}: '{desc}' from '{old_section}' to '{prev_section}'")
        return assigned_lines

    def filter_out_totals_for_llm(self, bs_lines: List[dict]) -> Tuple[List[dict], List[dict]]:
        """
        Filter out total/subtotal items before sending to LLM for section assignment.
        Returns (non_total_items, total_items).
        """
        non_total_items = []
        total_items = []
        
        for line in bs_lines:
            description = line.get('description', '')
            if self.is_total_or_net_row(description):
                total_items.append(line)
            else:
                non_total_items.append(line)
        
        if total_items:
            print(f"[DEBUG] Filtered out {len(total_items)} total/subtotal items before LLM assignment:")
            for item in total_items:
                print(f"  - {item.get('description', '')}")
        
        return non_total_items, total_items

    def convert_llm_section_to_template_section(self, llm_section: str) -> str:
        """Convert LLM section names to template section names."""
        section_mapping = {
            'Current Assets': 'current_assets',
            'current_assets': 'current_assets',
            'Non-Current Assets': 'noncurrent_assets',
            'noncurrent_assets': 'noncurrent_assets',
            'Noncurrent Assets': 'noncurrent_assets',
            'Current Liabilities': 'current_liabilities',
            'current_liabilities': 'current_liabilities',
            'Non-Current Liabilities': 'noncurrent_liabilities',
            'noncurrent_liabilities': 'noncurrent_liabilities',
            'Noncurrent Liabilities': 'noncurrent_liabilities',
            'Equity': 'equity',
            'equity': 'equity',
            'Uncategorized': None  # Will be handled by fallback logic
        }
        return section_mapping.get(llm_section, None)

    def assign_sections_with_llm(self, bs_lines: List[dict]) -> List[dict]:
        """
        Assigns BS sections using LLM in batches of 3 (chunked) for all items.
        Pre-filters totals/subtotals before LLM assignment.
        """
        print("\n[DEBUG] Assigning sections using LLM (batch mode, chunked)...")
        print(f"[DEBUG] Total items to process: {len(bs_lines)}")
        assigned_lines = []
        if not self.use_llm:
            print("[INFO] LLM not available, falling back to rule-based section assignment.")
            return self.assign_sections_by_context(bs_lines)

        # Pre-filter totals/subtotals
        non_total_items, total_items = self.filter_out_totals_for_llm(bs_lines)
        print(f"[DEBUG] After filtering: {len(non_total_items)} non-total items, {len(total_items)} total items")
        
        chunk_size = 3
        all_assignments = {}
        
        # Process non-total items with LLM
        for i in range(0, len(non_total_items), chunk_size):
            chunk = non_total_items[i:i + chunk_size]
            descriptions = [line['description'] for line in chunk]
            
            print(f"\n[DEBUG] Processing chunk {i//chunk_size + 1}: {descriptions}")
            
            # Get LLM assignments for this chunk
            chunk_assignments = self.llm_mapper.assign_sections_batch_with_llm(descriptions)
            
            if chunk_assignments:
                print(f"[DEBUG] LLM batch assignment successful: {chunk_assignments}")
                all_assignments.update(chunk_assignments)
            else:
                print(f"[WARN] LLM failed for chunk: {descriptions}. Falling back to rule-based for this chunk.")
                # Fall back to rule-based for this chunk
                for line in chunk:
                    section = self.analyze_subsections(line['description'])
                    all_assignments[line['description']] = section

        # Process total items (assign to None or skip)
        for line in total_items:
            all_assignments[line['description']] = None

        print(f"\n[DEBUG] Final assignments summary:")
        print(f"[DEBUG] Total assignments: {len(all_assignments)}")
        for desc, section in all_assignments.items():
            print(f"[DEBUG] '{desc}' -> {section}")

        # Assign sections to all items
        for line in bs_lines:
            desc = line['description']
            section = None
            source = "rule-based fallback"
            
            # Try to find exact match in LLM assignments
            if desc in all_assignments:
                llm_section = all_assignments[desc]
                # Convert LLM section name to template section name
                section = self.convert_llm_section_to_template_section(llm_section)
                source = "LLM"
            else:
                # Try fuzzy matching for minor differences
                for llm_desc, llm_section in all_assignments.items():
                    if self.get_similarity(desc, llm_desc) > 0.8:
                        section = self.convert_llm_section_to_template_section(llm_section)
                        source = f"LLM (fuzzy match: \"{llm_desc}\")"
                        break
                
                if section is None:
                    # Fall back to rule-based
                    subsection = self.analyze_subsections(desc)
                    # Map subsection to template section
                    subsection_map = {
                        'Current Assets': 'current_assets',
                        'Non-Current Assets': 'noncurrent_assets',
                        'Current Liabilities': 'current_liabilities',
                        'Non-Current Liabilities': 'noncurrent_liabilities',
                        'Equity': 'equity'
                    }
                    section = subsection_map.get(subsection, None)
                    source = "rule-based fallback"
            
            # Create assigned line with proper value handling
            assigned_line = {
                'description': desc,
                'idx': line.get('idx', 0),
                'section': section,
                'value': line.get('numbers', line.get('value', []))
            }
            assigned_lines.append(assigned_line)
            
            print(f"[SECTION ASSIGN] '{desc}' -> {section} [source: {source}]")

        assigned_lines = self.smooth_section_assignments(assigned_lines)
        return assigned_lines

    def assign_sections_by_context_is(self, extracted_lines: list) -> list:
        """
        Assigns IS sections to each line item based on its content and context.
        Uses explicit header detection to improve accuracy.
        """
        assigned = []
        current_section = None # Start with no section
        
        # Section headers to look for (case-insensitive)
        section_keywords = {
            'revenue': ['net sales', 'revenue'],
            'operating_expenses': ['operating costs and expenses', 'operating expenses'],
            'operating_income': ['operating income'],
            'other_income_expense': ['other income (expense)', 'other income'],
            'tax_net_income': ['net income', 'income before', 'comprehensive income']
        }

        for line in extracted_lines:
            desc_lower = line['description'].lower()
            
            # --- 1. Detect section based on headers ---
            found_section = False
            for section, keywords in section_keywords.items():
                if any(re.search(r'\b' + keyword + r'\b', desc_lower) for keyword in keywords):
                    current_section = section
                    found_section = True
                    break
            
            # --- 2. If no header, check if it's a known non-op item ---
            if not found_section:
                # This part is simplified; can be expanded with more keywords
                if 'interest' in desc_lower:
                    current_section = 'other_income_expense'

            # --- 3. Assign section ---
                        assigned_section = current_section
            
            assigned.append({
                'description': line['description'],
                'value': line.get('numbers', line.get('value', [])),  # Handle both 'numbers' and 'value' keys
                'section': assigned_section
            })
        
        assigned_lines = self.smooth_section_assignments(assigned)
        return assigned_lines

    def assign_sections_by_context_cfs(self, extracted_lines: list) -> list:
        """
        Assigns CFS sections to each line item using a more robust, keyword-based state machine.
        This correctly handles the final cash reconciliation section.
        """
        assigned = []
        current_section = None  # Start with no section

        # Define keywords that signal the start of a new section
        section_keywords = {
            'operating_activities': [
                'cash flow from operating activities',
                'cash flows from operating activities',
                'operating activities'
            ],
            'investing_activities': [
                'cash flow from investing activities',
                'cash flows from investing activities',
                'investing activities'
            ],
            'financing_activities': [
                'cash flow from financing activities',
                'cash flows from financing activities',
                'financing activities'
            ],
            'cash_reconciliation': [
                'net change in cash',
                'net increase in cash',
                'net decrease in cash',
                'effect of exchange rate',
                'cash at beginning',
                'cash and cash equivalents at beginning',
                'supplemental disclosure' # Often marks the end of main sections
            ]
        }
        
        for line in extracted_lines:
            desc_lower = line['description'].lower().strip()
            
            # Check if the line description indicates a new section
            found_new_section = False
            for section, keywords in section_keywords.items():
                if any(keyword in desc_lower for keyword in keywords):
                    current_section = section
                    self.extraction_logger.info(f"[SECTION SWITCH] '{line['description']}' -> {current_section}")
                    found_new_section = True
                    break
            
            # If no section is ever found, we can try a fallback
            if current_section is None:
                if 'net income' in desc_lower or 'net loss' in desc_lower:
                    current_section = 'operating_activities'
            
            assigned.append({
                'description': line['description'],
                'value': line.get('numbers', []),
                'section': current_section
            })
            
        # Add the new section to the template definition for mapping
        if 'cash_reconciliation' not in CFS_SECTION_TEMPLATE:
             CFS_SECTION_TEMPLATE['cash_reconciliation'] = {
                 'template': [
                     'Net change in Cash',
                     'Starting Cash',
                     'Ending Cash'
                 ]
             }
        
        assigned_lines = self.smooth_section_assignments(assigned)
        return assigned_lines

    def get_is_row_map(self, worksheet, section: str) -> dict:
        """
        Returns a mapping of template item descriptions to their row number 
        for a given IS section by reading directly from the worksheet.
        This is more robust than using a pre-processed list.
        """
        row_map = {}
        # Correct, non-overlapping sections based on template visual layout
        section_rows = {
            'revenue': [(6, 6)],
            'operating_expenses': [(10, 12)], # Dep, Amort, Impair
            'other_income_expense': [(13, 15), (19, 19)], # Interest, Other Income, Other
            'tax_net_income': [(18, 18)] # Tax Expense only
        }
        
        # Rows that are headers or calculated totals, and should not be mapped to
        skip_rows = {
            7: "Operating Expenses", 
            8: "Operating Income", 
            17: "Income Before Taxes",
            20: "Net Income"
        }

        ranges = section_rows.get(section, [])
        if not ranges: return {}
        
        for start, end in ranges:
            for row_num in range(start, end + 1):
                if row_num in skip_rows:
                    continue
                item = worksheet.cell(row=row_num, column=1).value
                if item:
                    row_map[item] = row_num
        return row_map

    def get_cfs_row_map(self, template_items: list, section: str) -> dict:
        """
        Returns a mapping of template item descriptions to their row number for a given CFS section.
        """
        cfs_row_mapping = {
            'Net profit (loss)': 23,
            'Adjustments to reconcile net profit': 24,
            'Depreciation': 25,
            'Deferred income taxes': 26,
            'Impairment and other losses': 27,
            'Changes in operating assets and liabilities': 28,
            'Net cash provided by (used in) operating activities': 29,
            'Other Operating Activities': 30,
            'Purchases of property and equipment': 32,
            'Proceeds from sale of assets': 33,
            'Net cash used in investing activities': 34,
            'Other Investing Activities': 35,
            'Proceeds from issuance': 37,
            'Principal payments': 38,
            'Net cash provided by (used in) financing activities': 39,
            'Other Financing Activities': 40,
            'Net change in Cash': 42,
            'Starting Cash': 43,
            'Ending Cash': 44,
            'Other': 45
        }
        
        return {item: cfs_row_mapping.get(item) for item in template_items if item in cfs_row_mapping}

    def get_is_row_map_full(self, sheet):
        """Helper to get row map for the entire IS."""
        row_map = {}
        # Iterate over all defined IS sections and their templates
        for section_info in IS_SECTION_TEMPLATE.values():
            for item in section_info['template']:
                # Find this item in column A
                for row in range(1, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == item:
                        row_map[item] = row
                        break
        return row_map
        
    def get_cfs_row_map_full(self, sheet):
        """Helper to get row map for the entire CFS."""
        row_map = {}
        for section_info in CFS_SECTION_TEMPLATE.values():
            for item in section_info['template']:
                for row in range(1, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == item:
                        row_map[item] = row
                        break
        return row_map

    def get_other_category_for_is_section(self, section: str) -> Optional[str]:
        if section == 'operating_expenses': return 'Other Operating Expenses'
        if section == 'other_income_expense': return 'Other Income/Expense'
        return None

    def get_other_category_for_cfs_section(self, section: str) -> Optional[str]:
        if section == 'operating_activities': return 'Other Operating Activities'
        if section == 'investing_activities': return 'Other Investing Activities'
        if section == 'financing_activities': return 'Other Financing Activities'
        return None

    def print_section_assignments(self, assigned_lines: list, year: str, statement_type: str):
        print(f"\nSection assignments for {statement_type} {year}:")
        for entry in assigned_lines:
            print(f"  {entry['description']}  -->  {entry['section']}  (value: {entry['value']})")

    def is_total_or_net_row(self, description: str) -> bool:
        """Check if description is a total or net row (expanded and strict)."""
        desc_lower = description.lower().strip()
        # Expanded patterns for totals/subtotals
        total_patterns = [
            r'^total(\s|$)',
            r'(\s|^)total\s',
            r'(\s|^)sum(\s|$)',
            r'(\s|^)subtotal(\s|$)',
            r'(\s|^)net(\s|$)',
            r'(\s|^)aggregate(\s|$)',
            r'(\s|^)grand total(\s|$)',
            r'(\s|^)overall(\s|$)',
            r'(\s|^)balance(\s|$)',
            r'(\s|^)ending balance(\s|$)',
            r'(\s|^)beginning balance(\s|$)',
            r'(\s|^)ending cash(\s|$)',
            r'(\s|^)starting cash(\s|$)',
            r'(\s|^)net change(\s|$)',
            r'(\s|^)net increase(\s|$)',
            r'(\s|^)net decrease(\s|$)'
        ]
        for pat in total_patterns:
            if re.search(pat, desc_lower):
                return True
        # Also catch if the description is just a number or empty
        if desc_lower in ('', '-', 'n/a', 'na', 'none'):
            return True
        return False

    def apply_rule_based_mapping(self, description: str) -> tuple[str, float]:
        """Apply rule-based mapping for balance sheet items"""
        import re
        desc_lower = description.lower()
        
        # Balance sheet rule-based mapping - EXPANDED
        bs_rules = {
            r'cash\s+(?:and\s+)?(?:cash\s+)?equivalents?': 'Cash and equivalents',
            r'accounts?\s+receivable(?:[—-]net)?': 'Accounts Receivable',
            r'notes?\s+receivable': 'Accounts Receivable',  # Added
            r'prepaid\s+expenses?': 'Prepaid Expenses',
            r'inventor(?:y|ies)(?:[—-]net)?': 'Inventory',
            r'property\s+(?:and\s+)?equipment(?:[—-]net)?': 'Net PPE',
            r'property\s+(?:and\s+)?equipment\s+at\s+cost': 'Net PPE',  # Added
            r'less\s+accumulated\s+depreciation': 'Net PPE',  # Added
            r'accumulated\s+depreciation': 'Net PPE',  # Added
            r'net\s+ppe': 'Net PPE',
            r'goodwill(?:[—-]net)?': 'Goodwill',
            r'(?:other\s+)?intangible\s+assets?(?:[—-]net)?': 'Intangibles',
            r'accounts?\s+payable': 'Accounts Payable',
            r'accrued\s+(?:liabilities?|interest)': 'Accrued Interest',
            r'long[- ]term\s+debt(?!.*current)': 'Long Term Debt',
            r'deferred\s+income\s+taxes?': 'Deferred income taxes',
            r'common\s+stock': 'Common Stock',
            r'retained\s+earnings?': 'Retained Earnings',
            r'paid[- ]in\s+capital': 'Paid in Capital',
            # Additional patterns for common items
            r'margin\s+deposits?': 'Investments',  # Added
            r'derivative\s+assets?': 'Investments',  # Added
            r'derivative\s+liabilities?': 'Other',  # Added
            r'right\s+of\s+use\s+assets?': 'Net PPE',  # Added
            r'finance\s+lease\s+assets?': 'Net PPE',  # Added
            r'finance\s+lease\s+liability': 'Long Term Debt',  # Added
            r'operating\s+lease\s+liability': 'Long Term Debt',  # Added
            r'current\s+portion\s+of\s+long[- ]term\s+debt': 'Current Portion of Long Term Debt',  # Added
            r'revolving\s+lines?\s+of\s+credit': 'Short term Borrowing',  # Added
            r'long[- ]term\s+incentive': 'Other',  # Added
            r'deferred\s+compensation': 'Other',  # Added
            r'contingent\s+consideration': 'Other',  # Added
            r'subchapter\s+s\s+income\s+tax': 'Other',  # Added
            r'noncontrolling\s+interests?': 'Other',  # Added
            r'other\s+(?:current|noncurrent)\s+assets?': 'Other',  # Added
            r'other\s+(?:current|noncurrent)\s+liabilities?': 'Other',  # Added
        }
        
        for pattern, template_item in bs_rules.items():
            if re.search(pattern, desc_lower):
                return template_item, 0.9
        
        return None, 0.0

    def apply_rule_based_mapping_is(self, description: str) -> tuple[str, float]:
        """Apply rule-based mapping for income statement items"""
        import re
        desc_lower = description.lower()
        
        # Income statement rule-based mapping
        is_rules = {
            r'revenue(?:s)?': 'Revenue',
            r'cost\s+of\s+(?:goods\s+)?sales?': 'Cost of Sales',
            r'gross\s+profit': 'Gross Profit',
            r'operating\s+expenses?': 'Operating Expenses',
            r'operating\s+income': 'Operating Income',
            r'interest\s+expense': 'Interest Expense',
            r'income\s+tax(?:es)?': 'Income Tax Expense',
            r'net\s+income': 'Net Income'
        }
        
        for pattern, template_item in is_rules.items():
            if re.search(pattern, desc_lower):
                return template_item, 0.9
        
        return None, 0.0

    def apply_rule_based_mapping_cfs(self, description: str) -> tuple[str, float]:
        """Apply rule-based mapping for cash flow statement items"""
        import re
        desc_lower = description.lower()
            
        # Cash flow statement rule-based mapping - EXPANDED
        cfs_rules = {
            r'net\s+(?:income|profit|loss)': 'Net profit (loss)',
            r'depreciation': 'Depreciation',
            r'amortization': 'Depreciation',  # Map to same field
            r'deferred\s+income\s+taxes?': 'Deferred income taxes',
            r'impairment\s+and\s+other\s+losses?': 'Impairment and other losses',
            r'changes?\s+in\s+operating\s+assets?\s+and\s+liabilities?': 'Changes in operating assets and liabilities',
            r'net\s+cash\s+provided\s+by\s+\(used\s+in\)\s+operating\s+activities?': 'Net cash provided by (used in) operating activities',
            r'purchases?\s+of\s+property\s+and\s+equipment': 'Purchases of property and equipment',
            r'proceeds?\s+from\s+sale\s+of\s+assets?': 'Proceeds from sale of assets',
            r'net\s+cash\s+used\s+in\s+investing\s+activities?': 'Net cash used in investing activities',
            r'proceeds?\s+from\s+issuance': 'Proceeds from issuance',
            r'principal\s+payments?': 'Principal payments',
            r'net\s+cash\s+provided\s+by\s+\(used\s+in\)\s+financing\s+activities?': 'Net cash provided by (used in) financing activities',
            # Additional patterns for common items
            r'\(increase\)\s+decrease\s+in\s+inventories?': 'Changes in operating assets and liabilities',
            r'\(increase\)\s+decrease\s+in\s+prepaid\s+expenses?': 'Changes in operating assets and liabilities',
            r'\(increase\)\s+decrease\s+in\s+receivables?': 'Changes in operating assets and liabilities',
            r'\(increase\)\s+decrease\s+in\s+accounts?\s+payable': 'Changes in operating assets and liabilities',
            r'accrued\s+liabilities?\s+and\s+taxes?': 'Changes in operating assets and liabilities',
            r'dividends?\s+paid': 'Other Financing Activities',
            r'purchase\s+of\s+marketable\s+securities?': 'Other Investing Activities',
            r'proceeds?\s+from\s+marketable\s+securities?': 'Other Investing Activities',
            # Cash reconciliation items
            r'cash\s+and\s+cash\s+equivalents?\s+at\s+beginning\s+of\s+year': 'Starting Cash',
            r'cash\s+and\s+cash\s+equivalents?\s+at\s+end\s+of\s+year': 'Ending Cash',
            r'net\s+(?:change|increase|decrease)\s+in\s+cash': 'Net change in Cash',
            r'net\s+(?:change|increase|decrease)\s+in\s+cash\s+and\s+cash\s+equivalents?': 'Net change in Cash'
        }
        
        for pattern, template_item in cfs_rules.items():
            if re.search(pattern, desc_lower):
                return template_item, 0.9
        
        return None, 0.0

    def get_semantic_match(self, description: str, template_items: list, section_context: str) -> tuple[str, float]:
        """Get semantic match using sentence transformers"""
        if not template_items:
            return None, 0.0
        
        try:
            # Get embeddings
            desc_embedding = self.model.encode([description])
            template_embeddings = self.model.encode(template_items)
            
            # Calculate similarities
            similarities = cosine_similarity(desc_embedding, template_embeddings)[0]
            
            # Find best match
            best_idx = np.argmax(similarities)
            best_score = similarities[best_idx]
            
            if best_score > 0.3:  # Lower threshold for semantic matching
                return template_items[best_idx], best_score
            
            return None, 0.0
            
        except Exception as e:
            print(f"Error in semantic matching: {e}")
            return None, 0.0

    def hybrid_map_item_decoupled(self, description: str, template_items: List[str], 
                                 section_context: str, statement_type: str) -> Tuple[Optional[str], float, str]:
        """
        Decoupled hybrid mapping: try rule-based first regardless of section assignment,
        then LLM only if rule-based fails. This prevents LLM section misassignment from
        affecting template mapping accuracy.
        """
        # Step 1: Always try rule-based mapping first (most reliable for standard items)
        if statement_type == 'balance_sheet':
            rule_match, rule_confidence = self.apply_rule_based_mapping(description)
        elif statement_type == 'income_statement':
            rule_match, rule_confidence = self.apply_rule_based_mapping_is(description)
        elif statement_type == 'cash_flow':
            rule_match, rule_confidence = self.apply_rule_based_mapping_cfs(description)
        else:
            rule_match, rule_confidence = None, 0.0
        
        # If rule-based found a good match, use it
        if rule_match and rule_confidence >= 0.6:
            return rule_match, rule_confidence, "rule_based"
        
        # Step 2: Try semantic matching
        semantic_match, semantic_confidence = self.get_semantic_match(description, template_items, section_context)
        if semantic_match and semantic_confidence >= 0.7:
            return semantic_match, semantic_confidence, "semantic"
        
        # Step 3: Only use LLM if rule-based and semantic both failed
        if self.use_llm:
            llm_match, llm_confidence, reasoning = self.llm_mapper.map_with_llm(
                description, template_items, section_context, statement_type
            )
            if llm_match and llm_confidence >= 0.5:
                return llm_match, llm_confidence, f"llm_fallback: {reasoning}"
        
        # Step 4: If all else fails, return None (will be added to "Other")
        return None, 0.0, "no_match"

    def map_balance_sheet_decoupled(self, extracted_data: Dict, template_path: str) -> str:
        """
        Decoupled balance sheet mapping: section assignment and template mapping are separate.
        This allows correct template mapping even if section assignment is wrong.
        """
        # Load template
        workbook = openpyxl.load_workbook(template_path)
        bs_sheet = workbook["Balance Sheet"]
        
        # Get year columns
        year_cols = self.get_year_columns(bs_sheet)
        years = list(year_cols.keys())
        
        # Extract balance sheet data
        bs_data = extracted_data.get('balance_sheet', {})
        if not bs_data:
            print("No balance sheet data found")
            return template_path
        
        # Convert to list format for processing
        bs_lines = []
        for year in years:
            year_data = bs_data.get(year, [])
            for item in year_data:
                bs_lines.append({
                    'description': item.get('description', ''),
                    'idx': item.get('idx', 0),
                    'numbers': item.get('numbers', []),
                    'year': year
                })
        
        # Step 1: Assign sections (this can be wrong, but we'll handle it)
        assigned_lines = self.assign_sections_with_llm(bs_lines)
        
        # Step 2: Group by section for mapping
        section_data = {}
        for line in assigned_lines:
            section = line.get('section')
            if section:
                if section not in section_data:
                    section_data[section] = []
                section_data[section].append(line)
        
        # Step 3: Map each section using decoupled approach
        for section, items in section_data.items():
            print(f"\n[DEBUG] Mapping section '{section}' with batch approach...")
            
            # Get template items for this section
            row_map = self.get_bs_row_map(bs_sheet, *self.get_section_row_range(section))
            if not row_map:
                print(f"Warning: No row map found for section '{section}'")
                continue
                
            template_items = list(row_map.keys())
            print(f"[DEBUG] Template items for {section}: {template_items}")
            
            # Use batch mapping for efficiency
            batch_mappings = self.map_section_with_batching(items, template_items, section, 'balance_sheet')
            
            # Process batch results
            for item in items:
                desc = item['description']
                numbers = item.get('numbers', [])
                
                # Skip totals (should have been filtered, but double-check)
                if self.is_total_or_net_row(desc):
                    print(f"  [SKIP TOTAL] '{desc}' is a total/subtotal row.")
                    continue
                
                # Get mapping from batch results
                if desc in batch_mappings:
                    template_item, confidence, method = batch_mappings[desc]
                    print(f"  [MAP-BATCH] '{desc}' -> {template_item} (confidence: {confidence:.2f}, method: {method})")
                else:
                    # Fall back to individual mapping
                    template_item, confidence, method = self.hybrid_map_item_decoupled(
                        desc, template_items, section, 'balance_sheet'
                    )
                    print(f"  [MAP-INDIVIDUAL] '{desc}' -> {template_item} (confidence: {confidence:.2f}, method: {method})")
                
                if template_item:
                    # Write to template (implementation depends on your existing logic)
                    # ... (continue with existing mapping logic)
                    pass
                else:
                    print(f"  [MAP-OTHER] '{desc}' -> Other (no match found)")
                    # Add to "Other" category
                    # ... (continue with existing logic)
        
        # Save and return
        else:
            workbook.save(template_path)
            return template_path

    def get_section_row_range(self, section: str) -> Tuple[int, int]:
        """Get the row range for a given section."""
        ranges = {
            'current_assets': (7, 13),
            'noncurrent_assets': (16, 19),
            'current_liabilities': (24, 29),
            'noncurrent_liabilities': (31, 34),
            'equity': (38, 43)
        }
        return ranges.get(section, (1, 1))

    def ensure_dict_of_years_format(self, extracted_data):
        """Convert list-of-lines format to dict-of-years format if needed."""
        result = {}
        for stmt_type, lines in extracted_data.items():
            if isinstance(lines, list):
                year_dict = {}
                for line in lines:
                    desc = line['description']
                    for year, value in line['numbers'].items():
                        if year not in year_dict:
                            year_dict[year] = {}
                        year_dict[year][desc] = value
                result[stmt_type] = year_dict
            else:
                result[stmt_type] = lines
        return result

    def map_to_template(self, extracted_data: Dict, template_path: str) -> str:
        # Convert to dict-of-years format if needed
        extracted_data = self.ensure_dict_of_years_format(extracted_data)
        # Load the template workbook
        shutil.copy(template_path, "temp_template.xlsx")
        wb = load_workbook("temp_template.xlsx")

        # Get sheets for BS, IS, CFS
        bs_sheet = wb['BS']
        is_cf_sheet = wb['IS.CF']

        # Determine year columns from the template (assuming they are in B, C, D, E starting from row 6)
        year_cols = {}
        print("DEBUG: Starting year column detection...")
        # Heuristic: Find years in row 6 of the Balance Sheet
        for col_idx in range(2, 6): # Check columns B, C, D, E
            cell_val = bs_sheet.cell(row=6, column=col_idx).value
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"DEBUG: Checking column {col_letter} (row 6): {cell_val} (type: {type(cell_val)})")
            
            if isinstance(cell_val, int) and 1990 <= cell_val <= 2050:
                year_cols[str(cell_val)] = col_letter
                print(f"DEBUG: Found direct year {cell_val} in column {col_letter}")
            elif isinstance(cell_val, str) and cell_val.startswith('='):
                # Handle Excel formulas like "=B6+1" - extract the year
                print(f"DEBUG: Found formula {cell_val}, attempting to parse...")
                try:
                    # Simple parsing for formulas like "=B6+1" where B6 contains 2020
                    if '+1' in cell_val:
                        base_cell = cell_val.split('+')[0][1:]  # Extract "B6" from "=B6+1"
                        base_col = base_cell[0]  # "B"
                        base_row = int(base_cell[1:])  # 6
                        base_year = bs_sheet[f"{base_col}{base_row}"].value
                        print(f"DEBUG: Base cell {base_col}{base_row} contains: {base_year}")
                        if isinstance(base_year, int):
                            # Calculate the year for this column
                            col_offset = col_idx - openpyxl.utils.column_index_from_string(base_col)
                            calculated_year = base_year + col_offset
                            print(f"DEBUG: Calculated year: {base_year} + {col_offset} = {calculated_year}")
                            if 1990 <= calculated_year <= 2050:
                                year_cols[str(calculated_year)] = col_letter
                                print(f"DEBUG: Found calculated year {calculated_year} in column {col_letter}")
                except Exception as e:
                    print(f"DEBUG: Error parsing formula {cell_val}: {e}")
                    pass

        if not year_cols:
            print("ERROR: Could not determine year columns from template. Aborting.")
            return ""

        print(f"DEBUG: Found year columns in template: {year_cols}")

        # Map extracted years to template years if they don't match
        template_years = list(year_cols.keys())
        year_mapping = {}
        
        # Check if we need to map years
        if 'balance_sheet' in extracted_data:
            extracted_years = list(extracted_data['balance_sheet'].keys())
            print(f"[DEBUG] Extracted years: {extracted_years}")
            print(f"[DEBUG] Template years: {template_years}")
            
            if set(extracted_years) != set(template_years):
                print(f"[DEBUG] Year mismatch detected. Mapping extracted years to template years...")
                # Map extracted years to template years (first to first, second to second, etc.)
                for i, extracted_year in enumerate(extracted_years):
                    if i < len(template_years):
                        template_year = template_years[i]
                        year_mapping[extracted_year] = template_year
                        print(f"[DEBUG] Mapping {extracted_year} -> {template_year}")
                    else:
                        print(f"[DEBUG] Warning: No template year available for {extracted_year}")
            else:
                # Years match, create direct mapping
                year_mapping = {year: year for year in extracted_years}
                print(f"[DEBUG] Years match, using direct mapping")
        
        print(f"[DEBUG] Final year mapping: {year_mapping}")

        # --- Balance Sheet Mapping ---
        if 'balance_sheet' in extracted_data:
            print("\n--- Processing Balance Sheet ---")
            print(f"[DEBUG] Template year columns: {year_cols}")
            print(f"[DEBUG] Extracted balance sheet years: {list(extracted_data['balance_sheet'].keys())}")
            
            # --- Get Balance Sheet row maps ---
            # Updated row ranges to match the template screenshots
            row_maps = {
                'current_assets': self.get_bs_row_map(bs_sheet, 7, 13),           # Includes 'Other' at 12, Total at 13
                'noncurrent_assets': self.get_bs_row_map(bs_sheet, 16, 19),      # Includes 'Other' at 18, Total at 19
                'non_current_assets': self.get_bs_row_map(bs_sheet, 16, 19),     # Alternative naming
                'current_liabilities': self.get_bs_row_map(bs_sheet, 24, 29),    # Includes 'Other' at 28, Total at 29
                'noncurrent_liabilities': self.get_bs_row_map(bs_sheet, 31, 34), # Includes 'Other' at 33, Total at 34
                'equity': self.get_bs_row_map(bs_sheet, 38, 43)                  # Includes 'Other' at 42, Total at 43
            }
            # NOTE: If you update the template, update these row ranges accordingly.

            # Collect all years and columns
            years = [str(y) for y in extracted_data['balance_sheet'].keys() if str(y) in year_cols]
            if not years:
                print("Warning: No matching years found in template columns for balance sheet.")
            
            # Use year mapping instead of direct year matching
            mapped_years = []
            for extracted_year in extracted_data['balance_sheet'].keys():
                if extracted_year in year_mapping:
                    mapped_year = year_mapping[extracted_year]
                    if mapped_year in year_cols:
                        mapped_years.append(mapped_year)
                        print(f"[DEBUG] Using mapped year: {extracted_year} -> {mapped_year}")
            
            if not mapped_years:
                print("Warning: No years could be mapped to template columns for balance sheet.")
                return template_path
            
            years = mapped_years
            print(f"[DEBUG] Processing balance sheet for mapped years: {years}")
            
            # Build a dict: {desc: {year: value}}
            item_year_values = defaultdict(dict)
            for extracted_year, mapped_year in year_mapping.items():
                if mapped_year in year_cols:
                    for desc, val in extracted_data['balance_sheet'][extracted_year].items():
                        item_year_values[desc][mapped_year] = val

            # Assign sections to all items (once, using the first available year for values)
            bs_lines = [{'description': d, 'numbers': [next(iter(v.values()))]} for d, v in item_year_values.items()]
            assigned_bs_lines = self.assign_sections_with_llm(bs_lines)
            
            # Post-process section assignments to fix common misassignments
            for item in assigned_bs_lines:
                desc = item['description'].lower()
                if 'accumulated depreciation' in desc or 'less accumulated depreciation' in desc:
                    if item.get('section') != 'noncurrent_assets':
                        print(f"[SECTION FIX] Correcting 'accumulated depreciation' from '{item.get('section')}' to 'noncurrent_assets'")
                        item['section'] = 'noncurrent_assets'
                elif 'property' in desc and ('equipment' in desc or 'plant' in desc):
                    if item.get('section') != 'noncurrent_assets':
                        print(f"[SECTION FIX] Correcting 'property/equipment' from '{item.get('section')}' to 'noncurrent_assets'")
                        item['section'] = 'noncurrent_assets'
                elif 'taxes and other receivables' in desc:
                    if item.get('section') != 'current_assets':
                        print(f"[SECTION FIX] Correcting 'taxes and other receivables' from '{item.get('section')}' to 'current_assets'")
                        item['section'] = 'current_assets'
            
            self.print_section_assignments(assigned_bs_lines, ','.join(years), 'balance_sheet')

            # Track used items to prevent double counting
            self.print_section_assignments(assigned_bs_lines, ','.join(years), 'balance_sheet')

            # Track used items to prevent double counting
            used_items = set()

                # Group items by their assigned section
                section_data = defaultdict(list)
            for idx, item in enumerate(assigned_bs_lines):
                if item.get('section'):
                        section_data[item['section']].append({
                            'description': item['description'],
                        'idx': idx  # Keep index to map back to item_year_values
                        })

                # Map each section
                for section, items in section_data.items():
                # --- Print all items assigned to each section before mapping ---
                print(f"\n[DEBUG] Items assigned to section '{section}':")
                for entry in items:
                    desc = entry['description']
                    idx = entry['idx']
                    year_vals = item_year_values[desc]
                    print(f"  - {desc}: {year_vals}")

                    row_map = row_maps.get(section)
                    if not row_map:
                        print(f"Warning: No row map found for section '{section}'. Skipping.")
                        continue
                    template_items = list(row_map.keys())
                    print(f"\n[DEBUG] Mapping section '{section}'. Template items: {template_items}")

                # Track which items are mapped to which template row (for double-counting check)
                item_to_template = defaultdict(list)
                template_to_items = defaultdict(list)
                other_items = defaultdict(list)  # year -> list of (desc, value)

                # Track accumulated values for this section to prevent double counting
                section_accumulated = defaultdict(float)
                section_other_sum = defaultdict(float)  # Track by year

                # For each item, map to template and write values for all years
                    for entry in items:
                        desc = entry['description']
                    if desc in used_items:
                        print(f"  [SKIP USED] '{desc}' already mapped.")
                            continue
                        if self.is_total_or_net_row(desc):
                            print(f"  [SKIP TOTAL] '{desc}' is a total/subtotal row.")
                    continue
                    idx = entry['idx']
                    year_vals = item_year_values[desc]
                    target_item, score, method = self.hybrid_map_item_decoupled(
                        desc, template_items, section, 'balance_sheet'
                    )
                        if target_item and score >= 0.4:
                        used_items.add(desc)
                        for year, val in year_vals.items():
                            if str(year) in year_cols:
                                try:
                                    val_float = float(re.sub(r'[^\d\.-]', '', str(val)))
                                except (ValueError, TypeError):
                                    print(f"Warning: Could not convert value to float for '{desc}': {val}")
                                    continue
                                section_accumulated[f"{target_item}_{year}"] += val_float
                                item_to_template[desc].append((target_item, year, val_float))
                                template_to_items[(target_item, year)].append((desc, val_float))
                                self.log_mapping_decision(desc, year, section, target_item, val_float, method, "ACCUMULATE HYBRID")
                            else:
                                for year, val in year_vals.items():
                                    if str(year) in year_cols:
                                        try:
                                            val_float = float(re.sub(r'[^\d\.-]', '', str(val)))
                                        except (ValueError, TypeError):
                                            continue
                                        section_other_sum[year] += val_float
                                        other_items[year].append((desc, val_float))
                                        self.log_mapping_decision(desc, year, section, 'Other', val_float, method, "ACCUMULATE OTHER")
                        used_items.add(desc)

                # --- Print mapping summary for this section ---
                print(f"\n[DEBUG] Mapping summary for section '{section}':")
                for (target_item, year), items_list in template_to_items.items():
                    total = sum(val for _, val in items_list)
                    print(f"  {target_item} [{year}]: {total} (from: {[desc for desc, _ in items_list]})")
                for year, items_list in other_items.items():
                    total = sum(val for _, val in items_list)
                    print(f"  Other [{year}]: {total} (from: {[desc for desc, _ in items_list]})")

                # --- Double-counting check ---
                # Only warn if an item is mapped to more than one template row for the SAME year
                for desc, mappings in item_to_template.items():
                    year_targets = [(target, year) for (target, year, _) in mappings]
                    # Count how many unique (target, year) pairs there are
                    year_target_counts = {}
                    for target, year in year_targets:
                        year_target_counts.setdefault(year, set()).add(target)
                    for year, targets in year_target_counts.items():
                        if len(targets) > 1:
                            print(f"[WARNING] Double-counted item: '{desc}' mapped to multiple template rows {targets} for year {year}")

                    # Write accumulated values to template
                for key, total_val in section_accumulated.items():
                    if total_val != 0:
                        target_item, year = key.rsplit('_', 1)
                        if target_item in row_map and str(year) in year_cols:
                            col = year_cols[str(year)]
                            row_idx = row_map[target_item]
                            bs_sheet[f"{col}{row_idx}"] = total_val
                            self.log_mapping_decision(desc, year, section, target_item, total_val, method, "WRITE")

                # Write 'Other' sums for this section
                # Ensure 'Other' is included in the equity row map as well
                if 'Other' in row_map:
                    for year, other_sum in section_other_sum.items():
                        if other_sum != 0 and str(year) in year_cols:
                            col = year_cols[str(year)]
                        row_idx = row_map['Other']
                        existing_val = bs_sheet[f"{col}{row_idx}"].value or 0
                        if isinstance(existing_val, str): existing_val = 0
                        bs_sheet[f"{col}{row_idx}"] = existing_val + other_sum
                            self.log_mapping_decision(desc, year, section, 'Other', other_sum, method, "OTHER SUM")

                # --- After writing, print sum check for each section and year ---
                for year in years:
                    mapped_sum = 0
                    for t_item in template_items:
                        if t_item in row_map and str(year) in year_cols:
                            col = year_cols[str(year)]
                            row_idx = row_map[t_item]
                            val = bs_sheet[f"{col}{row_idx}"].value
                            try:
                                mapped_sum += float(val) if val not in [None, ''] else 0
                            except (ValueError, TypeError):
                                continue
                    # Find template total row (e.g., 'Total Current Assets')
                    total_row = None
                    for t_item in template_items:
                        if t_item.lower().startswith('total'):
                            total_row = row_map[t_item]
                            break
                    template_total = None
                    if total_row and str(year) in year_cols:
                        col = year_cols[str(year)]
                        template_total = bs_sheet[f"{col}{total_row}"].value
                    # Only check if template_total is a number (not a formula)
                    if template_total is not None and isinstance(template_total, str) and template_total.strip().startswith('='):
                        self.extraction_logger.info(f"[CHECK] Section '{section}' [{year}]: template total is a formula ('{template_total}'), skipping numeric check.")
                        print(f"[CHECK] Section '{section}' [{year}]: template total is a formula ('{template_total}'), skipping numeric check.")
                        ok = None
                    else:
                        try:
                            ok = (template_total is not None and abs(mapped_sum - float(template_total)) < 1e-2)
                        except (ValueError, TypeError):
                            ok = None
                    self.log_section_sum_check(section, year, mapped_sum, template_total, ok)

            # --- For equity section, print all items considered and mapping reasons ---
            print(f"\n[DEBUG] Equity section analysis:")
            for entry in section_data.get('equity', []):
                desc = entry['description']
                idx = entry['idx']
                year_vals = item_year_values[desc]
                print(f"  - {desc}: {year_vals}")

        # --- Process Income Statement ---
        if 'income_statement' in extracted_data:
            print("\n--- Processing Income Statement ---")
            print(f"[DEBUG] Template year columns: {year_cols}")
            print(f"[DEBUG] Extracted income statement years: {list(extracted_data['income_statement'].keys())}")
            is_data = extracted_data['income_statement']
            print(f"[DEBUG] IS data structure: {type(is_data)}")
            print(f"[DEBUG] IS data keys: {list(is_data.keys()) if isinstance(is_data, dict) else 'Not a dict'}")
            
            # Get IS row maps
            is_row_maps = {
                'revenue': self.get_is_row_map(is_cf_sheet, 'revenue'),
                'operating_expenses': self.get_is_row_map(is_cf_sheet, 'operating_expenses'),
                'operating_income': self.get_is_row_map(is_cf_sheet, 'operating_income'),
                'other_income_expense': self.get_is_row_map(is_cf_sheet, 'other_income_expense'),
                'tax_net_income': self.get_is_row_map(is_cf_sheet, 'tax_net_income')
            }
            
            # Debug: Print available template items
            print(f"[DEBUG] Available IS template items:")
            for section, row_map in is_row_maps.items():
                print(f"  {section}: {list(row_map.keys()) if row_map else 'No items'}")
            
            # Process each year
            for extracted_year in is_data.keys():
                if extracted_year not in year_mapping:
                    print(f"[DEBUG] Skipping IS year {extracted_year} - not in year mapping")
                    continue
                
                mapped_year = year_mapping[extracted_year]
                if mapped_year not in year_cols:
                    print(f"[DEBUG] Skipping IS year {extracted_year} -> {mapped_year} - not in template")
                    continue
                        
                col = year_cols[mapped_year]
                year_data = is_data[extracted_year]
                print(f"[DEBUG] IS year {extracted_year} -> {mapped_year} data: {type(year_data)}, length: {len(year_data) if isinstance(year_data, dict) else 'N/A'}")
                
                if not isinstance(year_data, dict):
                    print(f"[WARN] IS year {extracted_year} data is not a dict, skipping")
                    continue
                        
                print(f"\n[DEBUG] Processing IS for year {extracted_year} -> {mapped_year}:")
                
                # Group items by section for batch processing
                section_items = defaultdict(list)
                for desc, val in year_data.items():
                    # Skip totals
                    if self.is_total_or_net_row(desc):
                        print(f"  [SKIP TOTAL] '{desc}' is a total/subtotal row.")
                        continue
                    
                    # Handle different value types
                    if isinstance(val, list):
                        numbers = val
                    elif isinstance(val, (int, float, str)):
                        numbers = [val]
                    else:
                        print(f"[WARN] Unknown IS value type: {type(val)}, skipping")
                        continue
                    
                    if not numbers:
                        continue
                    
                    try:
                        val_float = float(re.sub(r'[^\d\.-]', '', str(numbers[0])))
                        except (ValueError, TypeError):
                        print(f"[WARN] Could not convert IS value to float: {numbers[0]}")
                            continue
                        
                    # Assign section using rule-based approach for grouping
                    assigned_section = None
                    desc_lower = desc.lower()
                    
                    # Simple section assignment for grouping
                    if any(keyword in desc_lower for keyword in ['revenue', 'sales', 'income']):
                        assigned_section = 'revenue'
                    elif any(keyword in desc_lower for keyword in ['cost', 'expense', 'depreciation', 'amortization']):
                        assigned_section = 'operating_expenses'
                    elif any(keyword in desc_lower for keyword in ['interest']):
                        assigned_section = 'other_income_expense'
                    elif any(keyword in desc_lower for keyword in ['tax']):
                        assigned_section = 'tax_net_income'
                    else:
                        assigned_section = 'operating_expenses'  # Default
                    
                    section_items[assigned_section].append({
                        'description': desc,
                        'value': val_float
                    })
                
                # Process each section with batch mapping
                for section, items in section_items.items():
                    if not items:
                            continue
                        
                    row_map = is_row_maps.get(section)
                    if not row_map:
                        print(f"[WARN] No row map found for IS section '{section}'")
                        continue
                    
                    template_items = list(row_map.keys())
                    print(f"\n[DEBUG] Processing IS section '{section}' with batch mapping ({len(items)} items)")
                    
                    # Use batch mapping
                    batch_mappings = self.map_section_with_batching(items, template_items, section, 'income_statement')
                    
                    # Process batch results
                    for item in items:
                        desc = item['description']
                        val_float = item['value']
                        
                        if desc in batch_mappings:
                            template_item, confidence, method = batch_mappings[desc]
                            print(f"  [MAP-BATCH-IS] '{desc}' -> {section}::{template_item} ({val_float}) [method: {method}]")
                        else:
                            # Fall back to individual mapping
                            template_item, confidence, method = self.hybrid_map_item_decoupled(
                                desc, template_items, section, 'income_statement'
                            )
                            print(f"  [MAP-INDIVIDUAL-IS] '{desc}' -> {section}::{template_item} ({val_float}) [method: {method}]")
                        
                        if template_item and confidence >= 0.4:
                            if template_item in row_map:
                                row_idx = row_map[template_item]
                                is_cf_sheet[f"{col}{row_idx}"] = val_float
                            else:
                                print(f"  [WARN] Template item '{template_item}' not found in row map. Available items: {list(row_map.keys())}")
                                # Try to find a similar item
                                for available_item in row_map.keys():
                                    if self.get_similarity(template_item.lower(), available_item.lower()) > 0.7:
                                        row_idx = row_map[available_item]
                                        is_cf_sheet[f"{col}{row_idx}"] = val_float
                                        print(f"  [MAP-IS-SIMILAR] '{desc}' -> {section}::{available_item} ({val_float}) [method: {method}]")
                                        break
                        else:
                            # Add to "Other" for this section
                            other_row = self.get_other_category_for_is_section(section)
                            if other_row and other_row in row_map:
                                row_idx = row_map[other_row]
                        existing_val = is_cf_sheet[f"{col}{row_idx}"].value or 0
                        if isinstance(existing_val, str): existing_val = 0
                                is_cf_sheet[f"{col}{row_idx}"] = existing_val + val_float
                                print(f"  [MAP-OTHER-IS] '{desc}' -> {section}::Other ({val_float})")

        # --- Process Cash Flow Statement ---
        if 'cash_flow' in extracted_data:
            print("\n--- Processing Cash Flow Statement ---")
            print(f"[DEBUG] Template year columns: {year_cols}")
            print(f"[DEBUG] Extracted cash flow years: {list(extracted_data['cash_flow'].keys())}")
            cf_data = extracted_data['cash_flow']
            print(f"[DEBUG] CFS data structure: {type(cf_data)}")
            print(f"[DEBUG] CFS data keys: {list(cf_data.keys()) if isinstance(cf_data, dict) else 'Not a dict'}")
            
            # Get CFS row maps
            cf_row_maps = {
                'operating_activities': self.get_cfs_row_map(['Net profit (loss)', 'Depreciation', 'Changes in operating assets and liabilities', 'Other Operating Activities'], 'operating_activities'),
                'investing_activities': self.get_cfs_row_map(['Purchases of property and equipment', 'Proceeds from sale of assets', 'Other Investing Activities'], 'investing_activities'),
                'financing_activities': self.get_cfs_row_map(['Proceeds from issuance', 'Principal payments', 'Other Financing Activities'], 'financing_activities'),
                'cash_reconciliation': self.get_cfs_row_map(['Net change in Cash', 'Starting Cash', 'Ending Cash'], 'cash_reconciliation')
            }
            
            # Process each year
            for extracted_year in cf_data.keys():
                if extracted_year not in year_mapping:
                    print(f"[DEBUG] Skipping CFS year {extracted_year} - not in year mapping")
                    continue
                
                mapped_year = year_mapping[extracted_year]
                if mapped_year not in year_cols:
                    print(f"[DEBUG] Skipping CFS year {extracted_year} -> {mapped_year} - not in template")
                    continue
                    
                col = year_cols[mapped_year]
                year_data = cf_data[extracted_year]
                print(f"[DEBUG] CFS year {extracted_year} -> {mapped_year} data: {type(year_data)}, length: {len(year_data) if isinstance(year_data, dict) else 'N/A'}")
                
                if not isinstance(year_data, dict):
                    print(f"[WARN] CFS year {extracted_year} data is not a dict, skipping")
                    continue
                
                print(f"\n[DEBUG] Processing CFS for year {extracted_year} -> {mapped_year}:")
                
                # Group items by section for batch processing
                section_items = defaultdict(list)
                for desc, val in year_data.items():
                    # Skip totals
                    if self.is_total_or_net_row(desc):
                        print(f"  [SKIP TOTAL] '{desc}' is a total/subtotal row.")
                        continue
                    
                    # Handle different value types
                    if isinstance(val, list):
                        numbers = val
                    elif isinstance(val, (int, float, str)):
                        numbers = [val]
                    else:
                        print(f"[WARN] Unknown CFS value type: {type(val)}, skipping")
                        continue
                    
                    if not numbers:
                        continue
                    
                    try:
                        val_float = float(re.sub(r'[^\d\.-]', '', str(numbers[0])))
                    except (ValueError, TypeError):
                        print(f"[WARN] Could not convert CFS value to float: {numbers[0]}")
                            continue
                        
                    # Assign section using rule-based approach for grouping
                    assigned_section = None
                    desc_lower = desc.lower()
                    
                    # Simple section assignment for grouping
                    if any(keyword in desc_lower for keyword in ['operating', 'net income', 'depreciation', 'amortization']):
                        assigned_section = 'operating_activities'
                    elif any(keyword in desc_lower for keyword in ['investing', 'purchase', 'proceeds', 'acquisition']):
                        assigned_section = 'investing_activities'
                    elif any(keyword in desc_lower for keyword in ['financing', 'debt', 'stock', 'dividend']):
                        assigned_section = 'financing_activities'
                    elif any(keyword in desc_lower for keyword in ['cash', 'beginning', 'ending', 'change']):
                        assigned_section = 'cash_reconciliation'
                            else:
                        assigned_section = 'operating_activities'  # Default
                    
                    section_items[assigned_section].append({
                        'description': desc,
                        'value': val_float
                    })
                
                # Process each section with batch mapping
                for section, items in section_items.items():
                    if not items:
                        continue
                    
                    row_map = cf_row_maps.get(section)
                    if not row_map:
                        print(f"[WARN] No row map found for CFS section '{section}'")
                        continue
                    
                    template_items = list(row_map.keys())
                    print(f"\n[DEBUG] Processing CFS section '{section}' with batch mapping ({len(items)} items)")
                    
                    # Use batch mapping
                    batch_mappings = self.map_section_with_batching(items, template_items, section, 'cash_flow')
                    
                    # Process batch results
                    for item in items:
                        desc = item['description']
                        val_float = item['value']
                        
                        if desc in batch_mappings:
                            template_item, confidence, method = batch_mappings[desc]
                            print(f"  [MAP-BATCH-CFS] '{desc}' -> {section}::{template_item} ({val_float}) [method: {method}]")
                        else:
                            # Fall back to individual mapping
                            template_item, confidence, method = self.hybrid_map_item_decoupled(
                                desc, template_items, section, 'cash_flow'
                            )
                            print(f"  [MAP-INDIVIDUAL-CFS] '{desc}' -> {section}::{template_item} ({val_float}) [method: {method}]")
                        
                        if template_item and confidence >= 0.4:
                            if template_item in row_map:
                                row_idx = row_map[template_item]
                                is_cf_sheet[f"{col}{row_idx}"] = val_float
                            else:
                                print(f"  [WARN] Template item '{template_item}' not found in row map. Available items: {list(row_map.keys())}")
                                # Try to find a similar item
                                for available_item in row_map.keys():
                                    if self.get_similarity(template_item.lower(), available_item.lower()) > 0.7:
                                        row_idx = row_map[available_item]
                                        is_cf_sheet[f"{col}{row_idx}"] = val_float
                                        print(f"  [MAP-CFS-SIMILAR] '{desc}' -> {section}::{available_item} ({val_float}) [method: {method}]")
                                        break
                        else:
                            # Add to "Other" for this section
                            other_row = self.get_other_category_for_cfs_section(section)
                            if other_row and other_row in row_map:
                                row_idx = row_map[other_row]
                        existing_val = is_cf_sheet[f"{col}{row_idx}"].value or 0
                        if isinstance(existing_val, str): existing_val = 0
                                is_cf_sheet[f"{col}{row_idx}"] = existing_val + val_float
                                print(f"  [MAP-OTHER-CFS] '{desc}' -> {section}::Other ({val_float})")

        # Save the populated template
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        current_dir = Path(__file__).resolve().parent
        project_root = current_dir.parent.parent
        output_dir = project_root / "output_excel"
        output_dir.mkdir(exist_ok=True)
        
        populated_template_path = output_dir / f"populated_template_{timestamp}.xlsx"
        wb.save(populated_template_path)
        
        # Close the workbook before trying to delete the temp file
        wb.close()
        
        # Try to delete the temporary file with error handling
        try:
        Path("temp_template.xlsx").unlink()
        except (PermissionError, FileNotFoundError) as e:
            print(f"Warning: Could not delete temporary file temp_template.xlsx: {e}")
            # The file will be cleaned up later or can be deleted manually
        
        print(f"\nTemplate populated and saved to: {populated_template_path}")
        return str(populated_template_path)

    def get_bs_row_map(self, sheet, start_row, end_row):
        """Helper to get row map for a BS section."""
        row_map = {}
        for row in range(start_row, end_row + 1):
            cell_val = sheet.cell(row=row, column=1).value
            if cell_val:
                row_map[cell_val] = row
        return row_map
    
    def log_mapping_decision(self, desc, year, section, target_item, value, method, action, double_counted=False):
        msg = f"[MAP-{action}] '{desc}' [{year}] -> {section}::{target_item} ({value}) [method: {method}]"
        if double_counted:
            msg += " [DOUBLE COUNTED WARNING]"
        print(msg)
        self.extraction_logger.info(msg)

    def log_skip_decision(self, desc, year, section, reason):
        msg = f"[SKIP] '{desc}' [{year}] in {section}: {reason}"
        print(msg)
        self.extraction_logger.info(msg)

    def log_section_sum_check(self, section, year, mapped_sum, template_total, ok):
        msg = f"[CHECK] Section '{section}' [{year}]: mapped sum = {mapped_sum}, template total = {template_total} -> {'OK' if ok else 'MISMATCH'}"
        print(msg)
        self.extraction_logger.info(msg)

    def batch_map_items_with_llm(self, items: List[dict], template_items: List[str], 
                                section_context: str, statement_type: str) -> Dict[str, Tuple[str, float, str]]:
        """
        Map multiple items to template using LLM in a single batch call.
        Returns dict mapping item description to (template_item, confidence, method).
        """
        if not self.use_llm:
            return {}
        
        # Create batch prompt
        items_formatted = "\n".join([f'- "{item["description"]}"' for item in items])
        
        prompt = f"""You are a financial statement mapping expert. Your task is to map financial line items to the most appropriate template items.

STATEMENT TYPE: {statement_type.upper()}
SECTION: {section_context.upper()}

FINANCIAL LINE ITEMS TO MAP:
{items_formatted}

AVAILABLE TEMPLATE ITEMS:
{chr(10).join(f"- {item}" for item in template_items)}

INSTRUCTIONS:
1. Analyze each financial line item and find the best match from the template items
2. Consider synonyms, abbreviations, and common variations
3. Pay attention to the section context
4. If no good match exists, return "Other"
5. Provide a confidence score from 0.0 to 1.0

IMPORTANT: Your response must be a JSON object where keys are the line item descriptions and values are [template_item_name, confidence_score, reasoning].

Example Response Format:
{{
  "Cash and cash equivalents": ["Cash and equivalents", 1.0, "Direct match"],
  "Trade receivables": ["Accounts Receivable", 0.9, "Synonym match"],
  "Unknown item": ["Other", 0.2, "No good match"]
}}

Your response:"""
        
        try:
            response = requests.post(
                f"{self.llm_mapper.ollama_url}/api/generate",
                json={
                    "model": self.llm_mapper.model_name,
                    "prompt": prompt,
                    "stream": False,
                    "options": {
                        "temperature": 0.1,
                        "top_p": 0.9,
                        "max_tokens": 1024
                    }
                },
                timeout=30
            )
            
            if response.status_code == 200:
                result = response.json()
                response_text = result.get("response", "").strip()
                
                # Parse JSON response
                try:
                    import json
                    import re
                    
                    # Find JSON object in response
                    json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                    if json_match:
                        json_str = json_match.group(0)
                        mappings = json.loads(json_str)
                        
                        results = {}
                        for desc, mapping in mappings.items():
                            if isinstance(mapping, list) and len(mapping) >= 2:
                                template_item = str(mapping[0]).strip()
                                confidence = float(mapping[1])
                                reasoning = str(mapping[2]) if len(mapping) > 2 else ""
                                results[desc] = (template_item, confidence, f"llm_batch: {reasoning}")
                        
                        return results
                except Exception as e:
                    print(f"[WARN] Failed to parse LLM batch response: {e}")
                    return {}
            
        except Exception as e:
            print(f"[WARN] LLM batch call failed: {e}")
            return {}
        
        return {}

    def map_section_with_batching(self, section_data: List[dict], template_items: List[str], 
                                 section_context: str, statement_type: str) -> Dict[str, Tuple[str, float, str]]:
        """
        Map a section of items using batched LLM calls for efficiency.
        """
        if not section_data:
            return {}
        
        # Use batch mapping if LLM is available
        if self.use_llm:
            batch_results = self.batch_map_items_with_llm(section_data, template_items, section_context, statement_type)
            if batch_results:
                return batch_results
        
        # Fall back to individual mapping
        results = {}
        for item in section_data:
            desc = item['description']
            template_item, confidence, method = self.hybrid_map_item_decoupled(
                desc, template_items, section_context, statement_type
            )
            results[desc] = (template_item, confidence, method)
        
        return results

def main():
    # Get project root directory
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent.parent
    
    # Get paths
    template_path = project_root / "templates" / "financial_template.xlsx"
    if not template_path.exists():
        print(f"Template not found at {template_path}")
        return
        
    # Get most recent extracted Excel file or use command-line argument
    output_dir = project_root / "output_excel"
    if not output_dir.exists():
        print("No output directory found")
        return
        
    if len(sys.argv) > 1:
        latest_file = Path(sys.argv[1])
        if not latest_file.exists():
            print(f"Specified file does not exist: {latest_file}")
            return
    else:
    excel_files = [f for f in output_dir.glob("*.xlsx") if not f.name.startswith('~$')]
    if not excel_files:
        print("No valid (non-temporary) Excel files found in output directory")
        return
    # Sort by creation time and get most recent
    latest_file = max(excel_files, key=lambda x: x.stat().st_ctime)
    print(f"\nProcessing {latest_file}")
    
    # Read extracted data
    extracted_data = {}
    for sheet_name, df in pd.read_excel(latest_file, sheet_name=None).items():
        statement_type = sheet_name.lower().replace(' ', '_')
        extracted_data[statement_type] = {}
        
        # Find all year columns (exclude 'Description')
        year_cols = [col for col in df.columns if col != 'Description']
        print(f"[DEBUG] Found year columns for {statement_type}: {year_cols}")
        
        for year in year_cols:
            extracted_data[statement_type][str(year)] = {}
        for _, row in df.iterrows():
            desc = row['Description']
                if pd.notna(desc) and pd.notna(row.get(year)):
                    extracted_data[statement_type][str(year)][desc] = row[year]
        
        print(f"[DEBUG] Loaded {statement_type}: {len(extracted_data[statement_type])} years")
        for year, items in extracted_data[statement_type].items():
            print(f"[DEBUG]   {year}: {len(items)} items")
    
    # Map to template
    matcher = TemplateMatcher()
    output_path = matcher.map_to_template(extracted_data, str(template_path))
    print(f"\nTemplate populated and saved to: {output_path}")

if __name__ == "__main__":
    main()

