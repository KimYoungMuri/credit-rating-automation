import requests
import json
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import re
from typing import Dict, List, Tuple, Optional
from collections import defaultdict
import shutil

class SingleStatementMapper:
    """
    Mapper that processes one financial statement at a time with a single comprehensive LLM call.
    Takes entire extracted table and outputs complete mapped template.
    """
    
    def __init__(self, ollama_url: str = "http://localhost:11434"):
        self.ollama_url = ollama_url
        self.model_name = "mistral:latest"
        
    def check_ollama_available(self) -> bool:
        """Check if Ollama is running and available"""
        try:
            response = requests.get(f"{self.ollama_url}/api/tags", timeout=5)
            return response.status_code == 200
        except:
            return False
    
    def format_extracted_data(self, extracted_data: Dict) -> str:
        """Format extracted data into a clean, readable format for the prompt"""
        formatted_lines = []
        
        # Limit to first 20 items to keep prompt manageable
        item_count = 0
        max_items = 20
        
        for year, year_data in extracted_data.items():
            for desc, val in year_data.items():
                if item_count >= max_items:
                    break
                    
                # Skip totals
                if self.is_total_row(desc):
                    continue
                
                # Handle different value types
                if isinstance(val, list):
                    numbers = val
                elif isinstance(val, (int, float, str)):
                    numbers = [val]
                else:
                    continue
                
                if not numbers:
                    continue
                
                try:
                    val_float = float(re.sub(r'[^\d\.-]', '', str(numbers[0])))
                    # Format with shorter description if needed
                    short_desc = desc[:50] + "..." if len(desc) > 50 else desc
                    formatted_lines.append(f"- {short_desc} ({year}): {val_float:,.0f}")
                    item_count += 1
                except (ValueError, TypeError):
                    continue
            
            if item_count >= max_items:
                break
        
        if item_count >= max_items:
            formatted_lines.append(f"... and {sum(len(year_data) for year_data in extracted_data.values()) - max_items} more items")
        
        return "\n".join(formatted_lines)
    
    def get_template_structure(self, statement_type: str) -> Dict:
        """Get the complete template structure for a given statement type"""
        if statement_type == 'balance_sheet':
            return {
                "Current Assets": [
                    "Cash and equivalents",
                    "Accounts Receivable", 
                    "Prepaid Expenses",
                    "Inventory",
                    "Investments",
                    "Other"
                ],
                "Non-Current Assets": [
                    "Net PPE",
                    "Goodwill",
                    "Intangibles",
                    "Other"
                ],
                "Current Liabilities": [
                    "Accounts Payable",
                    "Accrued Interest",
                    "Short term Borrowing",
                    "Current Portion of Long Term Debt",
                    "Other"
                ],
                "Non-Current Liabilities": [
                    "Long Term Debt",
                    "Deferred income taxes",
                    "Other"
                ],
                "Equity": [
                    "Common Stock",
                    "Retained Earnings",
                    "Paid in Capital",
                    "Other"
                ]
            }
        elif statement_type == 'income_statement':
            return {
                "Revenue": ["Revenue"],
                "Operating Expenses": [
                    "Operating Expenses",
                    "Depreciation (-)",
                    "Amortization (-)",
                    "Assets gain(loss) impairments"
                ],
                "Other Income/Expense": [
                    "Interest Expense (-)",
                    "Interest Income (+)",
                    "Other income(expenses)"
                ],
                "Tax and Net Income": [
                    "Income Before Taxes",
                    "Tax expense",
                    "Net Income"
                ]
            }
        elif statement_type == 'cash_flow':
            return {
                "Operating Activities": [
                    "Net Income",
                    "Changes in noncash items",
                    "Changes in Assets and Liabilities",
                    "Net Cash from(used) Operating Activities"
                ],
                "Investing Activities": [
                    "CapEx",
                    "Proceeds from asset sales",
                    "Net cash from(used) for investing"
                ],
                "Financing Activities": [
                    "Issuance of Debt (long+short term)",
                    "Retirement of Debt (long+short term)",
                    "Issuance of Stock",
                    "Dividends Paid",
                    "Net cash from(used) for financing"
                ],
                "Cash Reconciliation": [
                    "Net change in Cash",
                    "Starting Cash",
                    "Ending Cash"
                ]
            }
        else:
            return {}
    
    def format_template_structure(self, template_structure: Dict) -> str:
        """Format template structure for the prompt"""
        formatted = []
        for section, items in template_structure.items():
            formatted.append(f"Section: {section}")
            for item in items:
                formatted.append(f"  - {item}")
            formatted.append("")
        return "\n".join(formatted)
    
    def create_comprehensive_prompt(self, extracted_data: Dict, statement_type: str) -> str:
        """
        Create a comprehensive prompt for mapping a financial statement to the template.
        """
        years = list(extracted_data.keys())
        years_str = ', '.join(years)
        extracted_formatted = "\n".join([
            f"- {desc} ({year}): {val}" for year, items in extracted_data.items() for desc, val in items.items()
        ])
        template_structure = self.get_template_structure(statement_type)
        template_formatted = "\n".join([
            f"Section: {section}\n  " + "\n  ".join(items) for section, items in template_structure.items()
        ])

        # Improved instructions for balance sheet
        if statement_type == 'balance_sheet':
            instructions = """
MAPPING RULES & INSTRUCTIONS:
- Use ONLY the following sections and line items (do not invent new rows):
  * Current Assets: Cash and equivalents, Accounts Receivable, Prepaid Expenses, Inventory, Investments, Other
  * Non-Current Assets: Net PPE, Goodwill, Intangibles, Other
  * Current Liabilities: Accounts Payable, Accrued Interest, Short term Borrowing, Current Portion of Long Term Debt, Other
  * Non-Current Liabilities: Long Term Debt, Deferred income taxes, Other
  * Equity: Common Stock, Retained Earnings, Paid in Capital, Other
- Map each extracted item to the most appropriate template row, using synonyms, abbreviations, and common accounting logic.
- If an item does not fit any row, assign it to 'Other' in the appropriate section.
- Do NOT map or generate values for any total/subtotal rows (e.g., 'Total Current Assets', 'Total Liabilities'). These are calculated by formulas.
- For items that could fit multiple rows, pick the most specific match.
- Group related items appropriately (e.g., multiple cash accounts → "Cash and equivalents").
- Map each value to the correct year (see year mapping if provided).

IMPORTANT:
- Output must be STRICT JSON. Do NOT include any comments, explanations, or calculations in the output.
- All values must be numbers, not expressions (e.g., use 33470, not 33275.0 + 195).
- Do NOT use // comments or any non-JSON syntax. Only output the JSON object as specified below.
"""
        elif statement_type == 'income_statement':
            instructions = """
MAPPING RULES:
- Revenue/Sales → "Revenue"
- Cost of goods → "Operating Expenses"
- Depreciation → "Depreciation (-)"
- Interest → "Interest Expense (-)"
- Tax → "Tax expense"
- Net income → "Net Income"
- Unmatched → "Other" in appropriate section
"""
        elif statement_type == 'cash_flow':
            instructions = """
MAPPING RULES:
- Net income → "Net Income"
- Depreciation → "Changes in noncash items"
- CapEx → "CapEx"
- Debt activities → "Issuance of Debt" or "Retirement of Debt"
- Dividends → "Dividends Paid"
- Cash change → "Net change in Cash"
- Unmatched → "Other" in appropriate section
"""
        else:
            instructions = ""

        prompt = f"""You are an expert financial analyst. Your task is to map extracted {statement_type.replace('_', ' ')} line items to a standardized template.\n\nEXTRACTED DATA (showing key items):\n{extracted_formatted}\n\nTEMPLATE SECTIONS AND ROWS:\n{template_formatted}\n\nINSTRUCTIONS:\n1. Map each item to the most appropriate template row\n2. Sum multiple items that map to the same row\n3. Use 'Other' for unmatched items\n4. Include all years ({years_str})\n\n{instructions}\nOUTPUT: JSON only with this structure:\n{{\n  \"mappings\": {{\n    \"Section Name\": {{\n      \"Template Row\": {{\n        \"{years[0]}\": value,\n        \"{years[1] if len(years) > 1 else years[0]}\": value\n      }}\n    }}\n  }},\n  \"unmapped_items\": [\n    {{\"description\": \"item\", \"year\": \"2022\", \"value\": 1000, \"reason\": \"no match\"}}\n  ]\n}}\n\nYour response:" """

        return prompt
    
    def is_total_row(self, description: str) -> bool:
        desc_lower = description.lower().strip()
        total_patterns = [
            r'^total\s',
            r'\stotal$',
            r'^sum\s',
            r'\ssum$',
            r'^subtotal',
            r'\ssubtotal$',
            r'^net\s(?!income|loss)',
            r'\snet$',
            r'^aggregate',
            r'^grand\s+total',
            r'^overall'
        ]
        return any(re.search(pattern, desc_lower) for pattern in total_patterns)
    
    def call_ollama_comprehensive(self, prompt: str, timeout: int = 180) -> Optional[Dict]:
        """Call Ollama with comprehensive prompt and parse JSON response"""
        try:
            print(f"[DEBUG] Sending comprehensive prompt to Ollama (length: {len(prompt)} chars)")
            print(f"[DEBUG] Timeout set to {timeout} seconds")
            
            response = requests.post(
                f"{self.ollama_url}/api/generate",
                json={
                    "model": self.model_name,
                    "prompt": prompt,
                    "stream": False,
                    "options": {
                        "temperature": 0.1,
                        "top_p": 0.9,
                        "max_tokens": 2048  # Reduced from 4096
                    }
                },
                timeout=timeout
            )
            
            if response.status_code != 200:
                print(f"[ERROR] Ollama API error: {response.status_code}")
                return None
            
            result = response.json()
            response_text = result.get("response", "").strip()
            
            print(f"[DEBUG] Received response (length: {len(response_text)} chars)")
            
            # Parse JSON response
            try:
                # Find JSON object in response
                json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                if json_match:
                    json_str = json_match.group(0)
                    parsed = json.loads(json_str)
                    print(f"[DEBUG] Successfully parsed comprehensive response")
                    return parsed
                else:
                    print(f"[ERROR] No JSON found in response")
                    print(f"[DEBUG] Response text: {response_text[:200]}...")
                    return None
                    
            except json.JSONDecodeError as e:
                print(f"[ERROR] Failed to parse JSON response: {e}")
                print(f"[DEBUG] Response text: {response_text[:200]}...")
                return None
                
        except requests.exceptions.Timeout:
            print(f"[ERROR] Ollama request timed out after {timeout} seconds")
            print(f"[SUGGESTION] Try reducing the number of items in the prompt")
            return None
        except requests.exceptions.ConnectionError:
            print(f"[ERROR] Connection error - check if Ollama is running")
            return None
        except Exception as e:
            print(f"[ERROR] Comprehensive Ollama call failed: {e}")
            return None
    
    def map_statement_comprehensive(self, extracted_data: Dict, statement_type: str) -> Tuple[Dict, List]:
        """Map an entire statement comprehensively using a single LLM call"""
        if not self.check_ollama_available():
            print(f"[ERROR] Ollama not available for comprehensive mapping of {statement_type}")
            return {}, []
        
        print(f"\n[INFO] Processing {statement_type} with comprehensive mapping...")
        print(f"[INFO] Years: {list(extracted_data.keys())}")
        print(f"[INFO] Total items: {sum(len(year_data) for year_data in extracted_data.values())}")
        
        # Create comprehensive prompt
        prompt = self.create_comprehensive_prompt(extracted_data, statement_type)
        
        # Call Ollama
        result = self.call_ollama_comprehensive(prompt)
        
        if result:
            mappings = result.get("mappings", {})
            unmapped_items = result.get("unmapped_items", [])
            
            print(f"[SUCCESS] Comprehensive mapping successful for {statement_type}:")
            print(f"[INFO]   Mapped sections: {len(mappings)}")
            print(f"[INFO]   Unmapped items: {len(unmapped_items)}")
            
            # Print summary of mappings
            for section, section_mappings in mappings.items():
                print(f"[INFO]   {section}: {len(section_mappings)} rows mapped")
            
            return mappings, unmapped_items
        else:
            print(f"[ERROR] Comprehensive mapping failed for {statement_type}")
            return {}, []
    
    def apply_mappings_to_excel(self, mappings: Dict, template_path: str, statement_type: str, 
                               year_mapping: Dict = None) -> str:
        """Apply comprehensive mappings to the Excel template"""
        
        # Load template
        wb = openpyxl.load_workbook(template_path)
        
        if statement_type == 'balance_sheet':
            sheet = wb['BS']
            # Define row mappings for balance sheet - UPDATED to match actual template
            row_mappings = {
                "Current Assets": {
                    "Cash and equivalents": 7,
                    "Accounts Receivable": 8,
                    "Prepaid Expenses": 9,
                    "Inventory": 10,
                    "Investments": 11,
                    "Other": 12
                },
                "Non-Current Assets": {
                    "Net PPE": 16,
                    "Goodwill": 17,
                    "Intangibles": 18,
                    "Other": 19
                },
                "Current Liabilities": {
                    "Accounts Payable": 24,
                    "Accrued Interest": 25,
                    "Short term Borrowing": 26,
                    "Current Portion of Long Term Debt": 27,
                    "Other": 28
                },
                "Non-Current Liabilities": {
                    "Long Term Debt": 31,
                    "Deferred income taxes": 32,
                    "Other": 33
                },
                "Equity": {
                    "Common Stock": 38,
                    "Retained Earnings": 39,
                    "Paid in Capital": 40,
                    "Other": 41
                }
            }
            
            # Get template year columns
            year_cols = {}
            for col in range(2, 10):  # Check columns B through I
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).isdigit():
                    year_cols[str(cell_value)] = openpyxl.utils.get_column_letter(col)
            
            print(f"[INFO] Template year columns: {year_cols}")
            
            # Apply mappings
            applied_count = 0
            for section, section_mappings in mappings.items():
                if section not in row_mappings:
                    print(f"[WARN] Template section '{section}' not found")
                    continue
                    
                section_rows = row_mappings[section]
                for template_row, year_values in section_mappings.items():
                    if template_row not in section_rows:
                        print(f"[WARN] Template row '{template_row}' not found in section '{section}'")
                        continue
                        
                    row_idx = section_rows[template_row]
                    
                    for year, value in year_values.items():
                        # Handle year mapping
                        target_year = year
                        if year_mapping and year in year_mapping:
                            target_year = year_mapping[year]
                        
                        if target_year not in year_cols:
                            print(f"[WARN] Year '{target_year}' not found in template columns")
                            continue
                        
                        col = year_cols[target_year]
                        
                        # Handle different value types
                        if isinstance(value, dict):
                            # If value is a dict, try to get a numeric value
                            numeric_value = None
                            for k, v in value.items():
                                if isinstance(v, (int, float)) and v != 0:
                                    numeric_value = v
                                    break
                            if numeric_value is not None:
                                sheet[f"{col}{row_idx}"] = numeric_value
                                applied_count += 1
                        elif isinstance(value, (int, float)):
                            sheet[f"{col}{row_idx}"] = value
                            applied_count += 1
                        else:
                            print(f"[WARN] Non-numeric value for {template_row} {year}: {value}")
            
            print(f"[SUCCESS] Applied {applied_count} values to template")
            
            # Save populated template
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"single_statement_{statement_type}_{timestamp}.xlsx"
            wb.save(output_path)
            wb.close()
            
            return output_path
        elif statement_type == 'income_statement':
            sheet = wb['IS.CF']
            # Define row mappings for income statement
            row_mappings = {
                "Revenue": {
                    "Revenue": 6
                },
                "Operating Expenses": {
                    "Operating Expenses": 10,
                    "Depreciation (-)": 11,
                    "Amortization (-)": 12,
                    "Assets gain(loss) impairments": 13
                },
                "Other Income/Expense": {
                    "Interest Expense (-)": 15,
                    "Interest Income (+)": 16,
                    "Other income(expenses)": 17
                },
                "Tax and Net Income": {
                    "Income Before Taxes": 19,
                    "Tax expense": 20,
                    "Net Income": 21
                }
            }
        elif statement_type == 'cash_flow':
            sheet = wb['IS.CF']
            # Define row mappings for cash flow statement
            row_mappings = {
                "Operating Activities": {
                    "Net Income": 23,
                    "Changes in noncash items": 24,
                    "Changes in Assets and Liabilities": 25,
                    "Net Cash from(used) Operating Activities": 26
                },
                "Investing Activities": {
                    "CapEx": 28,
                    "Proceeds from asset sales": 29,
                    "Net cash from(used) for investing": 30
                },
                "Financing Activities": {
                    "Issuance of Debt (long+short term)": 32,
                    "Retirement of Debt (long+short term)": 33,
                    "Issuance of Stock": 34,
                    "Dividends Paid": 35,
                    "Net cash from(used) for financing": 36
                },
                "Cash Reconciliation": {
                    "Net change in Cash": 38,
                    "Starting Cash": 39,
                    "Ending Cash": 40
                }
            }
        else:
            print(f"[ERROR] Unknown statement type: {statement_type}")
            return ""
        
        # Determine year columns (assuming years are in columns B, C, D, E starting from row 6)
        year_cols = {}
        for col_idx in range(2, 6):
            cell_val = sheet.cell(row=6, column=col_idx).value
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            
            if isinstance(cell_val, int) and 1990 <= cell_val <= 2050:
                year_cols[str(cell_val)] = col_letter
            elif isinstance(cell_val, str) and cell_val.startswith('='):
                try:
                    if '+1' in cell_val:
                        base_cell = cell_val.split('+')[0][1:]
                        base_col = base_cell[0]
                        base_row = int(base_cell[1:])
                        base_year = sheet[f"{base_col}{base_row}"].value
                        if isinstance(base_year, int):
                            col_offset = col_idx - openpyxl.utils.column_index_from_string(base_col)
                            calculated_year = base_year + col_offset
                            if 1990 <= calculated_year <= 2050:
                                year_cols[str(calculated_year)] = col_letter
                except Exception as e:
                    pass
        
        if not year_cols:
            print("[ERROR] Could not determine year columns from template")
            return ""
        
        print(f"[INFO] Template year columns: {year_cols}")
        
        # Apply mappings to template
        items_written = 0
        for section, section_mappings in mappings.items():
            if section not in row_mappings:
                print(f"[WARN] Section '{section}' not found in row mappings")
                continue
            
            section_row_map = row_mappings[section]
            
            for template_row, template_mappings in section_mappings.items():
                if template_row not in section_row_map:
                    print(f"[WARN] Template row '{template_row}' not found in section '{section}'")
                    continue
                
                row_idx = section_row_map[template_row]
                
                for year, value in template_mappings.items():
                    if year in year_cols:
                        col = year_cols[year]
                        sheet[f"{col}{row_idx}"] = value
                        items_written += 1
                        print(f"[WRITE] {section}::{template_row} [{year}]: {value}")
        
        # Save populated template
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"single_statement_{statement_type}_{timestamp}.xlsx"
        wb.save(output_path)
        wb.close()
        
        print(f"[SUCCESS] Applied {items_written} values to template")
        print(f"[SUCCESS] Template saved to: {output_path}")
        
        return output_path

def main():
    # Test the single statement mapper
    mapper = SingleStatementMapper()
    
    if not mapper.check_ollama_available():
        print("❌ Ollama not available. Please install and run Ollama with Mistral model.")
        return
    
    print("✅ Ollama available for single statement mapping!")
    
    # Test with sample balance sheet data
    sample_bs_data = {
        "2022": {
            "Cash and cash equivalents": 1500000,
            "Accounts receivable": 2500000,
            "Inventory": 1800000,
            "Property, plant and equipment": 5000000,
            "Accounts payable": 1200000,
            "Long-term debt": 3000000,
            "Common stock": 1000000,
            "Retained earnings": 5600000
        },
        "2023": {
            "Cash and cash equivalents": 1800000,
            "Accounts receivable": 2800000,
            "Inventory": 2000000,
            "Property, plant and equipment": 5200000,
            "Accounts payable": 1400000,
            "Long-term debt": 2800000,
            "Common stock": 1000000,
            "Retained earnings": 6800000
        }
    }
    
    print("\n🧪 Testing single statement balance sheet mapping...")
    mappings, unmapped = mapper.map_statement_comprehensive(sample_bs_data, 'balance_sheet')
    
    print(f"\n📊 Results:")
    print(f"Mappings: {json.dumps(mappings, indent=2)}")
    print(f"Unmapped items: {unmapped}")
    
    # Test applying to template
    template_path = Path(__file__).parent.parent.parent / "templates" / "financial_template.xlsx"
    if template_path.exists():
        print(f"\n📝 Applying mappings to template...")
        output_path = mapper.apply_mappings_to_excel(mappings, str(template_path), 'balance_sheet')
        print(f"Template saved to: {output_path}")

if __name__ == "__main__":
    main() 