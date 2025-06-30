#!/usr/bin/env python3
"""
Template Populator - Maps KG results into original financial template
- Reads the original financial_template.xlsx structure
- Maps our extracted values into the correct cells
- Preserves the original template format and layout
"""

import pandas as pd
import json
import shutil
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
from typing import Dict, Any

class TemplatePopulator:
    """Populates the original financial template with mapped KG data"""
    
    def __init__(self):
        self.original_template_path = Path("../../templates/financial_template.xlsx")
        self.working_template_path = Path("./populated_financial_template.xlsx")
        
        # Template field mappings - based on the image you showed
        self.template_mappings = {
            # ASSETS section
            'Cash and equivalents': ('B', 7),  # Row 7, Column B for 2023, C for 2024
            'Accounts Receivable': ('B', 8),
            'Prepaid Expenses': ('B', 9),
            'Inventory': ('B', 10),
            'Investments': ('B', 11),
            'Other': ('B', 12),  # Other current assets
            # Total Current Assets calculated
            
            'Net PPE': ('B', 15),
            'Goodwill': ('B', 16),
            'Intangibles': ('B', 17),
            # Other non-current in row 18
            # Total Non Current Assets calculated
            # Total Assets calculated
            
            # LIABILITIES section  
            'Accounts Payable': ('B', 24),
            'Accrued Interest': ('B', 25),
            'Short term Borrowing': ('B', 26),
            'Current Portion of Long Term Debt': ('B', 27),
            # Other current liabilities in row 28
            # Total Current Liabilities calculated
            
            'Long Term Debt': ('B', 31),
            'Deferred income taxes': ('B', 32),
            # Other non-current liabilities in row 33
            # Total Non Current Liabilities calculated
            # Total Liabilities calculated
            
            # EQUITY section
            'Common Stock': ('B', 39),
            'Retained Earnings': ('B', 40),
            'Paid in Capital': ('B', 41),
            # Other equity in row 42
            # Total Equity calculated
        }
    
    def setup_template(self) -> bool:
        """Copy original template for population"""
        try:
            print(f"📋 Setting up template for population...")
            print(f"   Source: {self.original_template_path}")
            print(f"   Working copy: {self.working_template_path}")
            
            if not self.original_template_path.exists():
                print(f"❌ Original template not found: {self.original_template_path}")
                return False
            
            # Use shutil to copy the template
            shutil.copy2(self.original_template_path, self.working_template_path)
            
            if self.working_template_path.exists():
                print(f"✅ Template copied successfully")
                return True
            else:
                print(f"❌ Failed to copy template")
                return False
                
        except Exception as e:
            print(f"❌ Error setting up template: {e}")
            return False
    
    def load_kg_results(self, json_file: str) -> Dict[str, Any]:
        """Load the KG mapping results from JSON"""
        try:
            with open(json_file, 'r') as f:
                data = json.load(f)
            print(f"✅ Loaded KG results: {len(data)} mappings")
            return data
        except Exception as e:
            print(f"❌ Error loading KG results: {e}")
            return {}
    
    def populate_template(self, kg_data: Dict[str, Any]) -> str:
        """Populate the original template with KG data"""
        if not self.working_template_path.exists():
            print("❌ Working template not found")
            return ""
        
        try:
            print("\n📝 Populating original template structure...")
            
            # Load the Excel workbook
            wb = load_workbook(self.working_template_path)
            
            # Get the first worksheet (assuming financial data is in first sheet)
            ws = wb.active
            print(f"   Working with sheet: {ws.title}")
            
            # Track populated fields
            populated_fields = []
            
            # Process each KG mapping
            for key, mapping in kg_data.items():
                template_field = mapping.get('template_field')
                value_2023 = mapping.get('value_2023')
                value_2024 = mapping.get('value_2024')
                
                # Check if we have a mapping for this template field
                if template_field in self.template_mappings:
                    col_letter, row_num = self.template_mappings[template_field]
                    
                    # Populate 2023 value (column B)
                    if value_2023 is not None:
                        cell_2023 = f"B{row_num}"
                        ws[cell_2023] = value_2023
                        print(f"   ✅ {template_field} 2023: {cell_2023} = {value_2023:,.0f}")
                    
                    # Populate 2024 value (column C) 
                    if value_2024 is not None:
                        cell_2024 = f"C{row_num}"
                        ws[cell_2024] = value_2024
                        print(f"   ✅ {template_field} 2024: {cell_2024} = {value_2024:,.0f}")
                    
                    populated_fields.append(template_field)
                else:
                    print(f"   ⚠️ No template mapping for: {template_field}")
            
            # Handle "Other" consolidations for each section
            self.populate_other_fields(ws, kg_data)
            
            # Calculate totals (if formulas exist in template)
            self.update_totals(ws)
            
            # Save the populated template
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"populated_original_template_{timestamp}.xlsx"
            wb.save(output_file)
            
            print(f"\n✅ Template populated successfully!")
            print(f"   Output file: {output_file}")
            print(f"   Fields populated: {len(populated_fields)}")
            print(f"   Fields: {', '.join(populated_fields)}")
            
            return output_file
            
        except Exception as e:
            print(f"❌ Error populating template: {e}")
            return ""
    
    def populate_other_fields(self, ws, kg_data: Dict[str, Any]):
        """Handle 'Other' fields that consolidate multiple items"""
        # Find "Other" mappings by section
        other_mappings = {item: data for item, data in kg_data.items() 
                         if data.get('template_field') == 'Other'}
        
        for key, mapping in other_mappings.items():
            section = mapping.get('section', '')
            value_2023 = mapping.get('value_2023')
            value_2024 = mapping.get('value_2024')
            
            # Map to appropriate "Other" row based on section
            if section == 'current_assets':
                row_num = 12  # Other current assets
            elif section == 'noncurrent_assets':
                row_num = 18  # Other non-current assets
            elif section == 'current_liabilities':
                row_num = 28  # Other current liabilities
            elif section == 'noncurrent_liabilities':
                row_num = 33  # Other non-current liabilities
            elif section == 'equity':
                row_num = 42  # Other equity
            else:
                continue
            
            # Populate the values
            if value_2023 is not None:
                ws[f"B{row_num}"] = value_2023
                print(f"   ✅ Other {section} 2023: B{row_num} = {value_2023:,.0f}")
            
            if value_2024 is not None:
                ws[f"C{row_num}"] = value_2024
                print(f"   ✅ Other {section} 2024: C{row_num} = {value_2024:,.0f}")
    
    def update_totals(self, ws):
        """Update any total formulas in the template"""
        print("\n🔢 Updating calculated totals...")
        
        # The original template likely has SUM formulas for totals
        # We just need to trigger recalculation
        # Excel will automatically recalculate when opened
        
        # Could add specific total calculations here if needed
        # For now, let Excel handle the formulas
        print("   ✅ Totals will be calculated by Excel formulas")
    
    def cleanup_template(self):
        """Clean up working template file"""
        try:
            if self.working_template_path.exists():
                self.working_template_path.unlink()
                print(f"🧹 Cleaned up working template")
        except Exception as e:
            print(f"⚠️ Warning: Could not clean up template: {e}")

def main():
    """Main function to populate the original template"""
    populator = TemplatePopulator()
    
    # Find the latest KG results file
    kg_json_file = "final_kg_us_venture_bs_20250630_135345.json"
    
    if not Path(kg_json_file).exists():
        print(f"❌ KG results file not found: {kg_json_file}")
        return
    
    try:
        # Setup template
        if not populator.setup_template():
            print("❌ Failed to setup template")
            return
        
        # Load KG results
        kg_data = populator.load_kg_results(kg_json_file)
        if not kg_data:
            print("❌ No KG data to process")
            return
        
        # Populate template
        output_file = populator.populate_template(kg_data)
        
        if output_file:
            # Copy to main output directory
            import shutil
            main_output = f"../../output_excel/{output_file}"
            shutil.copy2(output_file, main_output)
            print(f"\n📁 Also saved to main output: {main_output}")
        
    finally:
        # Clean up
        populator.cleanup_template()

if __name__ == "__main__":
    main() 