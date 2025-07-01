#!/usr/bin/env python3
"""
Quick Cash Flow Statement Test - Just populate template with hardcoded values
"""

import shutil
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def test_cfs_template():
    """Quick test to populate CFS template with hardcoded Net Income values"""
    
    # Set up template paths
    original_template_path = Path("../../templates/financial_template.xlsx")
    working_template_path = Path("./working_financial_template.xlsx")
    
    print("🧪 Quick Cash Flow Statement Template Test")
    print("=" * 50)
    
    # Copy template
    print("📋 Setting up template...")
    if not original_template_path.exists():
        print(f"❌ Template not found: {original_template_path}")
        return
    
    shutil.copy2(original_template_path, working_template_path)
    print("✅ Template copied successfully")
    
    try:
        # Load the working template
        workbook = load_workbook(working_template_path)
        
        # Check for IS.CF sheet
        if 'IS.CF' not in workbook.sheetnames:
            print(f"❌ IS.CF sheet not found. Available sheets: {workbook.sheetnames}")
            return
        
        worksheet = workbook['IS.CF']
        print(f"   Working with sheet: IS.CF")
        
        # *** HARDCODE Net Income values as requested ***
        print("\n🧪 TESTING MODE: Hardcoding Net Income values")
        
        # User requested B20 and C20 specifically
        worksheet['B20'] = 62866   # 2023 Net Income
        worksheet['C20'] = 119074  # 2024 Net Income
        print(f"   ✅ Net Income 2023: B20 = 62,866")
        print(f"   ✅ Net Income 2024: C20 = 119,074")
        
        # Also populate the cash flow starting point (row 23)
        worksheet['B23'] = 62866   # 2023 Net Income (cash flow start)
        worksheet['C23'] = 119074  # 2024 Net Income (cash flow start)
        print(f"   ✅ Cash Flow Net Income 2023: B23 = 62,866")
        print(f"   ✅ Cash Flow Net Income 2024: C23 = 119,074")
        
        # Save the populated template
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"test_cfs_template_{timestamp}.xlsx"
        output_path = Path(output_filename)
        
        workbook.save(output_path)
        workbook.close()
        
        print(f"\n✅ Template populated successfully!")
        print(f"   Output file: {output_filename}")
        
        # Copy to output directory
        output_dir = Path("../../output_excel")
        if output_dir.exists():
            output_copy = output_dir / output_filename
            shutil.copy2(output_path, output_copy)
            print(f"📁 Output copied to: {output_copy}")
        
        print(f"\n🎉 Quick test complete!")
        
    except Exception as e:
        print(f"❌ Error: {e}")
    
    finally:
        # Cleanup
        try:
            if working_template_path.exists():
                os.remove(working_template_path)
                print(f"🧹 Cleaned up working template")
        except Exception as e:
            print(f"⚠️ Could not clean up template: {e}")

if __name__ == "__main__":
    test_cfs_template() 